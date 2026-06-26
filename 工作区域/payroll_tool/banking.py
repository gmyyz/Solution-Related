import glob
import os
from datetime import date, datetime
from decimal import Decimal
from itertools import combinations

from openpyxl import load_workbook

from .rules import MAX_BANK_MATCH_COMBO_SIZE
from .workbook_utils import (
    _extract_year_month,
    _format_code,
    _normalize_text,
    _to_money,
)


BANK_FILE_PATTERN = '银行流水*.xlsx'


def _find_bank_files(bank_dir):
    return sorted(glob.glob(os.path.join(bank_dir, BANK_FILE_PATTERN)))


def _normalize_bank_date_key(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _normalize_text(value)


def _load_bank_records_with_stats(bank_dir):
    bank_records = {}
    seen_rows = set()
    source_files = _find_bank_files(bank_dir)
    stats = {
        'source_file_count': len(source_files),
        'scanned_row_count': 0,
        'deduped_row_count': 0,
        'kept_row_count': 0,
    }

    for path in source_files:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        for row_idx in range(2, ws.max_row + 1):
            company_code = _format_code(ws.cell(row=row_idx, column=2).value)
            bank_account = _format_code(ws.cell(row=row_idx, column=3).value)
            trans_date = ws.cell(row=row_idx, column=4).value
            trans_period = _extract_year_month(trans_date)
            payment_name = _normalize_text(ws.cell(row=row_idx, column=13).value)
            outgoing_amt = _to_money(ws.cell(row=row_idx, column=15).value)
            usage = _normalize_text(ws.cell(row=row_idx, column=16).value)
            bank_subject = _format_code(ws.cell(row=row_idx, column=34).value)

            if not company_code or trans_period is None:
                continue

            stats['scanned_row_count'] += 1
            row_identity = (
                company_code,
                _normalize_bank_date_key(trans_date),
                outgoing_amt,
                usage,
                payment_name,
                bank_subject,
            )
            if row_identity in seen_rows:
                stats['deduped_row_count'] += 1
                continue
            seen_rows.add(row_identity)

            key = (company_code, trans_period[0], trans_period[1])
            bank_records.setdefault(key, []).append(
                {
                    'file': os.path.basename(path),
                    'row_idx': row_idx,
                    'company_code': company_code,
                    'bank_account': bank_account,
                    'trans_date': trans_date,
                    'payment_name': payment_name,
                    'outgoing_amt': outgoing_amt,
                    'usage': usage,
                    'bank_subject': bank_subject,
                }
            )
            stats['kept_row_count'] += 1

    return {'records': bank_records, 'stats': stats}


def _load_bank_records(bank_dir):
    return _load_bank_records_with_stats(bank_dir)['records']


def _get_treasury_records(records):
    return sorted(
        [
            record
            for record in records
            if '国家金库北京市分库' in record['payment_name'] and record['outgoing_amt'] > 0
        ],
        key=lambda item: (_normalize_bank_date_key(item['trans_date']), item['row_idx'], item['outgoing_amt']),
    )


def _summarize_bank_records(bank_records, company_code, year, month):
    records = bank_records.get((company_code, year, month), [])
    salary_records = [record for record in records if '代发工资' in record['usage']]
    salary_amount = sum((record['outgoing_amt'] for record in salary_records), Decimal('0.00'))
    fund_amount = sum(
        (
            record['outgoing_amt']
            for record in records
            if '北京住房公积金管理中心' in record['payment_name']
        ),
        Decimal('0.00'),
    )
    treasury_records = _get_treasury_records(records)
    treasury_amount = sum((record['outgoing_amt'] for record in treasury_records), Decimal('0.00'))

    return {
        'salary': salary_amount,
        'salary_records': salary_records,
        'fund': fund_amount,
        'tax': Decimal('0.00'),
        'social': Decimal('0.00'),
        'treasury': treasury_amount,
        'treasury_records': treasury_records,
        'files': sorted({record['file'] for record in records}),
        'accounts': sorted({record['bank_account'] for record in records if record['bank_account']}),
        'record_count': len(records),
    }


def _match_bank_record_combination(records, target_amount, label):
    target = _to_money(target_amount)
    if not records:
        return {
            'matched': False,
            'matched_amount': Decimal('0.00'),
            'all_amount': Decimal('0.00'),
            'matched_records': [],
            'note': f'未找到{label}',
        }

    all_amount = sum((record['outgoing_amt'] for record in records), Decimal('0.00'))
    max_size = min(MAX_BANK_MATCH_COMBO_SIZE, len(records))
    for size in range(1, max_size + 1):
        for combo in combinations(records, size):
            combo_amount = sum((record['outgoing_amt'] for record in combo), Decimal('0.00'))
            if combo_amount == target:
                combo_desc = ' + '.join(str(record['outgoing_amt']) for record in combo)
                return {
                    'matched': True,
                    'matched_amount': combo_amount,
                    'all_amount': all_amount,
                    'matched_records': list(combo),
                    'note': f'匹配到 {size} 条{label}组合：{combo_desc}',
                }

    return {
        'matched': False,
        'matched_amount': all_amount,
        'all_amount': all_amount,
        'matched_records': [],
        'note': f'未找到匹配组合；全部{label}合计={all_amount}',
    }


def _match_salary_combination(salary_records, target_amount):
    return _match_bank_record_combination(salary_records, target_amount, '代发工资')


def _match_treasury_combination(records, target_amount, label):
    return _match_bank_record_combination(_get_treasury_records(records), target_amount, label)
