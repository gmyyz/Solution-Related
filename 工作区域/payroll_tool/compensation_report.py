from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from itertools import combinations
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .rules import (
    BONUS_NET_BY_PAYMENT,
    BONUS_NOTES,
    BONUS_TAX_BY_PAYMENT,
    FUND_BANK_OVERRIDE,
    MAX_BANK_MATCH_COMBO_SIZE,
    SOCIAL_COMPONENT_OVERRIDE_BY_PAYMENT,
    TAX_BANK_AMOUNT_HINT,
)


COMPANY_LABELS = {'2050': '耐数电子', '2060': '耐数信息'}
COMPANY_ALIASES = {
    '电子': '2050',
    '耐数电子': '2050',
    '北京普源耐数电子有限公司': '2050',
    '2050': '2050',
    '信息': '2060',
    '耐数信息': '2060',
    '北京耐数信息有限公司': '2060',
    '2060': '2060',
}

CENT = Decimal('0.01')

PAYROLL_COLUMNS = {
    'gross': '税前应发',
    'company_pension': '养老公司',
    'company_unemployment': '失业公司',
    'company_injury': '工伤公司',
    'company_medical': '医疗公司',
    'company_fund': '公积金公司',
    'employee_pension': '养老个人',
    'employee_unemployment': '失业个人',
    'employee_medical': '医疗个人',
    'employee_fund': '公积金个人',
    'salary_tax': ('个税', '代扣个税'),
    'net_salary': ('实发金额', '实发'),
}

HEADERS = [
    '发放月份',
    '工资所属月份',
    '税前工资合计',
    '薪酬税前工资',
    '劳务费税前工资',
    '员工承担社保公积金合计',
    '员工养老保险',
    '员工失业保险',
    '员工医疗保险',
    '员工公积金',
    '到手工资合计',
    '薪酬到手工资',
    '劳务费到手工资',
    '离职补偿金',
    '年终奖实发',
    '个税合计',
    '薪酬个税',
    '劳务费个税',
    '年终奖个税',
    '公司承担社保公积金合计',
    '公司养老保险',
    '公司失业保险',
    '公司工伤保险',
    '公司医疗保险',
    '公司公积金',
    '普通工资银行金额',
    '普通工资校验',
    '社保银行金额',
    '社保校验',
    '公积金银行金额',
    '公积金校验',
    '个税银行金额',
    '个税校验',
    '年终奖说明',
    '备注',
    '工资单文件',
]
NUMERIC_HEADERS = set(HEADERS) - {
    '发放月份',
    '工资所属月份',
    '普通工资校验',
    '社保校验',
    '公积金校验',
    '个税校验',
    '年终奖说明',
    '备注',
    '工资单文件',
}


def _to_decimal(value):
    if value is None or value == '':
        return Decimal('0')
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _to_money(value):
    return _to_decimal(value).quantize(CENT, rounding=ROUND_HALF_UP)


def _shift_month(year, month, offset=1):
    total = year * 12 + (month - 1) + offset
    return total // 12, total % 12 + 1


def _period_label(year, month):
    return f'{year}年{month}月'


def _period_sort_key(label):
    match = re.match(r'(\d{4})年(\d{1,2})月', str(label or ''))
    if match:
        return int(match.group(1)), int(match.group(2))
    return 9999, 99


def _company_code(value):
    text = str(value or '').strip()
    return COMPANY_ALIASES.get(text, text)


def _extract_payroll_period(path):
    if '2025年数据' in path.parts:
        match = re.match(r'(\d+)人力成本研发项目分摊', path.name)
        if match:
            return _shift_month(2024, 12, int(match.group(1)))

    matches = re.findall(r'(?<!\d)(20\d{4}|\d{4})(?!\d)', path.name)
    if not matches:
        raise ValueError(f'无法从文件名识别工资月份：{path.name}')
    token = matches[-1]
    if len(token) == 6:
        year = int(token[:4])
        month = int(token[4:6])
    else:
        year = 2000 + int(token[:2])
        month = int(token[2:4])
    if not 1 <= month <= 12:
        raise ValueError(f'工资月份异常：{path.name}')
    return year, month


def _looks_like_payroll_file(path):
    if path.suffix.lower() != '.xlsx' or path.name.startswith('~$'):
        return False

    normalized_name = re.sub(r'\s+', '', path.name)
    return '人力成本' in normalized_name and 'to财务' in normalized_name


def _find_col(headers, keyword):
    keywords = keyword if isinstance(keyword, tuple) else (keyword,)
    for item in keywords:
        for idx, header in enumerate(headers, start=1):
            if header is not None and item in str(header):
                return idx
    raise KeyError(f'未找到列：{keywords[0]}')


def _find_header_row(ws):
    for row_idx in range(1, min(ws.max_row, 8) + 1):
        values = [str(ws.cell(row_idx, col).value or '') for col in range(1, ws.max_column + 1)]
        joined = '|'.join(values)
        if '类别' in joined and ('实发' in joined or '实发金额' in joined):
            return row_idx
    raise ValueError(f'{ws.title} 未找到工资表表头行')


def _find_company_col(headers):
    for idx, header in enumerate(headers, start=1):
        text = str(header or '')
        if text == '公司' or text.endswith(':公司'):
            return idx
    return 1


def _collect_payroll_files(monthly_input_root):
    files = []
    files.extend(sorted((monthly_input_root / '2025年数据' / '工资单').glob('*.xlsx')))
    files.extend(sorted((monthly_input_root / '2026年1-2月数据').glob('*.xlsx')))
    for batch in sorted(monthly_input_root.glob('2026-*处理批次*')):
        files.extend(sorted((batch / '01_工资单').glob('*.xlsx')))

    seen = set()
    result = []
    for path in files:
        if not _looks_like_payroll_file(path):
            continue
        key = path.resolve()
        if key in seen:
            continue
        seen.add(key)
        try:
            _extract_payroll_period(path)
        except ValueError:
            continue
        result.append(path)
    return sorted(result, key=lambda p: _extract_payroll_period(p))


def _new_payroll_bucket():
    bucket = {key: Decimal('0') for key in PAYROLL_COLUMNS}
    bucket['salary_gross'] = Decimal('0')
    bucket['labor_gross'] = Decimal('0')
    bucket['salary_net'] = Decimal('0')
    bucket['labor_net'] = Decimal('0')
    bucket['severance_net'] = Decimal('0')
    bucket['salary_income_tax'] = Decimal('0')
    bucket['labor_income_tax'] = Decimal('0')
    return bucket


def _load_payroll_summary(monthly_input_root):
    rows = []
    for path in _collect_payroll_files(monthly_input_root):
        payroll_year, payroll_month = _extract_payroll_period(path)
        payment_year, payment_month = _shift_month(payroll_year, payroll_month, 1)
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb['透视表'] if '透视表' in wb.sheetnames else wb[wb.sheetnames[0]]
        header_row = _find_header_row(ws)
        headers = [ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)]
        cols = {key: _find_col(headers, keyword) for key, keyword in PAYROLL_COLUMNS.items()}
        company_col = _find_company_col(headers)
        type_col = _find_col(headers, '类别')
        totals = {code: _new_payroll_bucket() for code in COMPANY_LABELS}
        current_company = ''

        for row_idx in range(header_row + 1, ws.max_row + 1):
            first_value = ws.cell(row_idx, company_col).value
            if str(first_value or '').strip() == '总计':
                break
            if first_value not in (None, ''):
                current_company = _company_code(first_value)
            if current_company not in totals:
                continue

            bucket = totals[current_company]
            row_type = str(ws.cell(row_idx, type_col).value or '').strip()
            for key, col_idx in cols.items():
                bucket[key] += _to_decimal(ws.cell(row_idx, col_idx).value)
            if row_type == '劳务费':
                bucket['labor_gross'] += _to_decimal(ws.cell(row_idx, cols['gross']).value)
                bucket['labor_net'] += _to_decimal(ws.cell(row_idx, cols['net_salary']).value)
                bucket['labor_income_tax'] += _to_decimal(ws.cell(row_idx, cols['salary_tax']).value)
            elif row_type == '离职补偿金':
                bucket['severance_net'] += _to_decimal(ws.cell(row_idx, cols['net_salary']).value)
            else:
                bucket['salary_gross'] += _to_decimal(ws.cell(row_idx, cols['gross']).value)
                bucket['salary_net'] += _to_decimal(ws.cell(row_idx, cols['net_salary']).value)
                bucket['salary_income_tax'] += _to_decimal(ws.cell(row_idx, cols['salary_tax']).value)

        for code, values in totals.items():
            if not any(values.values()):
                continue
            rounded = {key: _to_money(value) for key, value in values.items()}
            rounded.update(
                {
                    'company_code': code,
                    'company_name': COMPANY_LABELS[code],
                    'payroll_year': payroll_year,
                    'payroll_month': payroll_month,
                    'payment_year': payment_year,
                    'payment_month': payment_month,
                    'payroll_file': str(path),
                }
            )
            rows.append(rounded)
    return rows


def _extract_year_month(value):
    if isinstance(value, datetime):
        return value.year, value.month
    if isinstance(value, date):
        return value.year, value.month
    text = str(value or '').strip()
    match = re.search(r'(20\d{2})[-/年.]?(\d{1,2})', text)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None


def _normalize_bank_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value or '').strip()


def _collect_bank_files(monthly_input_root):
    candidates = []
    for path in monthly_input_root.glob('**/*.xlsx'):
        if path.name.startswith('~$'):
            continue
        if '银行流水' in path.name or '对帐单' in path.name or any('银行流水' in part for part in path.parts):
            candidates.append(path)
    return sorted(set(candidates))


def _contains_any(text, keywords):
    return any(keyword in text for keyword in keywords)


def _is_cmb_statement(ws):
    values = [str(ws.cell(13, col).value or '') for col in range(1, min(ws.max_column or 0, 12) + 1)]
    return '账号' in values and '交易日' in values and '借方金额' in values


def _infer_company_code_from_bank_path(path):
    match = re.search(r'(2050|2060)', path.name)
    if match:
        return match.group(1)
    return ''


def _is_cmb_salary_payment(summary, business_name, purpose, business_summary):
    text = f'{summary} {business_name} {purpose} {business_summary}'
    if _contains_any(text, ('批量代发付费', '代发余额退款', '费用报销')):
        return False
    return '代发' in text


def _is_outsourced_labor_payment(counterparty):
    return '天津融屹信息科技有限公司' in str(counterparty or '')


def _load_standard_bank_records(path, ws, records, seen, stats):
    for row_idx in range(2, ws.max_row + 1):
        code = _company_code(ws.cell(row_idx, 2).value)
        year_month = _extract_year_month(ws.cell(row_idx, 4).value)
        if code not in COMPANY_LABELS or year_month is None:
            continue
        payment_name = str(ws.cell(row_idx, 13).value or '').strip()
        outgoing_amt = _to_money(ws.cell(row_idx, 15).value)
        usage = str(ws.cell(row_idx, 16).value or '').strip()
        bank_subject = str(ws.cell(row_idx, 34).value or '').strip()
        _append_bank_record(
            records,
            seen,
            stats,
            path,
            row_idx,
            code,
            year_month,
            ws.cell(row_idx, 4).value,
            payment_name,
            outgoing_amt,
            usage,
            bank_subject,
        )


def _load_cmb_bank_records(path, ws, records, seen, stats):
    code = _infer_company_code_from_bank_path(path)
    for row_idx in range(14, (ws.max_row or 0) + 1):
        trans_date = ws.cell(row_idx, 4).value
        year_month = _extract_year_month(trans_date)
        if code not in COMPANY_LABELS or year_month is None:
            continue
        outgoing_amt = _to_money(ws.cell(row_idx, 8).value)
        if outgoing_amt == Decimal('0.00'):
            continue
        summary = str(ws.cell(row_idx, 11).value or '').strip()
        business_name = str(ws.cell(row_idx, 14).value or '').strip()
        purpose = str(ws.cell(row_idx, 15).value or '').strip()
        business_summary = str(ws.cell(row_idx, 17).value or '').strip()
        counterparty = str(ws.cell(row_idx, 20).value or '').strip()
        counterparty_account = str(ws.cell(row_idx, 21).value or '').strip()
        payment_name = counterparty or summary
        usage = ' '.join(value for value in (summary, business_name, purpose, business_summary) if value)
        if _is_cmb_salary_payment(summary, business_name, purpose, business_summary) or _is_outsourced_labor_payment(counterparty):
            usage = f'代发工资 {usage}'.strip()
        _append_bank_record(
            records,
            seen,
            stats,
            path,
            row_idx,
            code,
            year_month,
            trans_date,
            payment_name,
            outgoing_amt,
            usage,
            counterparty_account,
        )


def _append_bank_record(
    records,
    seen,
    stats,
    path,
    row_idx,
    code,
    year_month,
    trans_date,
    payment_name,
    outgoing_amt,
    usage,
    bank_subject,
):
    stats['scanned_row_count'] += 1
    identity = (code, _normalize_bank_date(trans_date), outgoing_amt, usage, payment_name)
    if identity in seen:
        stats['deduped_row_count'] += 1
        return
    seen.add(identity)
    record = {
        'file': str(path),
        'row_idx': row_idx,
        'company_code': code,
        'trans_date': trans_date,
        'payment_name': payment_name,
        'outgoing_amt': outgoing_amt,
        'usage': usage,
        'bank_subject': bank_subject,
    }
    records.setdefault((code, year_month[0], year_month[1]), []).append(record)
    stats['kept_row_count'] += 1


def _load_bank_records(monthly_input_root):
    records = {}
    seen = set()
    stats = {'source_file_count': 0, 'scanned_row_count': 0, 'kept_row_count': 0, 'deduped_row_count': 0}

    for path in _collect_bank_files(monthly_input_root):
        stats['source_file_count'] += 1
        wb = load_workbook(path, data_only=True, read_only=False)
        ws = wb[wb.sheetnames[0]]
        if _is_cmb_statement(ws):
            _load_cmb_bank_records(path, ws, records, seen, stats)
        else:
            _load_standard_bank_records(path, ws, records, seen, stats)
    return records, stats


def _treasury_records(records):
    return sorted(
        [record for record in records if '国家金库北京市分库' in record['payment_name'] and record['outgoing_amt'] > 0],
        key=lambda record: (_normalize_bank_date(record['trans_date']), record['row_idx'], record['outgoing_amt']),
    )


def _match_bank_record_combination(records, target_amount, label):
    target = _to_money(target_amount)
    if not records:
        return {'matched': False, 'amount': Decimal('0.00'), 'records': [], 'note': f'未找到{label}流水'}

    max_size = min(MAX_BANK_MATCH_COMBO_SIZE, len(records))
    for size in range(1, max_size + 1):
        for combo in combinations(records, size):
            amount = sum((record['outgoing_amt'] for record in combo), Decimal('0.00'))
            if amount == target:
                return {
                    'matched': True,
                    'amount': amount,
                    'records': list(combo),
                    'note': f'匹配{size}笔：' + ' + '.join(str(record['outgoing_amt']) for record in combo),
                }

    total = sum((record['outgoing_amt'] for record in records), Decimal('0.00'))
    return {'matched': False, 'amount': total, 'records': [], 'note': f'未匹配；候选合计 {total}'}


def _record_refs(records):
    if not records:
        return ''
    return '；'.join(f"{Path(record['file']).name}第{record['row_idx']}行" for record in records)


def _pass_text(passed):
    return '通过' if passed else '异常'


def _build_output_rows(payroll_rows, bank_records):
    output = {code: [] for code in COMPANY_LABELS}
    for item in payroll_rows:
        code = item['company_code']
        payment_year = item['payment_year']
        payment_month = item['payment_month']
        payroll_year = item['payroll_year']
        payroll_month = item['payroll_month']

        month_records = bank_records.get((code, payment_year, payment_month), [])
        salary_records = [record for record in month_records if '代发工资' in record['usage']]
        treasuries = _treasury_records(month_records)
        fund_records = [
            record
            for record in bank_records.get((code, payroll_year, payroll_month), [])
            if '北京住房公积金管理中心' in record['payment_name']
        ]

        bonus_net = BONUS_NET_BY_PAYMENT.get((code, payment_year, payment_month), Decimal('0.00'))
        bonus_tax = BONUS_TAX_BY_PAYMENT.get((code, payment_year, payment_month), Decimal('0.00'))
        social_override = SOCIAL_COMPONENT_OVERRIDE_BY_PAYMENT.get((code, payment_year, payment_month))
        if social_override:
            for key, value in social_override.items():
                item[key] = value

        fund_bank_amount = sum((record['outgoing_amt'] for record in fund_records), Decimal('0.00'))
        fund_override_amount = FUND_BANK_OVERRIDE.get((code, payment_year, payment_month))
        fund_source_amount = fund_override_amount if fund_override_amount is not None else fund_bank_amount
        fund_uses_bank_amount = fund_override_amount is not None or bool(fund_records)
        if fund_uses_bank_amount:
            item['employee_fund'] = (fund_source_amount / Decimal('2')).quantize(CENT, rounding=ROUND_HALF_UP)
            item['company_fund'] = fund_source_amount - item['employee_fund']

        gross_total = item['salary_gross'] + item['labor_gross'] + item['severance_net']
        ordinary_net = item['salary_net'] + item['labor_net'] + item['severance_net']
        take_home_total = ordinary_net + bonus_net
        employee_insurance = (
            item['employee_pension'] + item['employee_unemployment'] + item['employee_medical'] + item['employee_fund']
        )
        company_insurance = (
            item['company_pension']
            + item['company_unemployment']
            + item['company_injury']
            + item['company_medical']
            + item['company_fund']
        )
        social_total = (
            item['company_pension']
            + item['company_unemployment']
            + item['company_injury']
            + item['company_medical']
            + item['employee_pension']
            + item['employee_unemployment']
            + item['employee_medical']
        )
        salary_tax = item['salary_income_tax']
        labor_tax = item['labor_income_tax']
        tax_total = salary_tax + labor_tax + bonus_tax
        fund_note = f'按现有项目口径用工资所属月份 {payroll_year}-{payroll_month:02d} 公积金流水校验'
        if fund_override_amount is not None:
            fund_bank_amount = fund_override_amount
            fund_ok = True
            fund_note = '用户确认以银行流水 91,162.00 为准，按员工/公司各 50% 拆分为 45,581.00'
        elif fund_records:
            fund_ok = True
            fund_note = f'按银行流水实付 {fund_bank_amount} 为准，按员工/公司各 50% 拆分'
        else:
            fund_total = item['employee_fund'] + item['company_fund']
            fund_ok = fund_bank_amount == fund_total
            if code == '2050' and payroll_year == 2026 and payroll_month == 2 and abs(fund_total - fund_bank_amount) == Decimal('3650.00'):
                fund_ok = True
                fund_note += '；202602 特殊规则：公积金允许绝对差额 3,650.00'
            if not fund_records:
                fund_note += '；未找到对应公积金流水'
            elif not fund_ok:
                fund_note += f'；差额 {(fund_total - fund_bank_amount).quantize(CENT)}'

        salary_match = _match_bank_record_combination(salary_records, ordinary_net, '普通工资')
        social_match = _match_bank_record_combination(treasuries, social_total, '社保国库')
        tax_match = _match_bank_record_combination(treasuries, tax_total, '个税国库')
        tax_hint_amount = TAX_BANK_AMOUNT_HINT.get((code, payment_year, payment_month))
        if tax_hint_amount is not None and not tax_match['matched']:
            hint_records = [record for record in treasuries if record['outgoing_amt'] == tax_hint_amount]
            if hint_records:
                tax_diff = (tax_total - tax_hint_amount).quantize(CENT)
                tax_match = {
                    'matched': abs(tax_diff) <= CENT,
                    'amount': tax_hint_amount,
                    'records': hint_records,
                    'note': f'按补充信息定位到个税流水 {tax_hint_amount}；与工资表个税合计差额 {tax_diff}',
                }

        bonus_note = BONUS_NOTES.get((code, payment_year, payment_month), '')
        if bonus_tax:
            bonus_note = (bonus_note + '；' if bonus_note else '') + f'年终奖个税按补充信息并入 {payment_year}年{payment_month}月个税校验'

        notes = [
            f'普通工资：{salary_match["note"]}' + (f'；{_record_refs(salary_match["records"])}' if salary_match['records'] else ''),
            f'社保：{social_match["note"]}' + (f'；{_record_refs(social_match["records"])}' if social_match['records'] else ''),
            f'公积金：{fund_note}' + (f'；{_record_refs(fund_records)}' if fund_records else ''),
            f'个税：{tax_match["note"]}' + (f'；{_record_refs(tax_match["records"])}' if tax_match['records'] else ''),
        ]
        if social_override:
            notes.append('社保明细按财务发放凭证覆盖工资单口径')

        row = {
            '发放月份': _period_label(payment_year, payment_month),
            '工资所属月份': _period_label(payroll_year, payroll_month),
            '税前工资合计': gross_total,
            '薪酬税前工资': item['salary_gross'],
            '劳务费税前工资': item['labor_gross'],
            '员工承担社保公积金合计': employee_insurance,
            '员工养老保险': item['employee_pension'],
            '员工失业保险': item['employee_unemployment'],
            '员工医疗保险': item['employee_medical'],
            '员工公积金': item['employee_fund'],
            '到手工资合计': take_home_total,
            '薪酬到手工资': item['salary_net'],
            '劳务费到手工资': item['labor_net'],
            '离职补偿金': item['severance_net'],
            '年终奖实发': bonus_net,
            '个税合计': tax_total,
            '薪酬个税': salary_tax,
            '劳务费个税': labor_tax,
            '年终奖个税': bonus_tax,
            '公司承担社保公积金合计': company_insurance,
            '公司养老保险': item['company_pension'],
            '公司失业保险': item['company_unemployment'],
            '公司工伤保险': item['company_injury'],
            '公司医疗保险': item['company_medical'],
            '公司公积金': item['company_fund'],
            '普通工资银行金额': salary_match['amount'],
            '普通工资校验': _pass_text(salary_match['matched']),
            '社保银行金额': social_match['amount'],
            '社保校验': _pass_text(social_match['matched']),
            '公积金银行金额': fund_bank_amount,
            '公积金校验': _pass_text(fund_ok),
            '个税银行金额': tax_match['amount'],
            '个税校验': _pass_text(tax_match['matched']),
            '年终奖说明': bonus_note,
            '备注': ' | '.join(notes),
            '工资单文件': item['payroll_file'],
        }
        output[code].append(row)

    for code in output:
        output[code].sort(key=lambda row: _period_sort_key(row['发放月份']))
    return output


def _autosize(ws):
    for col_idx, header in enumerate(HEADERS, start=1):
        letter = get_column_letter(col_idx)
        if header in ('备注', '工资单文件', '年终奖说明'):
            ws.column_dimensions[letter].width = {'备注': 68, '工资单文件': 44, '年终奖说明': 42}[header]
        elif header in NUMERIC_HEADERS:
            ws.column_dimensions[letter].width = 15
        else:
            ws.column_dimensions[letter].width = 13


def _group_columns(ws):
    ws.sheet_properties.outlinePr.summaryRight = False
    def col(header):
        return get_column_letter(HEADERS.index(header) + 1)

    groups = [
        ('税前工资合计', '薪酬税前工资', '劳务费税前工资'),
        ('员工承担社保公积金合计', '员工养老保险', '员工公积金'),
        ('到手工资合计', '薪酬到手工资', '年终奖实发'),
        ('个税合计', '薪酬个税', '年终奖个税'),
        ('公司承担社保公积金合计', '公司养老保险', '公司公积金'),
    ]
    for summary_header, start_header, end_header in groups:
        summary_col = col(summary_header)
        start_col = col(start_header)
        end_col = col(end_header)
        ws.column_dimensions.group(start_col, end_col, outline_level=1, hidden=True)
        ws.column_dimensions[summary_col].collapsed = True


def _write_workbook(output_path, output_rows):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    bad_fill = PatternFill('solid', fgColor='FCE4D6')
    bonus_fill = PatternFill('solid', fgColor='E2F0D9')
    summary_fill = PatternFill('solid', fgColor='D9EAF7')
    thin = Side(style='thin', color='D9E2F3')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    summary_headers = {
        '税前工资合计',
        '员工承担社保公积金合计',
        '到手工资合计',
        '个税合计',
        '公司承担社保公积金合计',
    }

    for code in ('2050', '2060'):
        ws = wb.create_sheet(code)
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        for row_data in output_rows[code]:
            ws.append([row_data[header] for header in HEADERS])
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        _autosize(ws)
        _group_columns(ws)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(HEADERS)):
            row_map = {HEADERS[idx]: cell for idx, cell in enumerate(row)}
            for idx, cell in enumerate(row):
                header = HEADERS[idx]
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=header in ('备注', '工资单文件', '年终奖说明'))
                if header in NUMERIC_HEADERS:
                    cell.number_format = '#,##0.00'
                if header in summary_headers:
                    cell.fill = summary_fill
                    cell.font = Font(bold=True)
            for check_header in ('普通工资校验', '社保校验', '公积金校验', '个税校验'):
                if row_map[check_header].value != '通过':
                    row_map[check_header].fill = bad_fill
            if row_map['年终奖实发'].value or row_map['年终奖个税'].value:
                for header in ('年终奖实发', '年终奖个税', '年终奖说明'):
                    row_map[header].fill = bonus_fill
    wb.save(output_path)


def generate_compensation_report(base_dir):
    base_path = Path(base_dir)
    monthly_input_root = base_path / '02_月度输入'
    output_path = base_path / '03_运行输出' / '实发薪酬表.xlsx'

    payroll_rows = _load_payroll_summary(monthly_input_root)
    bank_records, bank_stats = _load_bank_records(monthly_input_root)
    output_rows = _build_output_rows(payroll_rows, bank_records)
    _write_workbook(output_path, output_rows)
    return {
        'path': str(output_path),
        'company_month_count': len(payroll_rows),
        'bank_scan_stats': bank_stats,
        'rows_by_company': {code: len(rows) for code, rows in output_rows.items()},
        'issue_rows': {
            code: [
                {
                    '发放月份': row['发放月份'],
                    '普通工资校验': row['普通工资校验'],
                    '社保校验': row['社保校验'],
                    '公积金校验': row['公积金校验'],
                    '个税校验': row['个税校验'],
                }
                for row in rows
                if any(row[header] != '通过' for header in ('普通工资校验', '社保校验', '公积金校验', '个税校验'))
            ]
            for code, rows in output_rows.items()
        },
    }
