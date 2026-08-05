import glob
import io
import os
import re
import shutil
import tkinter as tk
from copy import copy
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP, ROUND_DOWN
from tkinter import ttk

from openpyxl import Workbook, load_workbook

from .banking import (
    _get_treasury_records,
    _load_bank_records,
    _load_bank_records_with_stats,
    _match_bank_record_combination,
    _match_salary_combination,
    _match_treasury_combination,
    _normalize_bank_date_key,
    _summarize_bank_records,
)
from .compensation_report import generate_compensation_report
from .options import (
    ACTUAL_VOUCHERS,
    VOUCHER_DISPLAY_ORDER,
    build_batch_layout,
    format_period_label,
    normalize_run_options,
    requires_bank_data,
    requires_bonus_data,
    requires_co_data,
    requires_shared_expense_data,
)
from .rules import (
    BANK_REASON_CODE,
    BANK_VOUCHER_TYPE,
    BONUS_DEPARTMENT_ACCRUAL_OVERRIDES,
    BONUS_TAX_BY_PAYMENT_PERIOD,
    DEPARTMENT_DAILY_ORDER_OVERRIDES,
)
from .workbook_utils import (
    _copy_cell_style,
    _find_payroll_header_row,
    _find_total_row,
    _format_code,
    _is_blank,
    _is_blank_project,
    _is_total_row,
    _mark_range_red,
    _mark_row_red,
    _normalize_text,
    _recalculate_total_row,
    _to_decimal,
    _to_money,
    _to_text_money,
)


# ============================================================
# 配色常量（RIGOL 品牌风格：黄黑白）
# ============================================================

COLOR_PRIMARY = '#FFD700'
COLOR_SUCCESS = '#FFD700'
COLOR_DANGER = '#FF4444'
COLOR_WARNING = '#FFA500'
COLOR_BG = '#1A1A1A'
COLOR_CARD = '#2A2A2A'
COLOR_BORDER = '#3D3D3D'
COLOR_TEXT_MAIN = '#FFFFFF'
COLOR_TEXT_SUB = '#AAAAAA'
COLOR_HEADER_BG = '#111111'

RAW_FILE_PATTERNS = (
    '人力成本研发项目分摊* - to财务-原始.xlsx',
    '人力成本研发项目分摊* - to财务.xlsx',
    '耐数人力成本分摊*-to财务.xlsx',
    '耐数人力成本分摊* - to财务.xlsx',
)
RAW_FILE_PATTERN = RAW_FILE_PATTERNS[0]
COMPANY_NAME_TO_CODE = {'耐数电子': '2050', '耐数信息': '2060'}
COMPANY_CODE_TO_NAME = {value: key for key, value in COMPANY_NAME_TO_CODE.items()}
SOCIAL_DEBIT_ACCOUNT_COLUMNS = [
    (6, '2211020002'),
    (7, '2211030002'),
    (8, '2211040001'),
    (9, '2211060002'),
    (11, '2211020001'),
    (12, '2211030001'),
    (13, '2211060001'),
]
SOCIAL_TEXT_BY_ACCOUNT = {
    '2211020002': '支付{period}工作期间公司承担养老保险',
    '2211030002': '支付{period}工作期间公司承担失业保险',
    '2211040001': '支付{period}工作期间公司承担工伤保险',
    '2211060002': '支付{period}工作期间公司承担医疗保险',
    '2211020001': '支付{period}工作期间个人承担养老保险',
    '2211030001': '支付{period}工作期间个人承担失业保险',
    '2211060001': '支付{period}工作期间个人承担医疗保险',
}
ACCRUAL_DEBIT_SPECS = [
    (5, '6601010001', '薪酬'),
    (6, '6601030008', '养老保险'),
    (8, '6601030003', '工伤保险'),
    (7, '6601030002', '失业保险'),
    (9, '6601030005', '医疗保险'),
    (10, '6601030006', '住房公积金'),
]
ACCRUAL_CREDIT_SPECS = [
    (10, '2211070002', '公司承担住房公积金'),
    (9, '2211060002', '公司承担医疗保险'),
    (7, '2211030002', '公司承担失业保险'),
    (8, '2211040001', '公司承担工伤保险'),
    (6, '2211020002', '公司承担养老保险'),
    (14, '2211070001', '个人承担住房公积金'),
    (13, '2211060001', '个人承担医疗保险'),
    (12, '2211030001', '个人承担失业保险'),
    (11, '2211020001', '个人承担养老保险'),
]
ACCOUNTS_TO_CROSS_CHECK = [
    '2211060002',
    '2211030002',
    '2211040001',
    '2211020002',
    '2221070000',
    '2211010001',
    '2211060001',
    '2211030001',
    '2211020001',
]
CO_SPECIAL_COST_ELEMENTS = {
    '6601010001',
    '6601010003',
    '6601030002',
    '6601030003',
    '6601030005',
    '6601030006',
    '6601030008',
}
CO_DEBIT_GL_BY_SOURCE = {
    'special': '5001010011',
    'other': '5001010012',
}
CO_CREDIT_COST_CENTER = '20502020'
CO_CREDIT_ORDER = '9201856'
EXAMPLE_IMAGE_DIRNAME = '示例截图'


def _apply_style(root):
    style = ttk.Style(root)
    style.theme_use('clam')
    style.configure('TFrame', background=COLOR_BG)
    style.configure(
        'Horizontal.TProgressbar',
        background=COLOR_PRIMARY,
        troughcolor='#333333',
        borderwidth=0,
        thickness=8,
    )


def _btn(parent, text, bg, command, padx=20, pady=8):
    if bg in (COLOR_PRIMARY, COLOR_SUCCESS):
        fg_color = '#000000'
        active_bg = '#E6C200'
    elif bg == COLOR_DANGER:
        fg_color = '#FFFFFF'
        active_bg = '#CC2222'
    else:
        fg_color = '#FFFFFF'
        active_bg = '#555555'
    return tk.Button(
        parent,
        text=text,
        bg=bg,
        fg=fg_color,
        font=('微软雅黑', 11, 'bold'),
        relief='flat',
        padx=padx,
        pady=pady,
        cursor='hand2',
        activebackground=active_bg,
        activeforeground=fg_color,
        command=command,
    )


def _circle_label(parent, text, size=26, bg_color=None, fg_color='#000000', parent_bg=None):
    bg_color = bg_color or COLOR_PRIMARY
    parent_bg = parent_bg or COLOR_BG
    canvas = tk.Canvas(parent, width=size, height=size, bg=parent_bg, highlightthickness=0)
    canvas.create_oval(1, 1, size - 1, size - 1, fill=bg_color, outline='')
    canvas.create_text(
        size // 2,
        size // 2,
        text=text,
        font=('微软雅黑', max(size // 3, 8), 'bold'),
        fill=fg_color,
        anchor='center',
    )
    return canvas


def _get_raw_dir(base_dir, processing_year=None, processing_month=None):
    if processing_year is not None and processing_month is not None:
        return build_batch_layout(base_dir, processing_year, processing_month).raw_dir
    return os.path.join(base_dir, '原始数据', '工资单')


def _get_mapping_path(base_dir):
    primary = os.path.join(base_dir, '01_基础资料', '01_映射与规则', 'Mapping表.xlsx')
    if os.path.exists(primary):
        return primary
    return os.path.join(base_dir, 'Mapping表.xlsx')


def _get_bank_dir(base_dir, processing_year=None, processing_month=None):
    if processing_year is not None and processing_month is not None:
        return build_batch_layout(base_dir, processing_year, processing_month).bank_dir
    return os.path.join(base_dir, '原始数据', '银行流水')


def _get_timesheet_path(base_dir, processing_year=None, processing_month=None):
    if processing_year is not None and processing_month is not None:
        return build_batch_layout(base_dir, processing_year, processing_month).timesheet_path
    master_path = os.path.join(base_dir, '01_基础资料', '02_工时数据', '工时数据.xlsx')
    if os.path.exists(master_path):
        return master_path
    return os.path.join(base_dir, '原始数据', '工时数据', '工时数据.xlsx')


def _get_bonus_path(base_dir, processing_year=None, processing_month=None):
    if processing_year is not None and processing_month is not None:
        return build_batch_layout(base_dir, processing_year, processing_month).bonus_path
    return os.path.join(base_dir, '原始数据', '奖金数据', '年终奖计提2026_ - to财务.xlsx')


def _get_shared_expense_path(base_dir, payroll_year, payroll_month):
    post_year, post_month = _shift_month(payroll_year, payroll_month, 1)
    return build_batch_layout(base_dir, post_year, post_month).shared_expense_path


def _get_co_workorder_path(base_dir, payroll_year, payroll_month):
    post_year, post_month = _shift_month(payroll_year, payroll_month, 1)
    return build_batch_layout(base_dir, post_year, post_month).co_workorder_path


def _get_voucher_template_path(base_dir):
    return os.path.join(os.path.dirname(base_dir), '模版性文件', '2060总账凭证导入-实际202602薪酬.XLS')


def _get_logo_path(base_dir):
    candidates = [
        os.path.join(base_dir, 'rigol_logo.png'),
        os.path.join(os.path.dirname(base_dir), '模版性文件', 'rigol_logo.png'),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return None


def _find_raw_files(raw_dir, payroll_year=None, payroll_month=None):
    paths = []
    seen = set()
    for pattern in RAW_FILE_PATTERNS:
        for path in sorted(glob.glob(os.path.join(raw_dir, pattern))):
            normalized = os.path.normcase(path)
            if normalized in seen:
                continue
            seen.add(normalized)
            paths.append(path)
    if payroll_year is None or payroll_month is None:
        return paths
    matched = []
    for path in paths:
        try:
            file_year, file_month = _extract_payroll_period(path)
        except ValueError:
            continue
        if file_year == payroll_year and file_month == payroll_month:
            matched.append(path)
    return matched

def _build_company_output_path(base_dir, company, payroll_year, payroll_month, input_path):
    post_year, post_month = _shift_month(payroll_year, payroll_month, 1)
    layout = build_batch_layout(base_dir, post_year, post_month)
    company_dir = os.path.join(layout.company_output_root(company), '01_工资整理')
    os.makedirs(company_dir, exist_ok=True)

    output_name = f'{payroll_month}月工资单-整理后.xlsx'
    return os.path.join(company_dir, output_name)


def _build_voucher_output_path(base_dir, company, payroll_year, payroll_month):
    company_code = COMPANY_NAME_TO_CODE[company]
    post_year, post_month = _shift_month(payroll_year, payroll_month, 1)
    layout = build_batch_layout(base_dir, post_year, post_month)
    company_dir = os.path.join(layout.company_output_root(company), '02_实际凭证')
    os.makedirs(company_dir, exist_ok=True)
    return os.path.join(company_dir, f'{company_code}总账凭证导入-实际{post_year}{post_month:02d}薪酬.XLS')


def _build_accrual_voucher_output_path(base_dir, company, payroll_year, payroll_month):
    company_code = COMPANY_NAME_TO_CODE[company]
    post_year, post_month = _shift_month(payroll_year, payroll_month, 1)
    layout = build_batch_layout(base_dir, post_year, post_month)
    company_dir = os.path.join(layout.company_output_root(company), '03_计提凭证')
    os.makedirs(company_dir, exist_ok=True)
    return os.path.join(company_dir, f'{company_code}总账凭证导入-计提{post_year}{post_month:02d}薪酬.XLS')


def _build_bonus_voucher_output_path(base_dir, company, payroll_year, payroll_month):
    company_code = COMPANY_NAME_TO_CODE[company]
    post_year, post_month = _shift_month(payroll_year, payroll_month, 1)
    layout = build_batch_layout(base_dir, post_year, post_month)
    company_dir = os.path.join(layout.company_output_root(company), '04_年终奖凭证')
    os.makedirs(company_dir, exist_ok=True)
    return os.path.join(company_dir, f'{company_code}总账凭证导入-计提{_bonus_label_for_filename(post_year, post_month)}年终奖.XLS')


def _build_co_voucher_output_path(base_dir, payroll_year, payroll_month):
    post_year, post_month = _shift_month(payroll_year, payroll_month, 1)
    layout = build_batch_layout(base_dir, post_year, post_month)
    co_dir = os.path.join(layout.company_output_root('耐数电子'), '06_CO工单分摊')
    os.makedirs(co_dir, exist_ok=True)
    return os.path.join(co_dir, f'2050总账凭证导入-CO工单分摊{post_year}{post_month:02d}.XLS')


def _build_rd_allocation_voucher_output_path(base_dir, company, payroll_year, payroll_month):
    company_code = COMPANY_NAME_TO_CODE[company]
    post_year, post_month = _shift_month(payroll_year, payroll_month, 1)
    layout = build_batch_layout(base_dir, post_year, post_month)
    company_dir = os.path.join(layout.company_output_root(company), '05_研发费用分摊')
    os.makedirs(company_dir, exist_ok=True)
    return os.path.join(company_dir, f'{company_code}总账-研发费用分摊{post_year % 100:02d}{post_month:02d}.XLS')


def _build_key_tax_source_output_path(base_dir, company, processing_year, processing_month):
    company_code = COMPANY_NAME_TO_CODE[company]
    layout = build_batch_layout(base_dir, processing_year, processing_month)
    output_dir = os.path.join(layout.company_output_root(company), '07_重点税源采集信息')
    os.makedirs(output_dir, exist_ok=True)
    return os.path.join(output_dir, f'{company_code}重点税源采集信息{processing_year % 100:02d}{processing_month:02d}.xlsx')


def _ensure_directory(path):
    os.makedirs(path, exist_ok=True)
    return path


def _write_utf8_file(path, content):
    _ensure_directory(os.path.dirname(path))
    with open(path, 'w', encoding='utf-8') as fh:
        fh.write(content)
    return path


def _render_markdown_table(headers, rows):
    lines = ['| ' + ' | '.join(headers) + ' |', '| ' + ' | '.join(['---'] * len(headers)) + ' |']
    for row in rows:
        lines.append('| ' + ' | '.join(str(item) for item in row) + ' |')
    return '\n'.join(lines)


def _render_markdown_list(items, empty_text='无'):
    if not items:
        return f'- {empty_text}'
    return '\n'.join(f'- {item}' for item in items)


def _status_label(status):
    mapping = {
        'ok': '通过',
        'warn': '预警',
        'warning': '预警',
        'error': '阻断',
        'err': '阻断',
        'info': '信息',
    }
    return mapping.get(status, status or '信息')


def _voucher_detail_specs():
    return {
        'A1': {
            'title': 'A1 社保扣款凭证',
            'purpose': '社保发放',
            'debit': '借方取工资表社保相关应付科目汇总。',
            'credit': '贷方取处理月份银行流水中的社保国库流水，按实际流水逐行展开。',
            'sources': [
                '工资整理文件总计行 `F/G/H/I/K/L/M` 七列，分别对应公司养老、公司失业、公司工伤、公司医疗、个人养老、个人失业、个人医疗。',
                '银行流水首表中 `B列公司代码`、`D列交易日期`、`M列付款名称`、`O列付款金额`、`AH列银行科目`。',
            ],
            'timesheet': [],
            'special': ['凭证类型固定 `KZ`；银行行项目原因代码固定写 `202`。'],
            'debit_detail': [
                '借方直接取工资整理文件总计行 `F/G/H/I/K/L/M` 七列金额。',
                '科目映射固定为：`F->2211020002`、`G->2211030002`、`H->2211040001`、`I->2211060002`、`K->2211020001`、`L->2211030001`、`M->2211060001`。',
            ],
            'credit_detail': [
                '先在次月银行流水里筛出 `M列付款名称=国家金库北京市分库` 且 `O列>0` 的国库流水。',
                '以工资表总计行 `F+G+H+I+K+L+M` 为目标，遍历最多 6 笔国库流水组合，匹配成功的流水作为 A1 贷方。',
                '贷方会计科目取银行流水 `AH列`，金额取 `O列`，凭证日期取 `D列` 最大日期。',
            ],
            'checks': ['A1 借贷合计 = 工资表总计行 `F+G+H+I+K+L+M` = 银行社保流水合计。'],
        },
        'A2': {
            'title': 'A2 公积金扣款凭证',
            'purpose': '公积金发放',
            'debit': '借方取公积金应付科目，按公司/个人口径汇总。',
            'credit': '贷方取银行流水中公积金付款行的银行科目与金额。',
            'sources': [
                '工资整理文件总计行 `J列公司公积金` 与 `N列个人公积金` 只用于校验。',
                '银行流水首表中 `M列付款名称=北京住房公积金管理中心` 的付款行，金额取 `O列`，银行科目取 `AH列`。',
            ],
            'timesheet': [],
            'special': ['`2050 / 202602` 特殊月份允许差额 `3650`。', '凭证类型固定 `KZ`；银行行项目原因代码固定写 `202`。'],
            'debit_detail': [
                '脚本不是直接取工资表 `J/N` 入账，而是先汇总银行公积金流水 `O列` 金额。',
                '再将银行总额按 1:1 拆成两行：个人公积金 `2211070001` 与公司公积金 `2211070002`。',
            ],
            'credit_detail': [
                '贷方逐行展开公积金银行流水。',
                '每条贷方的会计科目取银行流水 `AH列`，金额取 `O列`，凭证日期取该批流水 `D列` 最大日期。',
            ],
            'checks': ['工资表总计行 `J+N` 与银行公积金流水 `O列合计` 做核对；若属于特殊月份则按例外规则提示。'],
        },
        'A3': {
            'title': 'A3 工资扣款凭证',
            'purpose': '工资、劳务费与个税发放',
            'debit': '借方取工资表中的劳务费、个税、应付工资，劳务费按行展开。',
            'credit': '贷方取代发工资组合流水和个税国库流水。',
            'sources': [
                '工资整理文件逐行的 `D列项目类型`、`E列金额`、`O列个税`、`Q列实发工资`、`S列成本中心`、`T列内部订单`。',
                '次月银行流水首表中 `P列用途含代发工资` 的工资行，以及 `M列付款名称=国家金库北京市分库` 的国库行。',
            ],
            'timesheet': [],
            'special': ['工资、个税、社保银行侧均支持最多 6 笔流水组合匹配。', '劳务费借方逐行带出成本中心和内部订单。'],
            'debit_detail': [
                '若 `D列=劳务费`，则逐行取 `E列` 税前金额生成借方 `6601110001`，并带出该行 `S列成本中心` 与 `T列内部订单`。',
                '借方 `2221070000` 取匹配到的个税银行流水扣除劳务费个税后的金额；银行个税包含薪资、劳务费和已确认的上月年终奖个税。',
                '若 `D列=薪资`，则把各行 `Q列` 汇总生成借方 `2211010001`。',
            ],
            'credit_detail': [
                '先用工资整理文件总计行 `Q列` 作为匹配目标，在银行流水中寻找最多 6 笔 `P列用途含代发工资` 的组合。',
                '工资贷方的会计科目取这些银行行的 `AH列`，金额取各自 `O列`。',
                '个税贷方用工资整理文件全部 `O列` 个税加上已确认的上月年终奖个税作为目标，在国库流水中寻找最多 6 笔组合；匹配到的流水逐行展开。',
            ],
            'checks': ['非劳务 `Q列汇总 -> 2211010001`。', '个税银行流水组合 -> `2221070000`。', '银行侧必须匹配“工资组合 + 个税国库行”。'],
        },
        'A4': {
            'title': 'A4 上月实际工资入账',
            'purpose': '上月实际工资入账',
            'debit': '借方按成本中心+内部订单归集工资、社保、公积金等费用。',
            'credit': '贷方挂应付社保、公积金、个税、实发工资及其他项目。',
            'sources': [
                '工资整理文件逐行的 `B列部门`、`C列项目`、`D列项目类型`、`E/F/G/H/I/J` 费用列、`O/P/Q` 往来列、`S/T` 核算维度列。',
                'A4 的内部订单默认取 `T列`，但 OSD 特例会结合工时匹配结果重算。',
            ],
            'timesheet': [
                '内部订单默认取工资整理文件 `T列`。',
                '若 `B列=OSD运营支持部` 且 `C列项目` 在工时数据中被识别为“自研”，则内部订单改取 `OSD运营支持部日常工作` 对应订单。',
            ],
            'special': ['排除劳务费。', '`OSD运营支持部` 且项目类型为“自研”时，内部订单按 `OSD运营支持部日常工作` 处理。'],
            'debit_detail': [
                '跳过 `D列=劳务费` 的行。',
                '借方逐行读取 `E/F/G/H/I/J` 六列金额，并按 `S列成本中心 + T列/重算后内部订单` 聚合。',
                '科目映射固定为：`E->6601010001`、`F->6601030008`、`G->6601030002`、`H->6601030003`、`I->6601030005`、`J->6601030006`。',
            ],
            'credit_detail': [
                '贷方的社保、公积金应付科目直接取工资整理文件总计行 `F/G/H/I/J/K/L/M/N`。',
                '个税贷方取所有 `D列=薪资` 行的 `O列` 汇总，实发工资贷方取这些行 `Q列` 汇总，其他项目贷方取这些行 `P列` 汇总。',
                '其中 `F/G/H/I/J/N/M/L/K` 分别映射到 `2211020002/2211030002/2211040001/2211060002/2211070002/2211070001/2211060001/2211030001/2211020001`。',
            ],
            'checks': ['A4 自身借贷必须平。', '若同时选择 A1-A4，会做重点往来科目对冲。'],
        },
        'A5': {
            'title': 'A5 当月工资计提',
            'purpose': '当月工资计提',
            'debit': '借方与 A4 同口径，但只汇总到成本中心层级。',
            'credit': '贷方按应付工资、社保、公积金、个税等科目汇总。',
            'sources': ['与 A4 使用同一份工资整理文件，仍读取 `D/E/F/G/H/I/J/O/P/Q/S`，但借方只保留 `S列成本中心` 维度。'],
            'timesheet': [],
            'special': ['A5 是 A6 的分摊前底稿。'],
            'debit_detail': [
                '跳过 `D列=劳务费` 的行。',
                '借方仍读取 `E/F/G/H/I/J` 六列费用，但仅按 `S列成本中心` 聚合，不再区分 `T列内部订单`。',
                '科目映射与 A4 相同：`E->6601010001`、`F->6601030008`、`G->6601030002`、`H->6601030003`、`I->6601030005`、`J->6601030006`。',
            ],
            'credit_detail': [
                '贷方逻辑与 A4 一致：总计行 `F/G/H/I/J/K/L/M/N` 进应付社保公积金，各薪资行 `O/P/Q` 分别汇总到个税、其他项目、实发工资。',
            ],
            'checks': ['A5 总额应与 A4 总额一致。'],
        },
        'A6': {
            'title': 'A6 研发工时分摊',
            'purpose': '研发工时分摊',
            'debit': '借方按研发工时比例把费用分摊到内部订单。',
            'credit': '贷方先冲回 A5 原计提的非 OSD 费用。',
            'sources': [
                'A5 分摊前底稿来自工资整理文件中非劳务行的 `E/F/G/H/I/J` 六列，并按 `S列成本中心` 聚合。',
                '工时数据 `工时汇总` 页读取 `B列年份`、`C列月份`、`E列部门`、`F列公司`、`K列工时`、`M列内部订单`。',
            ],
            'timesheet': [
                '只看处理月份当月工时，例如 2603 只看 2026年3月工时。',
                '只统计 `K > 0` 的有效工时行。',
                '按 `E列部门 + F列公司 + M列内部订单` 汇总工时，再在部门内计算分摊比例。',
                '`OSD运营支持部` 不参与 A6 分摊。',
            ],
            'special': [
                '科目不变、成本中心不变，贷方只是冲回 A5。',
                '统一文本：`根据研发工时分摊计提的YYYY年M月人工费用`。',
                '分摊结果统一保留两位小数，尾差回补到最后一个内部订单。',
            ],
            'checks': ['A6 冲回金额 = A6 分摊金额。', '若某部门无可用工时，则该部门不能正常分摊。'],
            'debit_detail': [
                '借方金额来源于 A5 已聚合好的 `E/F/G/H/I/J` 六列费用。',
                '每个成本中心先反查所属部门，再把同部门当月工时按内部订单比例分配；借方成本中心沿用原 `S列`，内部订单改写为工时表 `M列`。',
            ],
            'credit_detail': [
                '贷方直接冲回 A5 的原借方金额，科目和成本中心不变，内部订单留空。',
            ],
        },
        'A7': {
            'title': 'A7 年终奖计提',
            'purpose': '年终奖计提',
            'debit': '借方按成本中心计提年终奖费用。',
            'credit': '贷方汇总到应付工资科目。',
            'sources': [
                '奖金文件 `部门统计` 页中 `A列公司`、`B列部门`、`C:N列1-12月金额`。',
                '奖金文件 `计提比例` 页中 `A列月份`、`B列计提系数`。',
            ],
            'timesheet': [],
            'special': ['季度末取 `1-当月累计奖金 × 当月系数`。', '非季度末取 `当月奖金 × 当月系数`。', '系数不追溯历史月份。'],
            'checks': ['每家公司贷方通常汇总为 1 行。'],
            'debit_detail': [
                '先按处理月份判断取数范围：季度末取 `C:N` 中 1-当月累计，非季度末只取对应月份列。',
                '每个部门金额再乘 `计提比例` 页 `B列` 的当月系数，借方科目固定为 `6601010003`，成本中心来自 Mapping 反查结果。',
            ],
            'credit_detail': [
                '贷方按公司汇总为 1 行，科目固定 `2211010001`。',
            ],
        },
        'A8': {
            'title': 'A8 年终奖分摊',
            'purpose': '年终奖分摊',
            'debit': '借方按奖金口径下的工时把年终奖分配到内部订单。',
            'credit': '贷方先冲回 A7 中需分摊的奖金费用。',
            'sources': [
                'A7 已算出的部门奖金金额与成本中心。',
                '工时数据 `工时汇总` 页的 `B/C/E/F/K/M`，季度末按累计月数合并。',
            ],
            'timesheet': [
                '1 月只看 1 月工时。',
                '2 月只看 2 月工时。',
                '3/6/9/12 月看 `1-当月` 的累计工时。',
                '`OSD` 不参与分摊。',
            ],
            'special': [
                'A8 只重分摊需要分摊的奖金部分，不一定等于整张 A7。',
                '统一文本：`根据研发工时分摊计提的期间年终奖`，例如 `根据研发工时分摊计提的2026Q1年终奖`。',
            ],
            'checks': ['A8 分摊前总额 = A8 分摊后总额。'],
            'debit_detail': [
                '借方科目固定 `6601010003`，金额取 A7 每个部门的奖金金额。',
                '借方成本中心沿用该部门成本中心，内部订单按奖金口径工时比例写入工时表 `M列` 对应订单。',
            ],
            'credit_detail': [
                '贷方先按部门冲回 A7 中需要分摊的金额，科目同样固定 `6601010003`，只带成本中心不带内部订单。',
            ],
        },
        'A9': {
            'title': 'A9 CO工单分摊',
            'purpose': 'CO工单分摊',
            'debit': '借方按 CO 工单比例分摊到内部订单，科目按规则映射。',
            'credit': '贷方取待分摊费用，固定成本中心和固定内部订单冲销。',
            'sources': [
                '`CO工单分摊.xlsx` 中 `CO工单清单` 工作表的订单比例底稿。',
                '同文件 `待分摊费用` 工作表中的待分摊费用行。',
            ],
            'timesheet': [],
            'special': [
                '仅适用于 `2050`。',
                '`G列 != 0` 的明细行参与分摊，最后一行加总仅作校验。',
                '贷方固定成本中心 `20502020`、固定内部订单 `9201856`。',
                '部分成本要素映射到 `5001010011`，其余映射到 `5001010012`。',
            ],
            'checks': ['`CO工单分摊` 明细 G 列合计必须等于最后一行加总。', 'A9 借贷合计必须一致。'],
            'debit_detail': [
                '分摊比例底稿来自 `CO工单分摊` 页：`A列内部订单`、`E列实际成本借方`、`G列总的实际成本`。',
                '仅 `G列 != 0` 的明细参与分摊，比例分母取这些明细 `E列` 合计。',
                '借方内部订单取 `A列`，借方科目按 `待分摊费用` 页 `A列成本要素` 映射到 `5001010011/5001010012`。',
            ],
            'credit_detail': [
                '贷方逐行读取 `待分摊费用` 页 `A列成本要素` 与 `D列待分摊金额`。',
                '贷方成本中心固定 `20502020`，内部订单固定 `9201856`，并在 SAP 模板 `O列反记账` 写 `x`。',
            ],
        },
        'A10': {
            'title': 'A10 研发费用分摊',
            'purpose': '研发费用分摊',
            'debit': '借方按费控科目+部门聚合后的金额，再按研发工时分摊到项目订单。',
            'credit': '贷方先按原费用科目、原成本中心、原内部订单冲回待分摊费用。',
            'sources': [
                '`待分摊费用YYMM.xlsx` 中 E/J/L/S/T/U 等关键字段。',
                '工时数据中的部门工时与内部订单。',
                'Mapping 表中的成本中心反查部门规则。',
            ],
            'timesheet': [
                '分摊月份只看处理月份当月工时。',
                '绝不跨公司分摊。',
                '默认按部门内工时分摊。',
                '`OSD` 不参与分摊。',
                '部门先用成本中心编码反查，再去工时数据中取比例，而不是直接信任文本名称。',
            ],
            'special': [
                'A10 先按 `费控科目 + 部门` 聚合，再把同组正负金额抵减后分摊。',
                '`待分摊费用` 中 `BG列功能范围 = 1000` 的费用不参与分摊。',
                '`LHD硬件逻辑部` 与 `HLD硬件逻辑部` 视为同一部门。',
                '`HLD硬件逻辑部` 分摊时，`9201856` 不进入分母，也不作为分子。',
                '`2050 / 20502050` 的借方分摊按 `20502060` 工时口径处理，但贷方仍维持原成本中心。',
                '`2060 / 20602020` 若本部门无可用工时，则改按整个 `耐数信息` 公司工时分摊。',
                '统一文本：`根据研发工时分摊YYYY年M月研发费用`。',
            ],
            'checks': ['A10 借方合计 = 贷方合计。', '每组聚合后的贷方金额应等于借方分摊合计。'],
            'debit_detail': [
                '源数据先读 `E列公司代码`、`J列会计科目`、`L列金额`、`S列成本中心`、`T列部门文本`、`U列原内部订单`、`BG列功能范围`。',
                '脚本先用 `S列` 通过 Mapping 反查标准部门，再按 `J列会计科目 + 标准部门` 聚合 `L列` 金额；同组正负先抵减，净额为 0 的组直接跳过。',
                '借方科目仍取原 `J列`，借方成本中心默认取原 `S列`，若 `2050/20502050` 特例则改用 `20502060` 作为分摊成本中心；内部订单按工时表 `M列` 比例写入。',
            ],
            'credit_detail': [
                '贷方按聚合后的组逐组冲回，科目取原 `J列`，金额取聚合后的净额，成本中心取原 `S列`，内部订单取原 `U列`。',
                '凭证类型固定 `SA`，并保持“一个聚合组一条贷方”。',
            ],
        },
    }


def _voucher_summary_rows(run_options):
    descriptions = _voucher_detail_specs()
    return [
        (
            voucher,
            descriptions[voucher]['purpose'],
            descriptions[voucher]['debit'],
            descriptions[voucher]['credit'],
        )
        for voucher in run_options.vouchers
        if voucher in descriptions
    ]


def _voucher_example_assets():
    return {
        'A1': [
            ('a1-payroll-social.png', '工资整理文件：总计行中 F:I 与 K:M 七列社保金额。'),
            ('a1-bank-social.png', '银行流水：同批两条国库流水中，较大金额对应社保。'),
        ],
        'A2': [
            ('a2-payroll-fund.png', '工资整理文件：总计行中的公司/个人公积金。'),
            ('a2-bank-fund.png', '银行流水：北京住房公积金管理中心对应付款行。'),
        ],
        'A3': [
            ('a3-payroll-salary.png', '工资整理文件：劳务费逐行、总计行 O/Q 汇总、以及 S/T 映射结果。'),
            ('a3-bank-salary-tax.png', '银行流水：代发工资组合与个税国库行。'),
        ],
        'A4': [
            ('a4-payroll-source.png', '工资整理文件：按成本中心和内部订单聚合后的来源区域。'),
            ('a4-voucher.png', 'A4 凭证：借方费用与贷方应付项目。'),
        ],
        'A5': [
            ('a5-voucher.png', 'A5 凭证：借方只到成本中心层级。'),
        ],
        'A6': [
            ('a6-timesheet.png', 'A6 工时示例：处理月份当月工时。'),
            ('a6-detail.png', 'A5-A6 核对表：A5 金额到 A6 分摊金额的展开。'),
            ('a6-voucher.png', 'A6 凭证：先冲回，再按工时分配到内部订单。'),
        ],
        'A7': [
            ('a7-bonus-source.png', 'A7 奖金来源：季度累计奖金基础金额。'),
            ('a7-bonus-ratio.png', 'A7 奖金系数：3 月对应奖金系数。'),
            ('a7-voucher.png', 'A7 凭证：借方按成本中心，贷方公司汇总。'),
        ],
        'A8': [
            ('a8-timesheet-q1.png', 'A8 工时示例：季度末看 1-3 月累计工时。'),
            ('a8-voucher.png', 'A8 凭证：冲回 A7 后重新分配到内部订单。'),
        ],
        'A9': [
            ('a9-co-source.png', 'A9 来源示例：CO工单清单工作表中的订单与比例底稿。'),
            ('a9-co-expense.png', 'A9 来源示例：待分摊费用工作表中的成本要素与金额。'),
        ],
        'A10': [
            ('a10-shared-expense-source.png', 'A10 来源示例：待分摊费用中的公司、科目、金额、成本中心和原订单。'),
        ],
    }


def _copy_example_images_for_batch(base_dir, archive_root, batch_code, vouchers):
    if batch_code != '2603':
        return
    asset_map = _voucher_example_assets()
    target_dir = os.path.join(archive_root, EXAMPLE_IMAGE_DIRNAME)
    os.makedirs(target_dir, exist_ok=True)
    for voucher in vouchers:
        for file_name, _caption in asset_map.get(voucher, []):
            source_path = os.path.join(base_dir, 'doc_assets', file_name)
            target_path = os.path.join(target_dir, file_name)
            if os.path.exists(source_path):
                shutil.copyfile(source_path, target_path)


def _archive_timesheet_snapshot(layout, timesheet_path):
    if not timesheet_path or not os.path.isfile(timesheet_path):
        return ''
    snapshot_dir = os.path.join(layout.archive_root, '06_基础资料快照')
    _ensure_directory(snapshot_dir)
    snapshot_path = os.path.join(snapshot_dir, f'{layout.batch_code}-工时数据留档.xlsx')
    if os.path.abspath(timesheet_path) != os.path.abspath(snapshot_path):
        shutil.copyfile(timesheet_path, snapshot_path)
    return snapshot_path


def _render_example_images(base_dir, batch_code, voucher):
    if batch_code != '2603':
        return ''
    asset_map = _voucher_example_assets()
    examples = asset_map.get(voucher, [])
    if not examples:
        return ''
    lines = ['#### 2026年3月样例截图', '']
    for file_name, caption in examples:
        asset_path = os.path.join(base_dir, 'doc_assets', file_name)
        if os.path.exists(asset_path):
            lines.append(f'![{voucher} 示例](./{EXAMPLE_IMAGE_DIRNAME}/{file_name})')
            lines.append(f'_{caption}_')
            lines.append('')
    return '\n'.join(lines).strip()


def _render_voucher_detail_sections(run_options, base_dir, batch_code):
    specs = _voucher_detail_specs()
    blocks = []
    for voucher in run_options.vouchers:
        spec = specs.get(voucher)
        if not spec:
            continue
        blocks.extend(
            [
                f'### {spec["title"]}',
                '',
                f'- 用途：{spec["purpose"]}',
                '',
                '#### 取数来源',
                '',
                _render_markdown_list(spec.get('sources', [])),
                '',
                '#### 借方生成逻辑',
                '',
                _render_markdown_list(spec.get('debit_detail', [spec['debit']])),
                '',
                '#### 贷方生成逻辑',
                '',
                _render_markdown_list(spec.get('credit_detail', [spec['credit']])),
                '',
            ]
        )
        if spec.get('timesheet'):
            blocks.extend(
                [
                    '#### 工时选取逻辑',
                    '',
                    _render_markdown_list(spec['timesheet']),
                    '',
                ]
            )
        if spec.get('special'):
            blocks.extend(
                [
                    '#### 特殊规则',
                    '',
                    _render_markdown_list(spec['special']),
                    '',
                ]
            )
        if spec.get('checks'):
            blocks.extend(
                [
                    '#### 关键勾稽与校验',
                    '',
                    _render_markdown_list(spec['checks']),
                    '',
                ]
            )
        example_block = _render_example_images(base_dir, batch_code, voucher)
        if example_block:
            blocks.extend([example_block, ''])
    return '\n'.join(blocks).strip()
    return [
        (
            voucher,
            descriptions[voucher]['purpose'],
            descriptions[voucher]['debit'],
            descriptions[voucher]['credit'],
        )
        for voucher in run_options.vouchers
        if voucher in descriptions
    ]


def _write_run_artifacts(
    base_dir,
    run_options,
    input_path,
    mapping_path,
    bank_dir,
    timesheet_path,
    bonus_path,
    co_path,
    shared_path,
    output_paths,
    voucher_paths,
    voucher_validation_summary,
    log_entries,
    bank_scan_stats,
    compensation_report_path=None,
):
    layout = build_batch_layout(base_dir, run_options.processing_year, run_options.processing_month)
    input_manifest_path = os.path.join(layout.archive_root, f'{layout.batch_code}-输入清单.md')
    run_log_path = os.path.join(layout.archive_root, f'{layout.batch_code}-运行日志.md')
    process_note_path = os.path.join(layout.archive_root, f'{layout.batch_code}-过程说明.md')
    output_manifest_path = os.path.join(layout.archive_root, f'{layout.batch_code}-输出清单.md')
    warning_logs = [entry['text'] for entry in log_entries if entry.get('tag') == 'warn']
    error_logs = [entry['text'] for entry in log_entries if entry.get('tag') in ('error', 'err')]
    voucher_logic_rows = _voucher_summary_rows(run_options)
    _copy_example_images_for_batch(base_dir, layout.archive_root, layout.batch_code, run_options.vouchers)
    timesheet_snapshot_path = _archive_timesheet_snapshot(layout, timesheet_path)
    voucher_detail_markdown = _render_voucher_detail_sections(run_options, base_dir, layout.batch_code)

    input_rows = [
        ('原始工资单', '必须', input_path, '本次唯一工资源文件'),
        ('Mapping表', '必须', mapping_path, '基础主数据'),
        ('工时数据', '必须', timesheet_path, '基础资料来源；已按本批次单独留档'),
        ('银行流水目录', '按需', bank_dir if requires_bank_data(run_options) else '未使用', 'A1-A3 时启用'),
        ('奖金数据', '按需', bonus_path if requires_bonus_data(run_options) else '未使用', 'A7-A8 时启用'),
        ('CO工单分摊', '按需', co_path if requires_co_data(run_options) else '未使用', 'A9 时启用'),
        ('待分摊费用', '按需', shared_path if requires_shared_expense_data(run_options) else '未使用', 'A10 时启用'),
    ]
    snapshot_rows = [('工时数据留档', timesheet_snapshot_path)] if timesheet_snapshot_path else []
    input_content = '\n'.join(
        [
            f'# {layout.batch_name} 输入清单',
            '',
            '## 一页摘要',
            '',
            f'- 批次编码：`{layout.batch_code}`',
            f'- 处理月份：`{run_options.processing_label}`',
            f'- 工资所属月份：`{run_options.payroll_label}`',
            f'- 公司范围：`{"、".join(run_options.companies) or "未选"}`',
            f'- 凭证范围：`{"、".join(run_options.vouchers) or "未选"}`',
            f'- 已记录输入项：`{len(input_rows)}` 项',
            '',
            '## 批次定位',
            '',
            _render_markdown_table(
                ['项目', '值'],
                [
                    ('月度输入目录', layout.monthly_input_root),
                    ('运行输出目录', layout.run_output_root),
                    ('归档留痕目录', layout.archive_root),
                ],
            ),
            '',
            '## 实际使用文件',
            '',
            _render_markdown_table(['资料', '要求', '路径/值', '用途'], input_rows),
            '',
            '## 基础资料版本留档',
            '',
            _render_markdown_table(['资料', '留档路径'], snapshot_rows or [('无', '无')]),
            '',
            '## 银行流水扫描统计',
            '',
            _render_markdown_table(
                ['指标', '值'],
                [
                    ('扫描文件数', bank_scan_stats.get('source_file_count', 0)),
                    ('扫描记录数', bank_scan_stats.get('scanned_row_count', 0)),
                    ('保留记录数', bank_scan_stats.get('kept_row_count', 0)),
                    ('去重记录数', bank_scan_stats.get('deduped_row_count', 0)),
                ],
            ),
            '',
            '## 使用说明',
            '',
            '- 本文档用于回答“本次到底用了什么输入资料”。',
            '- 若后续重跑本批次，应优先核对本清单中的路径与月份口径。',
            '',
        ]
    )

    log_rows = [(idx, _status_label(entry.get('tag')), entry['text']) for idx, entry in enumerate(log_entries, start=1)]
    run_log_content = '\n'.join(
        [
            f'# {layout.batch_name} 运行日志',
            '',
            '## 执行摘要',
            '',
            f'- 运行时间：`{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}`',
            f'- 处理月份：`{run_options.processing_label}`',
            f'- 公司范围：`{"、".join(run_options.companies)}`',
            f'- 凭证范围：`{"、".join(run_options.vouchers)}`',
            f'- 记录日志：`{len(log_entries)}` 条',
            f'- 预警日志：`{len(warning_logs)}` 条',
            f'- 错误日志：`{len(error_logs)}` 条',
            '',
            '## 重点关注',
            '',
            _render_markdown_list(warning_logs, empty_text='本次运行未出现预警日志'),
            '',
            '## 运行时间线',
            '',
            _render_markdown_table(['序号', '级别', '内容'], log_rows),
            '',
        ]
    )

    process_note_content = '\n'.join(
        [
            f'# {layout.batch_name} 过程说明',
            '',
            '## 文档用途',
            '',
            '- 本文档用于解释本批次凭证是按什么口径生成的。',
            '- 重点服务于复核、交接和后续审计沟通。',
            '',
            '## 本次处理口径',
            '',
            _render_markdown_table(
                ['项目', '值'],
                [
                    ('处理月份', run_options.processing_label),
                    ('工资所属月份', run_options.payroll_label),
                    ('公司范围', '、'.join(run_options.companies)),
                    ('凭证范围', '、'.join(run_options.vouchers)),
                ],
            ),
            '',
            '## 处理主链路',
            '',
            '1. 读取工资单、Mapping、工时和按需资料。',
            '2. 对工资单执行 A/B 列填充、Q 列校验、S/T 列映射。',
            '3. 如选择 A1-A3，则扫描银行流水目录、筛选处理月份并自动去重。',
            '4. 按公司拆分整理后工资单，再生成所选凭证。',
            '5. 同步输出输入清单、预检报告、运行日志、过程说明和输出清单。',
            '',
            '## 凭证生成逻辑摘要',
            '',
            _render_markdown_table(['凭证', '用途', '借方生成逻辑', '贷方生成逻辑'], voucher_logic_rows),
            '',
            '## 凭证明细逻辑',
            '',
            voucher_detail_markdown,
            '',
            '## 关键规则',
            '',
            '- 银行流水按目录全量扫描后再按处理月份筛选，并对完全相同行自动去重。',
            '- A10 先按费控科目 + 部门聚合，再进行工时分摊。',
            '- 工时数据作为可滚动更新的基础资料维护；每次正式运行都会把当次使用版本复制到本批次留痕目录。',
            '- 所有比例分摊统一保留两位小数，尾差回补到最后一个分摊对象。',
            '',
            '## 建议复核点',
            '',
            _render_markdown_list(
                warning_logs if warning_logs else ['优先核对工资整理文件、银行核对摘要和跨凭证对冲结果。'],
                empty_text='优先核对工资整理文件、银行核对摘要和跨凭证对冲结果。',
            ),
            '',
        ]
    )

    payroll_output_rows = [(company, path) for company, path in output_paths.items()]
    voucher_output_rows = [(category, path) for category, path in voucher_paths.items()]
    supplemental_output_rows = (
        [('实发薪酬表', compensation_report_path)]
        if compensation_report_path
        else []
    )
    validation_rows = []
    for company, validation in voucher_validation_summary.items():
        issue_count = len(validation.get('cross_group_issues', []))
        validation_rows.append((company, '通过' if issue_count == 0 else f'异常 {issue_count} 项'))
    output_manifest_content = '\n'.join(
        [
            f'# {layout.batch_name} 输出清单',
            '',
            '## 输出摘要',
            '',
            f'- 输出根目录：`{layout.run_output_root}`',
            f'- 留痕根目录：`{layout.archive_root}`',
            f'- 工资整理文件：`{len(payroll_output_rows)}` 个',
            f'- 凭证文件：`{len(voucher_output_rows)}` 个',
            f'- 补充报表：`{len(supplemental_output_rows)}` 个',
            f'- 基础资料留档：`{len(snapshot_rows)}` 个',
            f'- 公司级校验结果：`{len(validation_rows)}` 项',
            '',
            '## 工资整理文件',
            '',
            _render_markdown_table(['公司', '路径'], payroll_output_rows or [('无', '无')]),
            '',
            '## 凭证文件',
            '',
            _render_markdown_table(['对象', '路径'], voucher_output_rows or [('无', '无')]),
            '',
            '## 补充报表',
            '',
            _render_markdown_table(['对象', '路径'], supplemental_output_rows or [('无', '无')]),
            '',
            '## 基础资料留档',
            '',
            _render_markdown_table(['资料', '路径'], snapshot_rows or [('无', '无')]),
            '',
            '## 校验结果',
            '',
            _render_markdown_table(['公司', '结果'], validation_rows or [('无', '无')]),
            '',
            '## 说明',
            '',
            '- 本文档用于回答“本次到底产出了什么文件”。',
            '- 若后续重跑，同名文件会被最新结果覆盖，应结合运行日志一起查看。',
            '',
        ]
    )

    return {
        'input_manifest': _write_utf8_file(input_manifest_path, input_content),
        'run_log': _write_utf8_file(run_log_path, run_log_content),
        'process_note': _write_utf8_file(process_note_path, process_note_content),
        'output_manifest': _write_utf8_file(output_manifest_path, output_manifest_content),
        'timesheet_snapshot': timesheet_snapshot_path,
    }


def _allocate_amount_by_weights(total_amount, weighted_items):
    total_amount = _to_money(total_amount)
    positive_items = [(key, _to_decimal(weight)) for key, weight in weighted_items if _to_decimal(weight) > Decimal('0')]
    if total_amount == Decimal('0.00') or not positive_items:
        return {}

    total_weight = sum((weight for _, weight in positive_items), Decimal('0'))
    if total_weight <= Decimal('0'):
        return {}

    allocations = {}
    remainders = []
    allocated = Decimal('0.00')
    cent = Decimal('0.01')

    for key, weight in positive_items:
        raw = total_amount * weight / total_weight
        floored = raw.quantize(cent, rounding=ROUND_DOWN)
        allocations[key] = floored
        allocated += floored
        remainders.append((raw - floored, key))

    remaining_cents = int(((total_amount - allocated) / cent).quantize(Decimal('1'), rounding=ROUND_HALF_UP))
    remainders.sort(key=lambda item: (item[0], item[1]), reverse=True)

    for idx in range(remaining_cents):
        _, key = remainders[idx % len(remainders)]
        allocations[key] += cent

    return {key: amount.quantize(cent) for key, amount in allocations.items() if amount != Decimal('0.00')}


def _extract_payroll_period(input_path):
    name = os.path.basename(input_path)
    match = re.search(r'(\d{6})', name)
    if not match:
        raise ValueError(f'无法从工资表文件名识别账期：{name}')
    yyyymm = match.group(1)
    return int(yyyymm[:4]), int(yyyymm[4:])


def _shift_month(year, month, offset):
    total = year * 12 + (month - 1) + offset
    return total // 12, total % 12 + 1


def _normalize_department_for_order(dept):
    text = _normalize_text(dept)
    aliases = {
        'NAISHU SOLUTION': 'OSD运营支持部',
        'LHD硬件逻辑部': 'HLD硬件逻辑部',
        '逻辑部': 'HLD硬件逻辑部',
        '运营支持部OSD': 'OSD运营支持部',
        'OSD运营支持部': 'OSD运营支持部',
    }
    return aliases.get(text, text)


def _build_internal_order_candidates(dept, project):
    if not _is_blank_project(project):
        return [_normalize_text(project)]

    dept_text = _normalize_department_for_order(dept)
    if not dept_text:
        return []

    return [f'{dept_text}日常工作']


def _lookup_internal_order(order_map, dept, project):
    for candidate in _build_internal_order_candidates(dept, project):
        value = order_map.get(candidate, '')
        if value:
            return value, candidate
    return '', (_build_internal_order_candidates(dept, project)[0] if _build_internal_order_candidates(dept, project) else '')


def _match_order_from_kok3(kok3_index, company_code, project_name, extra_lookup):
    """先对描述做精确匹配，再允许「搜索词包含于 KOK3 描述(I列)」的包含匹配。"""
    if not project_name:
        return '', project_name

    search_terms = [project_name]
    extra_term = extra_lookup.get(project_name, '')
    if extra_term and extra_term not in search_terms:
        search_terms.append(extra_term)

    rows = kok3_index.get(company_code, [])

    for search_text in search_terms:
        exact = [(desc, order) for desc, order in rows if search_text == desc]
        if exact:
            orders = {order for _, order in exact}
            if len(orders) == 1:
                return exact[0][1], search_text
            return '', search_text

    for search_text in search_terms:
        contain = [(desc, order) for desc, order in rows if search_text != desc and search_text in desc]
        if contain:
            orders = {order for _, order in contain}
            if len(orders) == 1:
                return contain[0][1], search_text
            return '', search_text

    return '', (search_terms[-1] if search_terms else project_name)


def _match_timesheet_internal_orders(timesheet_path, save_changes=True):
    wb = load_workbook(timesheet_path)
    summary_ws = wb['工时汇总']
    kok3_ws = wb['KOK3']
    extra_ws = wb['补充逻辑'] if '补充逻辑' in wb.sheetnames else None

    extra_lookup = {}
    if extra_ws is not None:
        for row_idx in range(1, extra_ws.max_row + 1):
            project_name = _normalize_text(extra_ws.cell(row=row_idx, column=1).value)
            search_text = _normalize_text(extra_ws.cell(row=row_idx, column=2).value)
            if project_name and search_text:
                extra_lookup[project_name] = search_text

    company_code_by_name = {'耐数电子': 2050, '耐数信息': 2060}
    kok3_index = {}
    for row_idx in range(2, kok3_ws.max_row + 1):
        order = _format_code(kok3_ws.cell(row=row_idx, column=1).value)
        desc = _normalize_text(kok3_ws.cell(row=row_idx, column=9).value)
        company_code = _format_code(kok3_ws.cell(row=row_idx, column=12).value)
        if not order or not desc or not company_code:
            continue
        try:
            company_key = int(company_code)
        except ValueError:
            continue
        kok3_index.setdefault(company_key, []).append((desc, order))

    matched_rows = 0
    unmatched_rows = []
    used_extra_rows = 0

    for row_idx in range(2, summary_ws.max_row + 1):
        company = _normalize_text(summary_ws.cell(row=row_idx, column=6).value)
        project = _normalize_text(summary_ws.cell(row=row_idx, column=7).value)
        target_cell = summary_ws.cell(row=row_idx, column=13)
        target_cell.value = None

        if not company or not project:
            continue

        company_code = company_code_by_name.get(company)
        if company_code is None:
            unmatched_rows.append({'row_idx': row_idx, 'company': company, 'project': project, 'reason': 'unknown_company'})
            continue

        order, used_key = _match_order_from_kok3(kok3_index, company_code, project, extra_lookup)
        if order:
            target_cell.value = order
            matched_rows += 1
            if used_key and used_key != project:
                used_extra_rows += 1
        else:
            unmatched_rows.append({'row_idx': row_idx, 'company': company, 'project': project, 'reason': 'no_match'})

    if save_changes:
        wb.save(timesheet_path)

    project_index = {}
    allocation_index = {}
    for row_idx in range(2, summary_ws.max_row + 1):
        year_value = summary_ws.cell(row=row_idx, column=2).value
        month_value = summary_ws.cell(row=row_idx, column=3).value
        dept = _normalize_department_for_order(summary_ws.cell(row=row_idx, column=5).value)
        company = _normalize_text(summary_ws.cell(row=row_idx, column=6).value)
        project = _normalize_text(summary_ws.cell(row=row_idx, column=7).value)
        hours = _to_decimal(summary_ws.cell(row=row_idx, column=11).value)
        order = _format_code(summary_ws.cell(row=row_idx, column=13).value)
        item_type = _normalize_text(summary_ws.cell(row=row_idx, column=10).value)
        if not company or not project or not order:
            continue
        key = (company, project)
        if key not in project_index:
            project_index[key] = {'order': order, 'types': set()}
        elif project_index[key]['order'] != order:
            raise ValueError(
                f'工时汇总中「{company}」+「{project}」对应多个不同内部订单：'
                f'{project_index[key]["order"]} 与 {order}，请核对工时数据'
            )
        if item_type:
            project_index[key]['types'].add(item_type)

        month_match = re.search(r'(\d{1,2})', _normalize_text(month_value))
        month_num = int(month_match.group(1)) if month_match else None
        try:
            year_num = int(year_value) if year_value is not None else None
        except (TypeError, ValueError):
            year_num = None
        if year_num and month_num and dept and hours > Decimal('0'):
            alloc_bucket = allocation_index.setdefault((year_num, month_num, company, dept), {})
            alloc_bucket[order] = alloc_bucket.get(order, Decimal('0')) + hours

    return project_index, {
        'matched_rows': matched_rows,
        'used_extra_rows': used_extra_rows,
        'unmatched_rows': unmatched_rows,
    }


def _load_timesheet_match_context(timesheet_path):
    wb = load_workbook(timesheet_path, data_only=True)
    summary_ws = wb['工时汇总']
    kok3_ws = wb['KOK3']
    extra_ws = wb['补充逻辑'] if '补充逻辑' in wb.sheetnames else None

    extra_lookup = {}
    if extra_ws is not None:
        for row_idx in range(1, extra_ws.max_row + 1):
            project_name = _normalize_text(extra_ws.cell(row=row_idx, column=1).value)
            search_text = _normalize_text(extra_ws.cell(row=row_idx, column=2).value)
            if project_name and search_text:
                extra_lookup[project_name] = search_text

    kok3_index = {}
    for row_idx in range(2, kok3_ws.max_row + 1):
        order = _format_code(kok3_ws.cell(row=row_idx, column=1).value)
        desc = _normalize_text(kok3_ws.cell(row=row_idx, column=9).value)
        company_code = _format_code(kok3_ws.cell(row=row_idx, column=12).value)
        if not order or not desc or not company_code:
            continue
        try:
            company_key = int(company_code)
        except ValueError:
            continue
        kok3_index.setdefault(company_key, []).append((desc, order))

    company_code_by_name = {'耐数电子': 2050, '耐数信息': 2060}
    project_index = {}
    allocation_index = {}
    for row_idx in range(2, summary_ws.max_row + 1):
        year_value = summary_ws.cell(row=row_idx, column=2).value
        month_value = summary_ws.cell(row=row_idx, column=3).value
        dept = _normalize_department_for_order(summary_ws.cell(row=row_idx, column=5).value)
        company = _normalize_text(summary_ws.cell(row=row_idx, column=6).value)
        project = _normalize_text(summary_ws.cell(row=row_idx, column=7).value)
        hours = _to_decimal(summary_ws.cell(row=row_idx, column=11).value)
        order = _format_code(summary_ws.cell(row=row_idx, column=13).value)
        item_type = _normalize_text(summary_ws.cell(row=row_idx, column=10).value)
        if not order:
            company_code = company_code_by_name.get(company)
            if company_code is not None:
                order, _ = _match_order_from_kok3(kok3_index, company_code, project, extra_lookup)
        if not company or not project or not order:
            continue
        key = (company, project)
        if key not in project_index:
            project_index[key] = {'order': order, 'types': set()}
        elif project_index[key]['order'] != order:
            raise ValueError(
                f'工时汇总中「{company}」+「{project}」对应多个不同内部订单：'
                f'{project_index[key]["order"]} 与 {order}，请核对工时数据'
            )
        if item_type:
            project_index[key]['types'].add(item_type)

        month_match = re.search(r'(\d{1,2})', _normalize_text(month_value))
        month_num = int(month_match.group(1)) if month_match else None
        try:
            year_num = int(year_value) if year_value is not None else None
        except (TypeError, ValueError):
            year_num = None
        if year_num and month_num and dept and hours > Decimal('0'):
            alloc_bucket = allocation_index.setdefault((year_num, month_num, company, dept), {})
            alloc_bucket[order] = alloc_bucket.get(order, Decimal('0')) + hours

    return {
        'project_index': project_index,
        'extra_lookup': extra_lookup,
        'kok3_index': kok3_index,
        'allocation_index': allocation_index,
    }


def _lookup_internal_order_from_timesheet(timesheet_context, company, dept, project):
    company_code_by_name = {'耐数电子': 2050, '耐数信息': 2060}
    cc = company_code_by_name.get(company)
    dept_text = _normalize_department_for_order(dept)
    candidates = _build_internal_order_candidates(dept, project)

    for candidate in candidates:
        entry = timesheet_context['project_index'].get((company, candidate))
        if entry and entry.get('order'):
            return entry['order'], candidate
        if cc is not None:
            order, used_key = _match_order_from_kok3(
                timesheet_context['kok3_index'],
                cc,
                candidate,
                timesheet_context['extra_lookup'],
            )
            if order:
                return order, used_key
    daily_project = f'{dept_text}日常工作' if dept_text else ''
    override_order = DEPARTMENT_DAILY_ORDER_OVERRIDES.get((company, dept_text), '')
    if override_order and daily_project in candidates:
        return override_order, daily_project
    return '', (candidates[0] if candidates else '')


def _resolve_a4_internal_order(timesheet_context, company, dept, project, default_order):
    dept_text = _normalize_department_for_order(dept)
    if dept_text != 'OSD运营支持部' or _is_blank_project(project):
        return default_order

    project_entry = timesheet_context['project_index'].get((company, _normalize_text(project)))
    types = project_entry.get('types', set()) if project_entry else set()
    if not project_entry or not any('自研' in t for t in types):
        return default_order

    daily_entry = timesheet_context['project_index'].get((company, 'OSD运营支持部日常工作'))
    if daily_entry and daily_entry.get('order'):
        return daily_entry['order']
    company_code_by_name = {'耐数电子': 2050, '耐数信息': 2060}
    order, _ = _match_order_from_kok3(
        timesheet_context['kok3_index'],
        company_code_by_name.get(company),
        'OSD运营支持部日常工作',
        timesheet_context['extra_lookup'],
    )
    if order:
        return order
    return default_order


def _load_bonus_context(bonus_path, cost_center_map):
    # 奖金源文件只用于读取基础数据，必须保持原始格式完全不被脚本触碰。
    wb = load_workbook(bonus_path, data_only=True, read_only=True)
    stats_ws = wb['部门统计']
    ratio_ws = wb['计提比例']

    ratio_map = {}
    for row_idx in range(1, ratio_ws.max_row + 1):
        month_text = _normalize_text(ratio_ws.cell(row=row_idx, column=1).value)
        ratio_value = ratio_ws.cell(row=row_idx, column=2).value
        month_match = re.search(r'(\d{1,2})', month_text)
        if not month_match or ratio_value in (None, ''):
            continue
        ratio_map[int(month_match.group(1))] = _to_decimal(ratio_value)

    rows_by_company = {'耐数电子': [], '耐数信息': []}
    for row_idx in range(2, stats_ws.max_row + 1):
        company = _bonus_company_alias(stats_ws.cell(row=row_idx, column=1).value)
        dept = _extract_bonus_department(stats_ws.cell(row=row_idx, column=2).value)
        if not company or not dept:
            continue

        monthly_amounts = {}
        for month in range(1, 13):
            monthly_amounts[month] = _to_money(stats_ws.cell(row=row_idx, column=2 + month).value)

        cost_center = cost_center_map.get(company, {}).get(dept, '')
        rows_by_company[company].append(
            {
                'row_idx': row_idx,
                'dept': dept,
                'cost_center': cost_center,
                'monthly_amounts': monthly_amounts,
            }
        )

    return {
        'rows_by_company': rows_by_company,
        'ratio_map': ratio_map,
    }


def _load_co_workorder_context(co_path):
    wb = load_workbook(co_path, data_only=True, read_only=True)
    alloc_ws = wb['CO工单分摊'] if 'CO工单分摊' in wb.sheetnames else wb[wb.sheetnames[0]]
    expense_ws = wb['待分摊费用'] if '待分摊费用' in wb.sheetnames else (wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else None)
    if expense_ws is None:
        raise ValueError('CO工单分摊文件缺少“待分摊费用”工作表')

    allocation_rows = []
    for row_idx in range(2, alloc_ws.max_row):
        order = _format_code(alloc_ws.cell(row=row_idx, column=1).value)
        total_actual = _to_money(alloc_ws.cell(row=row_idx, column=7).value)
        if total_actual == Decimal('0.00'):
            continue
        ratio_base = _to_money(alloc_ws.cell(row=row_idx, column=5).value)
        if not order:
            raise ValueError(f'CO工单分摊 第 {row_idx} 行 A列内部订单为空，无法分摊')
        if ratio_base <= Decimal('0.00'):
            raise ValueError(f'CO工单分摊 第 {row_idx} 行 E列实际成本借方必须大于 0')
        allocation_rows.append(
            {
                'row_idx': row_idx,
                'order': order,
                'ratio_base': ratio_base,
                'total_actual': total_actual,
            }
        )

    if not allocation_rows:
        raise ValueError('CO工单分摊中未找到 G列不为0 的可分摊明细行')

    total_ratio_base = sum((item['ratio_base'] for item in allocation_rows), Decimal('0.00'))
    if total_ratio_base <= Decimal('0.00'):
        raise ValueError('CO工单分摊可分摊明细行的 E列实际成本借方合计必须大于 0')

    summary_total = _to_money(alloc_ws.cell(row=alloc_ws.max_row, column=7).value)
    selected_total = sum((item['total_actual'] for item in allocation_rows), Decimal('0.00')).quantize(Decimal('0.01'))
    if summary_total != selected_total:
        raise ValueError(
            f'CO工单分摊校验失败：最后一行 G列合计 {summary_total} 与明细 G列合计 {selected_total} 不一致'
        )

    expense_rows = []
    for row_idx in range(2, expense_ws.max_row + 1):
        source_account = _format_code(expense_ws.cell(row=row_idx, column=1).value)
        amount = _to_money(expense_ws.cell(row=row_idx, column=4).value)
        if not source_account or amount == Decimal('0.00'):
            continue
        debit_gl = (
            CO_DEBIT_GL_BY_SOURCE['special']
            if source_account in CO_SPECIAL_COST_ELEMENTS
            else CO_DEBIT_GL_BY_SOURCE['other']
        )
        expense_rows.append(
            {
                'row_idx': row_idx,
                'source_account': source_account,
                'debit_gl': debit_gl,
                'amount': amount,
            }
        )

    if not expense_rows:
        raise ValueError('待分摊费用工作表中未找到 D列非0 的待分摊费用')

    return {
        'allocation_rows': allocation_rows,
        'total_ratio_base': total_ratio_base,
        'summary_total': summary_total,
        'expense_rows': expense_rows,
    }


def _load_shared_expense_context(expense_path, cost_center_map):
    wb = load_workbook(expense_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    reverse_map = _build_cost_center_reverse_map(cost_center_map)
    rows_by_company = {'耐数电子': [], '耐数信息': []}

    for row_idx in range(2, ws.max_row + 1):
        company_code = _format_code(ws.cell(row=row_idx, column=5).value)
        company = COMPANY_CODE_TO_NAME.get(company_code)
        if not company:
            continue

        gl_account = _format_code(ws.cell(row=row_idx, column=10).value)
        amount = _to_money(ws.cell(row=row_idx, column=12).value)
        cost_center = _format_code(ws.cell(row=row_idx, column=19).value)
        dept_text = _normalize_department_for_order(ws.cell(row=row_idx, column=20).value)
        credit_order = _format_code(ws.cell(row=row_idx, column=21).value)
        functional_area = _format_code(ws.cell(row=row_idx, column=59).value)
        if not gl_account or amount == Decimal('0.00') or not cost_center:
            continue
        if functional_area == '1000':
            continue

        allocation_cost_center = cost_center
        if company == '耐数电子' and cost_center == '20502050':
            allocation_cost_center = '20502060'

        mapped_dept = reverse_map.get(company, {}).get(cost_center, '')
        allocation_dept = reverse_map.get(company, {}).get(allocation_cost_center, '')
        dept = allocation_dept or mapped_dept or dept_text
        dept = _normalize_department_for_order(dept)
        if not dept:
            raise ValueError(f'待分摊费用 第 {row_idx} 行无法根据成本中心 {cost_center} 识别标准部门')

        rows_by_company[company].append(
            {
                'row_idx': row_idx,
                'gl_account': gl_account,
                'amount': amount,
                'cost_center': cost_center,
                'allocation_cost_center': allocation_cost_center,
                'dept': dept,
                'credit_order': credit_order,
            }
        )

    return {
        'rows_by_company': rows_by_company,
    }


def _build_bonus_department_amounts(company, payroll_year, payroll_month, bonus_context):
    posting_year, posting_month = _shift_month(payroll_year, payroll_month, 1)
    ratio = bonus_context['ratio_map'].get(posting_month)
    if ratio is None:
        raise ValueError(f'奖金计提比例表中未找到 {posting_month} 月系数')

    source_months = _bonus_months_for_period(posting_year, posting_month)
    rows = []
    for item in bonus_context['rows_by_company'].get(company, []):
        if not item['cost_center']:
            raise ValueError(f'{company} 奖金数据中部门 {item["dept"]} 未匹配到成本中心')
        base_amount = sum((item['monthly_amounts'].get(month, Decimal('0.00')) for month in source_months), Decimal('0.00'))
        amount = (base_amount * ratio).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        amount = BONUS_DEPARTMENT_ACCRUAL_OVERRIDES.get(
            (company, posting_year, posting_month, item['dept']),
            amount,
        )
        if amount == Decimal('0.00'):
            continue
        rows.append(
            {
                'dept': item['dept'],
                'cost_center': item['cost_center'],
                'base_amount': base_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
                'amount': amount,
            }
        )

    return {
        'posting_year': posting_year,
        'posting_month': posting_month,
        'ratio': ratio,
        'source_months': source_months,
        'rows': rows,
    }


def _load_mappings(mapping_path):
    """仅从 Mapping 读取成本中心；内部订单由工时数据.xlsx 匹配。"""
    wb = load_workbook(mapping_path, data_only=True)

    cost_center_ws = wb['成本中心']

    cost_center_map = {'耐数电子': {}, '耐数信息': {}}

    for row_idx in range(2, cost_center_ws.max_row + 1):
        dept_2050 = _normalize_text(cost_center_ws.cell(row=row_idx, column=2).value)
        code_2050 = _format_code(cost_center_ws.cell(row=row_idx, column=3).value)
        dept_2060 = _normalize_text(cost_center_ws.cell(row=row_idx, column=7).value)
        code_2060 = _format_code(cost_center_ws.cell(row=row_idx, column=8).value)

        if dept_2050 and code_2050 and dept_2050 not in cost_center_map['耐数电子']:
            cost_center_map['耐数电子'][dept_2050] = code_2050
        if dept_2060 and code_2060 and dept_2060 not in cost_center_map['耐数信息']:
            cost_center_map['耐数信息'][dept_2060] = code_2060

    return cost_center_map


def _build_cost_center_reverse_map(cost_center_map):
    reverse_map = {'耐数电子': {}, '耐数信息': {}}
    for company, dept_map in cost_center_map.items():
        for dept, cost_center in dept_map.items():
            dept_text = _normalize_department_for_order(dept)
            cc = _format_code(cost_center)
            if cc and dept_text and cc not in reverse_map[company]:
                reverse_map[company][cc] = dept_text
    return reverse_map


def _get_bonus_tax_adjustment(company_code, payment_year, payment_month):
    return BONUS_TAX_BY_PAYMENT_PERIOD.get(
        (_format_code(company_code), payment_year, payment_month),
        Decimal('0.00'),
    )


def _write_bank_summary(ws, bank_results, header_ref):
    start_col = 22  # V列
    headers = ['公司', '核对事项', '工资表金额', '银行流水金额', '差额', '核对月份', '结果', '说明']
    widths = [12, 12, 14, 14, 14, 12, 10, 42]

    for idx, header in enumerate(headers):
        cell = ws.cell(row=2, column=start_col + idx)
        cell.value = header
        _copy_cell_style(header_ref, cell)
        ws.column_dimensions[cell.column_letter].width = widths[idx]

    text_ref = ws['A3']
    num_ref = ws['Q3']
    for row_offset, result in enumerate(bank_results, start=3):
        values = [
            result['company'],
            result['item_label'],
            result['payroll_amount'],
            result['bank_amount'],
            result['diff'],
            result['period_label'],
            '通过' if result['passed'] else '异常',
            result['note'],
        ]

        for col_offset, value in enumerate(values):
            cell = ws.cell(row=row_offset, column=start_col + col_offset)
            cell.value = value
            if col_offset in (2, 3, 4):
                _copy_cell_style(num_ref, cell)
            else:
                _copy_cell_style(text_ref, cell)

        if not result['passed']:
            _mark_range_red(ws, row_offset, start_col, start_col + len(headers) - 1)


def _select_treasury_records(records):
    treasury_records = _get_treasury_records(records)
    if not treasury_records:
        return None, []
    return treasury_records[0], treasury_records[1:]


def _get_last_bank_date(records):
    dates = [record['trans_date'] for record in records if isinstance(record['trans_date'], (datetime, date))]
    if not dates:
        raise ValueError('未找到可用的银行交易日期')
    return max(dates)


def _sap_date_value(value):
    if not isinstance(value, (datetime, date)):
        raise ValueError(f'无法转换为凭证日期：{value}')
    return int(value.strftime('%Y%m%d'))


def _period_text(year, month):
    return f'{year}年{month}月'


def _bonus_months_for_period(year, month):
    if month in (3, 6, 9, 12):
        return list(range(1, month + 1))
    return [month]


def _bonus_label_for_filename(year, month):
    if month in (3, 6, 9, 12):
        if month == 3:
            return f'{year}Q1'
        return f'Q1-Q{month // 3}'
    return f'{year}{month:02d}'


def _bonus_label_for_text(year, month):
    return _bonus_label_for_filename(year, month)


def _extract_bonus_department(value):
    text = _normalize_text(value)
    if not text:
        return ''
    return _normalize_department_for_order(text.split('/')[-1])


def _bonus_company_alias(value):
    text = _normalize_text(value)
    mapping = {
        '北京普源耐数电子有限公司': '耐数电子',
        '北京耐数信息有限公司': '耐数信息',
    }
    return mapping.get(text, '')


def _month_end_date(year, month):
    next_year, next_month = _shift_month(year, month, 1)
    first_of_next = date(next_year, next_month, 1)
    return first_of_next.fromordinal(first_of_next.toordinal() - 1)


def _make_voucher_row(
    group,
    company_code,
    posting_date,
    gl_account,
    amount,
    text,
    posting_key,
    cost_center='',
    order='',
    voucher_type='ZZ',
):
    return {
        'A': group,
        'B': company_code,
        'C': voucher_type,
        'D': _sap_date_value(posting_date),
        'E': _sap_date_value(posting_date),
        'F': 'CNY',
        'G': posting_key,
        'H': '',
        'I': gl_account,
        'J': cost_center,
        'K': order,
        'L': '',
        'M': _to_text_money(amount),
        'N': text,
        'O': '',
        'P': '',
        'Q': '',
    }


def _rebalance_expense_debit_rows(rows, preferred_accounts, voucher_label):
    debit_total = Decimal('0.00')
    credit_total = Decimal('0.00')
    for row in rows:
        amount = _to_money(row.get('M'))
        if str(row.get('G')) == '40':
            debit_total += amount
        elif str(row.get('G')) == '50':
            credit_total += amount

    adjustment = (credit_total - debit_total).quantize(Decimal('0.01'))
    if adjustment == Decimal('0.00'):
        return rows

    preferred_set = set(preferred_accounts)
    for row in reversed(rows):
        if str(row.get('G')) != '40' or row.get('I') not in preferred_set:
            continue
        current_amount = _to_money(row.get('M'))
        new_amount = (current_amount + adjustment).quantize(Decimal('0.01'))
        if new_amount < Decimal('0.00'):
            continue
        row['M'] = _to_text_money(new_amount)
        return rows

    raise ValueError(f'{voucher_label} 存在 {adjustment} 尾差，但未找到可补差的费用借方行')


def _build_a1_rows(company, company_code, ws, total_row_idx, bank_records, payroll_year, payroll_month):
    next_year, next_month = _shift_month(payroll_year, payroll_month, 1)
    month_records = bank_records.get((company_code, next_year, next_month), [])
    social_target = sum(
        (_to_money(ws.cell(total_row_idx, col_idx).value) for col_idx, _ in SOCIAL_DEBIT_ACCOUNT_COLUMNS),
        Decimal('0.00'),
    )
    social_match = _match_treasury_combination(month_records, social_target, '社保国库流水')
    social_records = social_match['matched_records']
    if not social_match['matched'] or not social_records:
        raise ValueError(f'{company} 未找到匹配社保金额的银行流水组合：{social_match["note"]}')

    posting_date = _get_last_bank_date(social_records)
    text = f'支付{_period_text(payroll_year, payroll_month)}工作期间社保'
    rows = []
    period_text = _period_text(payroll_year, payroll_month)

    for col_idx, gl_account in SOCIAL_DEBIT_ACCOUNT_COLUMNS:
        amount = _to_money(ws.cell(total_row_idx, col_idx).value)
        if amount == Decimal('0.00'):
            continue
        debit_text = SOCIAL_TEXT_BY_ACCOUNT[gl_account].format(period=period_text)
        rows.append(
            _make_voucher_row(
                'A1',
                company_code,
                posting_date,
                gl_account,
                amount,
                debit_text,
                40,
                voucher_type=BANK_VOUCHER_TYPE,
            )
        )

    for record in social_records:
        credit_row = _make_voucher_row(
            'A1',
            company_code,
            posting_date,
            record['bank_subject'],
            record['outgoing_amt'],
            text,
            50,
            voucher_type=BANK_VOUCHER_TYPE,
        )
        credit_row['P'] = BANK_REASON_CODE
        rows.append(
            credit_row
        )

    return rows


def _build_a2_rows(company, company_code, bank_records, payroll_year, payroll_month):
    next_year, next_month = _shift_month(payroll_year, payroll_month, 1)
    month_records = bank_records.get((company_code, next_year, next_month), [])
    fund_records = [record for record in month_records if '北京住房公积金管理中心' in record['payment_name']]
    if not fund_records:
        raise ValueError(f'{company} 未找到公积金对应的银行流水')

    posting_date = _get_last_bank_date(fund_records)
    text = f'支付{_period_text(next_year, next_month)}工作期间公积金'
    total_amount = sum((record['outgoing_amt'] for record in fund_records), Decimal('0.00'))
    personal_amount = (total_amount / Decimal('2')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    company_amount = (total_amount - personal_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    period_text = _period_text(next_year, next_month)

    rows = [
        _make_voucher_row(
            'A2',
            company_code,
            posting_date,
            '2211070001',
            personal_amount,
            f'支付{period_text}工作期间个人承担公积金',
            40,
            voucher_type=BANK_VOUCHER_TYPE,
        ),
        _make_voucher_row(
            'A2',
            company_code,
            posting_date,
            '2211070002',
            company_amount,
            f'支付{period_text}工作期间公司承担公积金',
            40,
            voucher_type=BANK_VOUCHER_TYPE,
        ),
    ]

    for record in fund_records:
        credit_row = _make_voucher_row(
            'A2',
            company_code,
            posting_date,
            record['bank_subject'],
            record['outgoing_amt'],
            text,
            50,
            voucher_type=BANK_VOUCHER_TYPE,
        )
        credit_row['P'] = BANK_REASON_CODE
        rows.append(
            credit_row
        )

    return rows


def _build_a3_rows(company, company_code, ws, bank_records, payroll_year, payroll_month, timesheet_context):
    next_year, next_month = _shift_month(payroll_year, payroll_month, 1)
    month_records = bank_records.get((company_code, next_year, next_month), [])
    salary_records = [record for record in month_records if '代发工资' in record['usage']]

    labor_rows = []
    severance_rows = []
    salary_total = Decimal('0')
    tax_total = Decimal('0')
    labor_tax_total = Decimal('0')
    bank_tax_target = Decimal('0')
    total_row_idx = _find_total_row(ws)
    if total_row_idx is None:
        raise ValueError(f'{company} 工资单中未找到总计行')
    salary_match_target = _to_money(ws.cell(total_row_idx, 17).value)
    data_start_row = _find_payroll_header_row(ws) + 1
    for row_idx in range(data_start_row, ws.max_row + 1):
        if _is_total_row(ws.cell(row_idx, 1).value):
            break
        item_type = _normalize_text(ws.cell(row_idx, 4).value)
        if item_type == '劳务费':
            labor_tax_total += _to_decimal(ws.cell(row_idx, 15).value)
            bank_tax_target += _to_decimal(ws.cell(row_idx, 15).value)
            labor_rows.append(
                {
                    'amount': _to_money(ws.cell(row_idx, 5).value),
                    'cost_center': _format_code(ws.cell(row_idx, 19).value),
                    'order': _format_code(ws.cell(row_idx, 20).value),
                }
            )
        elif item_type == '离职补偿金':
            dept = _normalize_department_for_order(ws.cell(row_idx, 2).value)
            order, daily_project = _lookup_internal_order_from_timesheet(
                timesheet_context,
                company,
                dept,
                '',
            )
            if not order:
                raise ValueError(
                    f'{company} 离职补偿金第 {row_idx} 行未找到部门日常内部订单：'
                    f'部门“{dept}”，匹配项目“{daily_project}”'
                )
            severance_rows.append(
                {
                    'amount': _to_money(ws.cell(row_idx, 17).value),
                    'cost_center': _format_code(ws.cell(row_idx, 19).value),
                    'order': order,
                }
            )
        elif item_type == '薪资':
            salary_total += _to_decimal(ws.cell(row_idx, 17).value)
            tax_total += _to_decimal(ws.cell(row_idx, 15).value)
            bank_tax_target += _to_decimal(ws.cell(row_idx, 15).value)

    salary_total = salary_total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    tax_total = tax_total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    labor_tax_total = labor_tax_total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    bank_tax_target = bank_tax_target.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    bonus_tax_adjustment = _get_bonus_tax_adjustment(company_code, next_year, next_month)
    bank_tax_target = (bank_tax_target + bonus_tax_adjustment).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    salary_match = _match_salary_combination(salary_records, salary_match_target)
    if not salary_match['matched'] or not salary_match['matched_records']:
        raise ValueError(f'{company} A3 未找到匹配工资金额的银行流水组合')
    tax_match = _match_treasury_combination(month_records, bank_tax_target, '个税国库流水')
    tax_records = tax_match['matched_records']
    if not tax_match['matched'] or not tax_records:
        raise ValueError(f'{company} A3 未找到匹配个税金额的银行流水组合：{tax_match["note"]}')

    credit_records = salary_match['matched_records'] + tax_records
    posting_date = _get_last_bank_date(credit_records)
    text = f'支付{_period_text(payroll_year, payroll_month)}工作期间工资及劳务费'
    period_text = _period_text(payroll_year, payroll_month)

    rows = []
    for labor_row in labor_rows:
        rows.append(
            _make_voucher_row(
                'A3',
                company_code,
                posting_date,
                '6601110001',
                labor_row['amount'],
                f'支付{period_text}工作期间劳务费',
                40,
                labor_row['cost_center'],
                labor_row['order'],
                voucher_type=BANK_VOUCHER_TYPE,
            )
        )

    for severance_row in severance_rows:
        rows.append(
            _make_voucher_row(
                'A3',
                company_code,
                posting_date,
                '6601990001',
                severance_row['amount'],
                '离职补偿金',
                40,
                severance_row['cost_center'],
                severance_row['order'],
                voucher_type=BANK_VOUCHER_TYPE,
            )
        )

    tax_debit_amount = (tax_match['matched_amount'] - labor_tax_total).quantize(
        Decimal('0.01'),
        rounding=ROUND_HALF_UP,
    )
    if tax_debit_amount < Decimal('0.00'):
        raise ValueError(f'{company} A3 个税银行流水小于劳务费个税，无法生成 2221070000 借方')
    if tax_debit_amount != Decimal('0.00'):
        rows.append(
            _make_voucher_row(
                'A3',
                company_code,
                posting_date,
                '2221070000',
                tax_debit_amount,
                f'支付{period_text}工作期间个人所得税',
                40,
                voucher_type=BANK_VOUCHER_TYPE,
            )
        )
    if salary_total != Decimal('0.00'):
        rows.append(
            _make_voucher_row(
                'A3',
                company_code,
                posting_date,
                '2211010001',
                salary_total,
                f'支付{period_text}工作期间薪酬',
                40,
                voucher_type=BANK_VOUCHER_TYPE,
            )
        )

    for record in salary_match['matched_records']:
        credit_row = _make_voucher_row(
            'A3',
            company_code,
            posting_date,
            record['bank_subject'],
            record['outgoing_amt'],
            text,
            50,
            voucher_type=BANK_VOUCHER_TYPE,
        )
        credit_row['P'] = BANK_REASON_CODE
        rows.append(
            credit_row
        )

    for tax_record in tax_records:
        tax_credit_row = _make_voucher_row(
            'A3',
            company_code,
            posting_date,
            tax_record['bank_subject'],
            tax_record['outgoing_amt'],
            f'支付{period_text}工作期间个人所得税',
            50,
            voucher_type=BANK_VOUCHER_TYPE,
        )
        tax_credit_row['P'] = BANK_REASON_CODE
        rows.append(tax_credit_row)

    return _rebalance_expense_debit_rows(rows, ('6601110001', '6601990001'), f'{company} A3')


def _amount_to_ten_thousand_yuan(amount):
    return (amount / Decimal('10000')).quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)


def _describe_bank_records(records):
    if not records:
        return ''
    return '；'.join(
        f'{record["file"]} 第{record["row_idx"]}行 {record["outgoing_amt"]}'
        for record in records
    )


def _append_key_tax_source_row(rows, company, company_code, processing_label, payroll_label, category, item, amount, source, note):
    amount = _to_money(amount)
    rows.append(
        {
            '处理月份': processing_label,
            '工资所属月份': payroll_label,
            '公司代码': company_code,
            '公司': company,
            '类别': category,
            '项目': item,
            '金额（元）': amount,
            '金额（万元）': _amount_to_ten_thousand_yuan(amount),
            '取数口径': source,
            '备注': note,
        }
    )


def _build_company_key_tax_source_rows(company, payroll_path, bank_records, payroll_year, payroll_month):
    company_code = COMPANY_NAME_TO_CODE[company]
    processing_year, processing_month = _shift_month(payroll_year, payroll_month, 1)
    processing_label = f'{processing_year}年{processing_month}月'
    payroll_label = f'{payroll_year}年{payroll_month}月'

    wb = load_workbook(payroll_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    total_row_idx = _find_total_row(ws)
    if total_row_idx is None:
        raise ValueError(f'{company} 工资整理文件中未找到总计行，无法生成重点税源采集信息')

    company_pension = _to_money(ws.cell(total_row_idx, 6).value)
    company_unemployment = _to_money(ws.cell(total_row_idx, 7).value)
    company_injury = _to_money(ws.cell(total_row_idx, 8).value)
    company_medical = _to_money(ws.cell(total_row_idx, 9).value)
    social_total = sum(
        (_to_money(ws.cell(total_row_idx, col_idx).value) for col_idx in (6, 7, 8, 9, 11, 12, 13)),
        Decimal('0.00'),
    )

    salary_tax = Decimal('0.00')
    labor_tax = Decimal('0.00')
    data_start_row = _find_payroll_header_row(ws) + 1
    for row_idx in range(data_start_row, ws.max_row + 1):
        if _is_total_row(ws.cell(row_idx, 1).value):
            break
        item_type = _normalize_text(ws.cell(row_idx, 4).value)
        if item_type == '薪资':
            salary_tax += _to_decimal(ws.cell(row_idx, 15).value)
        elif item_type == '劳务费':
            labor_tax += _to_decimal(ws.cell(row_idx, 15).value)
    salary_tax = salary_tax.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    labor_tax = labor_tax.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    bonus_tax = _get_bonus_tax_adjustment(company_code, processing_year, processing_month)
    salary_tax_with_bonus = (salary_tax + bonus_tax).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    tax_total = (salary_tax_with_bonus + labor_tax).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    month_records = bank_records.get((company_code, processing_year, processing_month), [])
    social_match = _match_treasury_combination(month_records, social_total, '社保国库流水')
    social_note = social_match['note']
    if social_match['matched_records']:
        social_note = f'{social_note}；来源：{_describe_bank_records(social_match["matched_records"])}'

    fund_records = [record for record in month_records if '北京住房公积金管理中心' in record['payment_name']]
    fund_total = sum((record['outgoing_amt'] for record in fund_records), Decimal('0.00'))
    company_fund = (fund_total / Decimal('2')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    fund_note = f'银行流水公积金实付合计 {fund_total}，按现有凭证口径公司承担 50%'
    if fund_records:
        fund_note = f'{fund_note}；来源：{_describe_bank_records(fund_records)}'

    tax_match = _match_treasury_combination(month_records, tax_total, '个税国库流水')
    tax_note = tax_match['note']
    if tax_match['matched_records']:
        tax_note = f'{tax_note}；来源：{_describe_bank_records(tax_match["matched_records"])}'

    rows = []
    _append_key_tax_source_row(
        rows,
        company,
        company_code,
        processing_label,
        payroll_label,
        '社保',
        '养老保险（公司承担）',
        company_pension,
        '工资整理文件总计行 F列，公司承担养老；社保国库流水用于实缴匹配',
        social_note,
    )
    _append_key_tax_source_row(
        rows,
        company,
        company_code,
        processing_label,
        payroll_label,
        '社保',
        '医疗保险（公司承担）',
        company_medical,
        '工资整理文件总计行 I列，公司承担医疗；社保国库流水用于实缴匹配',
        social_note,
    )
    _append_key_tax_source_row(
        rows,
        company,
        company_code,
        processing_label,
        payroll_label,
        '社保',
        '失业保险（公司承担）',
        company_unemployment,
        '工资整理文件总计行 G列，公司承担失业；社保国库流水用于实缴匹配',
        social_note,
    )
    _append_key_tax_source_row(
        rows,
        company,
        company_code,
        processing_label,
        payroll_label,
        '社保',
        '工伤保险（公司承担）',
        company_injury,
        '工资整理文件总计行 H列，公司承担工伤；社保国库流水用于实缴匹配',
        social_note,
    )
    _append_key_tax_source_row(
        rows,
        company,
        company_code,
        processing_label,
        payroll_label,
        '公积金',
        '住房公积金（公司承担）',
        company_fund,
        '银行流水付款名称含“北京住房公积金管理中心”的实付金额，按公司承担 50% 拆分',
        fund_note,
    )
    _append_key_tax_source_row(
        rows,
        company,
        company_code,
        processing_label,
        payroll_label,
        '个人所得税',
        '个人所得税-工资薪金',
        salary_tax_with_bonus,
        '工资整理文件薪资行 O列个税；若存在本月已确认年终奖个税调整，则并入工资薪金口径',
        f'薪资个税 {salary_tax}，年终奖个税调整 {bonus_tax}；{tax_note}',
    )
    _append_key_tax_source_row(
        rows,
        company,
        company_code,
        processing_label,
        payroll_label,
        '个人所得税',
        '个人所得税-劳务',
        labor_tax,
        '工资整理文件劳务费行 O列个税',
        tax_note,
    )
    _append_key_tax_source_row(
        rows,
        company,
        company_code,
        processing_label,
        payroll_label,
        '个人所得税',
        '个人所得税合计',
        tax_total,
        '工资薪金个税与劳务个税合计，并与个税国库流水匹配',
        tax_note,
    )
    return rows


def _write_key_tax_source_workbook(report_path, rows, processing_year, processing_month):
    wb = Workbook()
    ws = wb.active
    ws.title = '重点税源采集信息'
    headers = ['处理月份', '工资所属月份', '公司代码', '公司', '类别', '项目', '金额（元）', '金额（万元）', '取数口径', '备注']
    ws.append(headers)
    for row in rows:
        ws.append([row[header] for header in headers])

    widths = [14, 14, 10, 12, 12, 24, 14, 14, 46, 70]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width
    ws.freeze_panes = 'A2'
    for cell in ws[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, 7).number_format = '#,##0.00'
        ws.cell(row_idx, 8).number_format = '#,##0.0000'

    note_ws = wb.create_sheet('口径说明')
    note_ws.append(['项目', '说明'])
    note_ws.append(['输出位置', report_path])
    note_ws.append(['统计月份', f'{processing_year}年{processing_month}月银行流水'])
    note_ws.append(['社保公司承担', '养老、医疗、失业、工伤取工资整理文件公司承担列，并用对应社保国库流水匹配实缴总额。'])
    note_ws.append(['公积金公司承担', '取北京住房公积金管理中心银行实付金额的 50%，沿用现有 A2 凭证拆分口径。'])
    note_ws.append(['个人所得税', '工资薪金个税包含薪资行个税及本月已确认年终奖个税调整；劳务个税单独列示。'])
    note_ws.column_dimensions['A'].width = 18
    note_ws.column_dimensions['B'].width = 100
    wb.save(report_path)


def _write_key_tax_source_report(base_dir, output_paths, bank_records, payroll_year, payroll_month):
    processing_year, processing_month = _shift_month(payroll_year, payroll_month, 1)
    report_paths = {}
    for company, payroll_path in output_paths.items():
        rows = _build_company_key_tax_source_rows(company, payroll_path, bank_records, payroll_year, payroll_month)
        report_path = _build_key_tax_source_output_path(base_dir, company, processing_year, processing_month)
        _write_key_tax_source_workbook(report_path, rows, processing_year, processing_month)
        report_paths[company] = report_path
    return report_paths


def _build_a4_rows(company, company_code, ws, total_row_idx, payroll_year, payroll_month, timesheet_context):
    posting_year, posting_month = _shift_month(payroll_year, payroll_month, 1)
    posting_date = _month_end_date(posting_year, posting_month)
    period_text = _period_text(payroll_year, payroll_month)

    grouped_amounts = {}
    debit_totals = {}
    salary_total = Decimal('0')
    tax_total = Decimal('0')
    other_total = Decimal('0')

    data_start_row = _find_payroll_header_row(ws) + 1
    for row_idx in range(data_start_row, ws.max_row + 1):
        if _is_total_row(ws.cell(row_idx, 1).value):
            break

        item_type = _normalize_text(ws.cell(row_idx, 4).value)
        if item_type in ('劳务费', '离职补偿金'):
            continue

        dept = _normalize_text(ws.cell(row_idx, 2).value)
        project = _normalize_text(ws.cell(row_idx, 3).value)
        cost_center = _format_code(ws.cell(row_idx, 19).value)
        order = _resolve_a4_internal_order(
            timesheet_context,
            company,
            dept,
            project,
            _format_code(ws.cell(row_idx, 20).value),
        )
        key = (cost_center, order)
        bucket = grouped_amounts.setdefault(key, {})

        for col_idx, account, label in ACCRUAL_DEBIT_SPECS:
            amount = _to_decimal(ws.cell(row_idx, col_idx).value)
            if amount == Decimal('0'):
                continue
            bucket[account] = bucket.get(account, Decimal('0')) + amount
            debit_totals[account] = debit_totals.get(account, Decimal('0')) + amount

        if item_type == '薪资':
            salary_total += _to_decimal(ws.cell(row_idx, 17).value)
            tax_total += _to_decimal(ws.cell(row_idx, 15).value)
            other_total += _to_decimal(ws.cell(row_idx, 16).value)

    salary_total = salary_total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    tax_total = tax_total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    other_total = other_total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    rows = []
    sorted_groups = sorted(grouped_amounts.items())
    rounded_grouped_amounts = {}
    account_last_keys = {}
    rounded_account_totals = {account: Decimal('0.00') for _, account, _ in ACCRUAL_DEBIT_SPECS}

    for key, account_map in sorted_groups:
        rounded_account_map = {}
        for _, account, _ in ACCRUAL_DEBIT_SPECS:
            raw_amount = account_map.get(account, Decimal('0'))
            rounded_amount = raw_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            rounded_account_map[account] = rounded_amount
            rounded_account_totals[account] += rounded_amount
            if raw_amount != Decimal('0'):
                account_last_keys[account] = key
        rounded_grouped_amounts[key] = rounded_account_map

    for _, account, _ in ACCRUAL_DEBIT_SPECS:
        target_total = debit_totals.get(account, Decimal('0')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        diff = (target_total - rounded_account_totals.get(account, Decimal('0.00'))).quantize(Decimal('0.01'))
        if diff != Decimal('0.00') and account in account_last_keys:
            rounded_grouped_amounts[account_last_keys[account]][account] += diff

    for (cost_center, order), _ in sorted_groups:
        account_map = rounded_grouped_amounts[(cost_center, order)]
        for _, account, label in ACCRUAL_DEBIT_SPECS:
            amount = account_map.get(account, Decimal('0.00'))
            if amount == Decimal('0.00'):
                continue
            rows.append(
                _make_voucher_row(
                    'A4',
                    company_code,
                    posting_date,
                    account,
                    amount,
                    f'实际入账{period_text}工作期间{label}',
                    40,
                    cost_center,
                    order,
                )
            )

    for col_idx, account, label in ACCRUAL_CREDIT_SPECS:
        amount = _to_money(ws.cell(total_row_idx, col_idx).value)
        if amount == Decimal('0.00'):
            continue
        rows.append(
            _make_voucher_row(
                'A4',
                company_code,
                posting_date,
                account,
                amount,
                f'实际入账{period_text}工作期间{label}',
                50,
            )
        )

    if tax_total != Decimal('0.00'):
        rows.append(
            _make_voucher_row(
                'A4',
                company_code,
                posting_date,
                '2221070000',
                tax_total,
                f'实际入账{period_text}工作期间应交个税',
                50,
            )
        )
    if salary_total != Decimal('0.00'):
        rows.append(
            _make_voucher_row(
                'A4',
                company_code,
                posting_date,
                '2211010001',
                salary_total,
                f'实际入账{period_text}工作期间实发工资',
                50,
            )
        )
    if other_total != Decimal('0.00'):
        rows.append(
            _make_voucher_row(
                'A4',
                company_code,
                posting_date,
                '2211010001',
                other_total,
                f'实际入账{period_text}工作期间其他项目',
                50,
            )
        )

    expense_accounts = tuple(account for _, account, _ in ACCRUAL_DEBIT_SPECS)
    return _rebalance_expense_debit_rows(rows, expense_accounts, f'{company} A4')


def _should_skip_one_time_june_2026_accrual(company, posting_year, posting_month, dept):
    # The only 2060/WLS employee left in June 2026. This department therefore
    # does not accrue or allocate payroll costs in June or July 2026.
    return (
        company == '耐数信息'
        and posting_year == 2026
        and posting_month in (6, 7)
        and dept == 'WLS无线解决方案'
    )


def _collect_a5_base_data(ws, company, posting_year, posting_month):
    grouped_amounts = {}
    debit_totals = {}
    dept_by_cost_center = {}
    salary_total = Decimal('0')
    tax_total = Decimal('0')
    other_total = Decimal('0')

    data_start_row = _find_payroll_header_row(ws) + 1
    for row_idx in range(data_start_row, ws.max_row + 1):
        if _is_total_row(ws.cell(row_idx, 1).value):
            break

        item_type = _normalize_text(ws.cell(row_idx, 4).value)
        if item_type in ('劳务费', '离职补偿金'):
            continue

        dept = _normalize_department_for_order(ws.cell(row_idx, 2).value)
        if _should_skip_one_time_june_2026_accrual(company, posting_year, posting_month, dept):
            continue
        cost_center = _format_code(ws.cell(row_idx, 19).value)
        if cost_center and dept and cost_center not in dept_by_cost_center:
            dept_by_cost_center[cost_center] = dept

        bucket = grouped_amounts.setdefault(cost_center, {})
        for col_idx, account, _ in ACCRUAL_DEBIT_SPECS:
            amount = _to_decimal(ws.cell(row_idx, col_idx).value)
            if amount == Decimal('0'):
                continue
            bucket[account] = bucket.get(account, Decimal('0')) + amount
            debit_totals[account] = debit_totals.get(account, Decimal('0')) + amount

        if item_type == '薪资':
            salary_total += _to_decimal(ws.cell(row_idx, 17).value)
            tax_total += _to_decimal(ws.cell(row_idx, 15).value)
            other_total += _to_decimal(ws.cell(row_idx, 16).value)

    return {
        'grouped_amounts': grouped_amounts,
        'debit_totals': debit_totals,
        'dept_by_cost_center': dept_by_cost_center,
        'salary_total': salary_total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
        'tax_total': tax_total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
        'other_total': other_total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
    }


def _build_a5_rows(company, company_code, ws, total_row_idx, payroll_year, payroll_month):
    posting_year, posting_month = _shift_month(payroll_year, payroll_month, 1)
    accrual_data = _collect_a5_base_data(ws, company, posting_year, posting_month)
    posting_date = _month_end_date(posting_year, posting_month)
    period_text = _period_text(posting_year, posting_month)

    rows = []
    sorted_groups = sorted(accrual_data['grouped_amounts'].items())
    rounded_grouped_amounts = {}
    account_last_keys = {}
    rounded_account_totals = {account: Decimal('0.00') for _, account, _ in ACCRUAL_DEBIT_SPECS}

    for key, account_map in sorted_groups:
        rounded_account_map = {}
        for _, account, _ in ACCRUAL_DEBIT_SPECS:
            raw_amount = account_map.get(account, Decimal('0'))
            rounded_amount = raw_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            rounded_account_map[account] = rounded_amount
            rounded_account_totals[account] += rounded_amount
            if raw_amount != Decimal('0'):
                account_last_keys[account] = key
        rounded_grouped_amounts[key] = rounded_account_map

    for _, account, _ in ACCRUAL_DEBIT_SPECS:
        target_total = accrual_data['debit_totals'].get(account, Decimal('0')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        diff = (target_total - rounded_account_totals.get(account, Decimal('0.00'))).quantize(Decimal('0.01'))
        if diff != Decimal('0.00') and account in account_last_keys:
            rounded_grouped_amounts[account_last_keys[account]][account] += diff

    for cost_center, _ in sorted_groups:
        account_map = rounded_grouped_amounts[cost_center]
        for _, account, label in ACCRUAL_DEBIT_SPECS:
            amount = account_map.get(account, Decimal('0.00'))
            if amount == Decimal('0.00'):
                continue
            rows.append(
                _make_voucher_row(
                    'A5',
                    company_code,
                    posting_date,
                    account,
                    amount,
                    f'计提{period_text}工作期间{label}',
                    40,
                    cost_center,
                )
            )

    for col_idx, account, label in ACCRUAL_CREDIT_SPECS:
        amount = _to_money(ws.cell(total_row_idx, col_idx).value)
        if amount == Decimal('0.00'):
            continue
        rows.append(
            _make_voucher_row(
                'A5',
                company_code,
                posting_date,
                account,
                amount,
                f'计提{period_text}工作期间{label}',
                50,
            )
        )

    if accrual_data['tax_total'] != Decimal('0.00'):
        rows.append(
            _make_voucher_row(
                'A5',
                company_code,
                posting_date,
                '2221070000',
                accrual_data['tax_total'],
                f'计提{period_text}工作期间应交个税',
                50,
            )
        )
    if accrual_data['salary_total'] != Decimal('0.00'):
        rows.append(
            _make_voucher_row(
                'A5',
                company_code,
                posting_date,
                '2211010001',
                accrual_data['salary_total'],
                f'计提{period_text}工作期间实发工资',
                50,
            )
        )
    if accrual_data['other_total'] != Decimal('0.00'):
        rows.append(
            _make_voucher_row(
                'A5',
                company_code,
                posting_date,
                '2211010001',
                accrual_data['other_total'],
                f'计提{period_text}工作期间其他项目',
                50,
            )
        )

    expense_accounts = tuple(account for _, account, _ in ACCRUAL_DEBIT_SPECS)
    return _rebalance_expense_debit_rows(rows, expense_accounts, f'{company} A5')


def _build_a6_rows(company, company_code, ws, payroll_year, payroll_month, timesheet_context):
    posting_year, posting_month = _shift_month(payroll_year, payroll_month, 1)
    accrual_data = _collect_a5_base_data(ws, company, posting_year, posting_month)
    posting_date = _month_end_date(posting_year, posting_month)
    text = f'根据研发工时分摊计提的{posting_year}年{posting_month}月人工费用'

    rows = []
    alloc_rows_by_key = {}

    for cost_center, account_map in sorted(accrual_data['grouped_amounts'].items()):
        dept = accrual_data['dept_by_cost_center'].get(cost_center, '')
        if not dept or dept == 'OSD运营支持部':
            continue

        alloc_bucket = timesheet_context['allocation_index'].get((posting_year, posting_month, company, dept))
        if not alloc_bucket:
            raise ValueError(f'{company} {dept} 在 {posting_year}年{posting_month}月 未找到 K>0 的工时项目，无法生成 A6')

        order_hours = sorted(alloc_bucket.items())
        total_hours = sum((hours for _, hours in order_hours), Decimal('0'))
        if total_hours <= Decimal('0'):
            raise ValueError(f'{company} {dept} 在 {posting_year}年{posting_month}月 工时合计为 0，无法生成 A6')

        for _, account, _ in ACCRUAL_DEBIT_SPECS:
            amount = account_map.get(account, Decimal('0')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            if amount == Decimal('0.00'):
                continue

            rows.append(
                _make_voucher_row(
                    'A6',
                    company_code,
                    posting_date,
                    account,
                    amount,
                    text,
                    50,
                    cost_center,
                )
            )

            rounded_total = Decimal('0.00')
            last_key = None
            for order, hours in order_hours:
                alloc_amount = (amount * hours / total_hours).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                key = (cost_center, order, account)
                alloc_rows_by_key[key] = alloc_rows_by_key.get(key, Decimal('0.00')) + alloc_amount
                rounded_total += alloc_amount
                last_key = key

            diff = (amount - rounded_total).quantize(Decimal('0.01'))
            if diff != Decimal('0.00') and last_key is not None:
                alloc_rows_by_key[last_key] += diff

    for (cost_center, order, account), amount in sorted(alloc_rows_by_key.items()):
        amount = amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if amount == Decimal('0.00'):
            continue
        rows.append(
            _make_voucher_row(
                'A6',
                company_code,
                posting_date,
                account,
                amount,
                text,
                40,
                cost_center,
                order,
            )
        )

    return rows


def _bonus_allocation_months(year, month):
    if month in (3, 6, 9, 12):
        return list(range(1, month + 1))
    return [month]


def _get_bonus_allocation_bucket(timesheet_context, year, month, company, dept):
    bucket = {}
    for alloc_month in _bonus_allocation_months(year, month):
        month_bucket = timesheet_context['allocation_index'].get((year, alloc_month, company, dept), {})
        for order, hours in month_bucket.items():
            bucket[order] = bucket.get(order, Decimal('0')) + hours
    return bucket


def _build_a7_rows(company, company_code, payroll_year, payroll_month, bonus_context):
    bonus_data = _build_bonus_department_amounts(company, payroll_year, payroll_month, bonus_context)
    posting_date = _month_end_date(bonus_data['posting_year'], bonus_data['posting_month'])
    label = _bonus_label_for_text(bonus_data['posting_year'], bonus_data['posting_month'])
    text = f'计提{label}年终奖'

    rows = []
    total_amount = Decimal('0.00')
    for item in bonus_data['rows']:
        rows.append(
            _make_voucher_row(
                'A7',
                company_code,
                posting_date,
                '6601010003',
                item['amount'],
                text,
                40,
                item['cost_center'],
            )
        )
        total_amount += item['amount']

    if total_amount != Decimal('0.00'):
        rows.append(
            _make_voucher_row(
                'A7',
                company_code,
                posting_date,
                '2211010001',
                total_amount,
                text,
                50,
            )
        )

    return rows


def _build_a8_rows(company, company_code, payroll_year, payroll_month, bonus_context, timesheet_context):
    bonus_data = _build_bonus_department_amounts(company, payroll_year, payroll_month, bonus_context)
    posting_date = _month_end_date(bonus_data['posting_year'], bonus_data['posting_month'])
    label = _bonus_label_for_text(bonus_data['posting_year'], bonus_data['posting_month'])
    text = f'根据研发工时分摊计提的{label}年终奖'

    rows = []
    alloc_rows_by_key = {}

    for item in bonus_data['rows']:
        dept = item['dept']
        if dept == 'OSD运营支持部':
            continue

        amount = item['amount']
        cost_center = item['cost_center']
        alloc_bucket = _get_bonus_allocation_bucket(
            timesheet_context,
            bonus_data['posting_year'],
            bonus_data['posting_month'],
            company,
            dept,
        )
        if not alloc_bucket:
            raise ValueError(f'{company} {dept} 在奖金分摊口径下未找到可用工时，无法生成 A8')

        order_hours = sorted((order, hours) for order, hours in alloc_bucket.items() if hours > Decimal('0'))
        total_hours = sum((hours for _, hours in order_hours), Decimal('0'))
        if total_hours <= Decimal('0'):
            raise ValueError(f'{company} {dept} 在奖金分摊口径下工时合计为 0，无法生成 A8')

        rows.append(
            _make_voucher_row(
                'A8',
                company_code,
                posting_date,
                '6601010003',
                amount,
                text,
                50,
                cost_center,
            )
        )

        rounded_total = Decimal('0.00')
        last_key = None
        for order, hours in order_hours:
            alloc_amount = (amount * hours / total_hours).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            key = (cost_center, order)
            alloc_rows_by_key[key] = alloc_rows_by_key.get(key, Decimal('0.00')) + alloc_amount
            rounded_total += alloc_amount
            last_key = key

        diff = (amount - rounded_total).quantize(Decimal('0.01'))
        if diff != Decimal('0.00') and last_key is not None:
            alloc_rows_by_key[last_key] += diff

    for (cost_center, order), amount in sorted(alloc_rows_by_key.items()):
        amount = amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if amount == Decimal('0.00'):
            continue
        rows.append(
            _make_voucher_row(
                'A8',
                company_code,
                posting_date,
                '6601010003',
                amount,
                text,
                40,
                cost_center,
                order,
            )
        )

    return rows


def _build_a9_rows(company_code, payroll_year, payroll_month, co_context):
    posting_year, posting_month = _shift_month(payroll_year, payroll_month, 1)
    posting_date = _month_end_date(posting_year, posting_month)
    text = f'CO工单分摊{posting_year}年{posting_month}月'

    rows = []
    debit_allocations = {}
    allocation_rows = co_context['allocation_rows']
    total_ratio_base = co_context['total_ratio_base']

    for item in co_context['expense_rows']:
        credit_row = _make_voucher_row(
            'A9',
            company_code,
            posting_date,
            item['source_account'],
            item['amount'],
            text,
            50,
            CO_CREDIT_COST_CENTER,
            CO_CREDIT_ORDER,
        )
        credit_row['O'] = 'x'
        rows.append(
            credit_row
        )

        rounded_total = Decimal('0.00')
        last_key = None
        for alloc in allocation_rows:
            alloc_amount = (item['amount'] * alloc['ratio_base'] / total_ratio_base).quantize(
                Decimal('0.01'),
                rounding=ROUND_HALF_UP,
            )
            key = (item['debit_gl'], alloc['order'])
            debit_allocations[key] = debit_allocations.get(key, Decimal('0.00')) + alloc_amount
            rounded_total += alloc_amount
            last_key = key

        diff = (item['amount'] - rounded_total).quantize(Decimal('0.01'))
        if diff != Decimal('0.00') and last_key is not None:
            debit_allocations[last_key] += diff

    for (debit_gl, order), amount in sorted(debit_allocations.items()):
        amount = amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if amount == Decimal('0.00'):
            continue
        rows.append(
            _make_voucher_row(
                'A9',
                company_code,
                posting_date,
                debit_gl,
                amount,
                text,
                40,
                '',
                order,
            )
        )

    return rows


def _build_a10_rows(company, company_code, payroll_year, payroll_month, expense_context, timesheet_context):
    posting_year, posting_month = _shift_month(payroll_year, payroll_month, 1)
    posting_date = _month_end_date(posting_year, posting_month)
    text = f'根据研发工时分摊{posting_year}年{posting_month}月研发费用'

    rows = []
    alloc_rows_by_key = {}
    grouped_items = {}

    for item in expense_context['rows_by_company'].get(company, []):
        dept = item['dept']
        if dept == 'OSD运营支持部' or 'OSD' in dept:
            continue
        if _should_skip_one_time_june_2026_accrual(company, posting_year, posting_month, dept):
            continue

        key = (
            item['gl_account'],
            dept,
            item['cost_center'],
            item['allocation_cost_center'],
            item['credit_order'],
        )
        grouped = grouped_items.get(key)
        if grouped is None:
            grouped_items[key] = {
                'gl_account': item['gl_account'],
                'dept': dept,
                'amount': item['amount'],
                'cost_center': item['cost_center'],
                'allocation_cost_center': item['allocation_cost_center'],
                'credit_order': item['credit_order'],
                'row_indices': [item['row_idx']],
            }
            continue

        grouped['amount'] += item['amount']
        grouped['row_indices'].append(item['row_idx'])

    for item in grouped_items.values():
        item['amount'] = _to_money(item['amount'])
        if item['amount'] == Decimal('0.00'):
            continue
        if item['amount'] < Decimal('0.00'):
            raise ValueError(
                f'{company} A10 聚合后金额为负数：科目 {item["gl_account"]} / 部门 {item["dept"]} / 金额 {item["amount"]}'
            )

        dept = item['dept']
        alloc_bucket = dict(timesheet_context['allocation_index'].get((posting_year, posting_month, company, dept), {}))
        if dept == 'HLD硬件逻辑部':
            alloc_bucket.pop('9201856', None)

        order_hours = sorted((order, hours) for order, hours in alloc_bucket.items() if hours > Decimal('0'))
        if not order_hours and company == '耐数信息' and item['cost_center'] == '20602020':
            company_bucket = {}
            for (year_num, month_num, bucket_company, _), bucket in timesheet_context['allocation_index'].items():
                if year_num != posting_year or month_num != posting_month or bucket_company != company:
                    continue
                for order, hours in bucket.items():
                    if hours > Decimal('0'):
                        company_bucket[order] = company_bucket.get(order, Decimal('0')) + hours
            order_hours = sorted(company_bucket.items())
        if not order_hours:
            raise ValueError(f'{company} {dept} 在 {posting_year}年{posting_month}月 未找到可用于 A10 分摊的工时项目')

        total_hours = sum((hours for _, hours in order_hours), Decimal('0'))
        if total_hours <= Decimal('0'):
            raise ValueError(f'{company} {dept} 在 {posting_year}年{posting_month}月 工时合计为 0，无法生成 A10')

        rows.append(
            _make_voucher_row(
                'A10',
                company_code,
                posting_date,
                item['gl_account'],
                item['amount'],
                text,
                50,
                item['cost_center'],
                item['credit_order'],
                voucher_type='SA',
            )
        )

        allocations = _allocate_amount_by_weights(
            item['amount'],
            [((item['gl_account'], item['allocation_cost_center'], order), hours) for order, hours in order_hours],
        )
        for key, alloc_amount in allocations.items():
            alloc_rows_by_key[key] = alloc_rows_by_key.get(key, Decimal('0.00')) + alloc_amount

    for (gl_account, cost_center, order), amount in sorted(alloc_rows_by_key.items()):
        amount = amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if amount == Decimal('0.00'):
            continue
        rows.append(
            _make_voucher_row(
                'A10',
                company_code,
                posting_date,
                gl_account,
                amount,
                text,
                40,
                cost_center,
                order,
                voucher_type='SA',
            )
        )

    return rows


def _clear_voucher_template_rows(ws):
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, 18):
            ws.cell(row=row_idx, column=col_idx).value = None


def _write_voucher_rows_to_template(ws, voucher_rows):
    for row_offset, row_data in enumerate(voucher_rows, start=2):
        for col_letter, value in row_data.items():
            ws[f'{col_letter}{row_offset}'] = value


def _validate_voucher_groups(voucher_rows):
    group_totals = {}
    for row in voucher_rows:
        amount = _to_money(row['M'])
        group = row['A']
        totals = group_totals.setdefault(group, {'debit': Decimal('0.00'), 'credit': Decimal('0.00')})
        if str(row['G']) == '40':
            totals['debit'] += amount
        elif str(row['G']) == '50':
            totals['credit'] += amount

    for group, totals in group_totals.items():
        if totals['debit'].quantize(Decimal('0.01')) != totals['credit'].quantize(Decimal('0.01')):
            raise ValueError(f'{group} 借贷不平：借方 {totals["debit"]}，贷方 {totals["credit"]}')


def _validate_cross_group_accounts(voucher_rows):
    account_totals = {}
    other_project_credit = Decimal('0.00')

    for row in voucher_rows:
        account = _format_code(row['I'])
        amount = _to_money(row['M'])
        bucket = account_totals.setdefault(account, {'debit': Decimal('0.00'), 'credit': Decimal('0.00')})
        if str(row['G']) == '40':
            bucket['debit'] += amount
        elif str(row['G']) == '50':
            bucket['credit'] += amount

        if account == '2211010001' and str(row['G']) == '50' and row['N'].endswith('其他项目'):
            other_project_credit += amount

    issues = []
    for account in ACCOUNTS_TO_CROSS_CHECK:
        totals = account_totals.get(account, {'debit': Decimal('0.00'), 'credit': Decimal('0.00')})
        debit_total = totals['debit'].quantize(Decimal('0.01'))
        credit_total = totals['credit'].quantize(Decimal('0.01'))

        if account == '2211010001':
            adjusted_credit = (credit_total - other_project_credit).quantize(Decimal('0.01'))
            if debit_total != adjusted_credit:
                issues.append(
                    f'{account} 跨凭证不平：借方 {debit_total}，贷方(扣除其他项目后) {adjusted_credit}'
                )
        elif debit_total != credit_total:
            issues.append(f'{account} 跨凭证不平：借方 {debit_total}，贷方 {credit_total}')

    return issues


def _generate_voucher_files(
    base_dir,
    company_output_paths,
    bank_records,
    payroll_year,
    payroll_month,
    log,
    timesheet_context,
    bonus_context,
    run_options=None,
):
    run_options = normalize_run_options(run_options)
    template_path = _get_voucher_template_path(base_dir)
    template_bytes = open(template_path, 'rb').read()
    voucher_paths = {}
    validation_summary = {}
    co_context = None
    shared_expense_context = None
    co_workorder_path = _get_co_workorder_path(base_dir, payroll_year, payroll_month)
    shared_expense_path = _get_shared_expense_path(base_dir, payroll_year, payroll_month)

    if requires_co_data(run_options):
        if os.path.isfile(co_workorder_path):
            log('正在读取 CO工单分摊 数据…')
            co_context = _load_co_workorder_context(co_workorder_path)
            log(
                f'CO工单分摊读取完成：{len(co_context["allocation_rows"])} 个参与分摊订单，'
                f'{len(co_context["expense_rows"])} 行待分摊费用',
                'ok',
            )
        else:
            log(f'未找到 CO工单分摊 文件，已跳过 A9：{co_workorder_path}', 'warn')

    if requires_shared_expense_data(run_options):
        if os.path.isfile(shared_expense_path):
            log('正在读取 待分摊费用 数据…')
            cost_center_map = _load_mappings(_get_mapping_path(base_dir))
            shared_expense_context = _load_shared_expense_context(shared_expense_path, cost_center_map)
            log(
                f'待分摊费用读取完成：2050 {len(shared_expense_context["rows_by_company"]["耐数电子"])} 行，'
                f'2060 {len(shared_expense_context["rows_by_company"]["耐数信息"])} 行',
                'ok',
            )
        else:
            log(f'未找到 待分摊费用 文件，已跳过 A10：{shared_expense_path}', 'warn')

    for company, company_output_path in company_output_paths.items():
        if not run_options.wants_company(company):
            continue

        company_code = COMPANY_NAME_TO_CODE[company]
        company_wb = load_workbook(company_output_path, data_only=True)
        company_ws = company_wb[company_wb.sheetnames[0]]
        total_row_idx = _find_total_row(company_ws)
        if total_row_idx is None:
            raise ValueError(f'{company} 工资单中未找到总计行')

        actual_voucher_rows = []
        if run_options.wants_voucher('A1'):
            actual_voucher_rows.extend(
                _build_a1_rows(company, company_code, company_ws, total_row_idx, bank_records, payroll_year, payroll_month)
            )
        if run_options.wants_voucher('A2'):
            actual_voucher_rows.extend(_build_a2_rows(company, company_code, bank_records, payroll_year, payroll_month))
        if run_options.wants_voucher('A3'):
            actual_voucher_rows.extend(
                _build_a3_rows(
                    company,
                    company_code,
                    company_ws,
                    bank_records,
                    payroll_year,
                    payroll_month,
                    timesheet_context,
                )
            )
        if run_options.wants_voucher('A4'):
            actual_voucher_rows.extend(
                _build_a4_rows(company, company_code, company_ws, total_row_idx, payroll_year, payroll_month, timesheet_context)
            )

        accrual_voucher_rows = []
        if run_options.wants_voucher('A5'):
            accrual_voucher_rows.extend(
                _build_a5_rows(company, company_code, company_ws, total_row_idx, payroll_year, payroll_month)
            )
        if run_options.wants_voucher('A6'):
            accrual_voucher_rows.extend(
                _build_a6_rows(company, company_code, company_ws, payroll_year, payroll_month, timesheet_context)
            )

        bonus_voucher_rows = []
        if bonus_context is not None and run_options.wants_voucher('A7'):
            bonus_voucher_rows.extend(_build_a7_rows(company, company_code, payroll_year, payroll_month, bonus_context))
        if bonus_context is not None and run_options.wants_voucher('A8'):
            bonus_voucher_rows.extend(
                _build_a8_rows(company, company_code, payroll_year, payroll_month, bonus_context, timesheet_context)
            )

        co_voucher_rows = []
        if company == '耐数电子' and co_context is not None and run_options.wants_voucher('A9'):
            co_voucher_rows.extend(_build_a9_rows(company_code, payroll_year, payroll_month, co_context))

        rd_voucher_rows = []
        if shared_expense_context is not None and run_options.wants_voucher('A10'):
            company_expense_rows = shared_expense_context['rows_by_company'].get(company, [])
            if company_expense_rows:
                rd_voucher_rows.extend(
                    _build_a10_rows(company, company_code, payroll_year, payroll_month, shared_expense_context, timesheet_context)
                )

        _validate_voucher_groups(actual_voucher_rows)
        _validate_voucher_groups(accrual_voucher_rows)
        _validate_voucher_groups(bonus_voucher_rows)
        if co_voucher_rows:
            _validate_voucher_groups(co_voucher_rows)
        if rd_voucher_rows:
            _validate_voucher_groups(rd_voucher_rows)

        cross_group_checked = all(run_options.wants_voucher(voucher) for voucher in ACTUAL_VOUCHERS)
        cross_group_issues = _validate_cross_group_accounts(actual_voucher_rows) if cross_group_checked else []

        actual_voucher_output_path = ''
        if actual_voucher_rows:
            actual_voucher_wb = load_workbook(io.BytesIO(template_bytes))
            actual_voucher_ws = actual_voucher_wb[actual_voucher_wb.sheetnames[0]]
            _clear_voucher_template_rows(actual_voucher_ws)
            _write_voucher_rows_to_template(actual_voucher_ws, actual_voucher_rows)
            actual_voucher_output_path = _build_voucher_output_path(base_dir, company, payroll_year, payroll_month)
            actual_voucher_wb.save(actual_voucher_output_path)
            voucher_paths[f'{company}（实际）'] = actual_voucher_output_path
            log(f'{company} 实际凭证文件已生成：{actual_voucher_output_path}', 'ok')

        accrual_voucher_output_path = ''
        if accrual_voucher_rows:
            accrual_voucher_wb = load_workbook(io.BytesIO(template_bytes))
            accrual_voucher_ws = accrual_voucher_wb[accrual_voucher_wb.sheetnames[0]]
            _clear_voucher_template_rows(accrual_voucher_ws)
            _write_voucher_rows_to_template(accrual_voucher_ws, accrual_voucher_rows)
            accrual_voucher_output_path = _build_accrual_voucher_output_path(base_dir, company, payroll_year, payroll_month)
            accrual_voucher_wb.save(accrual_voucher_output_path)
            voucher_paths[f'{company}（计提）'] = accrual_voucher_output_path
            log(f'{company} 计提凭证文件已生成：{accrual_voucher_output_path}', 'ok')

        bonus_voucher_output_path = ''
        if bonus_voucher_rows:
            bonus_voucher_wb = load_workbook(io.BytesIO(template_bytes))
            bonus_voucher_ws = bonus_voucher_wb[bonus_voucher_wb.sheetnames[0]]
            _clear_voucher_template_rows(bonus_voucher_ws)
            _write_voucher_rows_to_template(bonus_voucher_ws, bonus_voucher_rows)
            bonus_voucher_output_path = _build_bonus_voucher_output_path(base_dir, company, payroll_year, payroll_month)
            bonus_voucher_wb.save(bonus_voucher_output_path)
            voucher_paths[f'{company}（年终奖）'] = bonus_voucher_output_path
            log(f'{company} 年终奖凭证文件已生成：{bonus_voucher_output_path}', 'ok')

        if co_voucher_rows:
            co_voucher_wb = load_workbook(io.BytesIO(template_bytes))
            co_voucher_ws = co_voucher_wb[co_voucher_wb.sheetnames[0]]
            _clear_voucher_template_rows(co_voucher_ws)
            _write_voucher_rows_to_template(co_voucher_ws, co_voucher_rows)
            co_voucher_output_path = _build_co_voucher_output_path(base_dir, payroll_year, payroll_month)
            co_voucher_wb.save(co_voucher_output_path)
            voucher_paths[f'{company}（CO工单分摊）'] = co_voucher_output_path
            log(f'{company} CO工单分摊凭证文件已生成：{co_voucher_output_path}', 'ok')

        if rd_voucher_rows:
            rd_voucher_wb = load_workbook(io.BytesIO(template_bytes))
            rd_voucher_ws = rd_voucher_wb[rd_voucher_wb.sheetnames[0]]
            _clear_voucher_template_rows(rd_voucher_ws)
            _write_voucher_rows_to_template(rd_voucher_ws, rd_voucher_rows)
            rd_voucher_output_path = _build_rd_allocation_voucher_output_path(base_dir, company, payroll_year, payroll_month)
            rd_voucher_wb.save(rd_voucher_output_path)
            voucher_paths[f'{company}（研发费用分摊）'] = rd_voucher_output_path
            log(f'{company} 研发费用分摊凭证文件已生成：{rd_voucher_output_path}', 'ok')

        validation_summary[company] = {
            'a4_balanced': True,
            'a5_balanced': True,
            'a6_balanced': True,
            'a7_balanced': True,
            'a8_balanced': True,
            'a9_balanced': True,
            'a10_balanced': True,
            'group_balances': {voucher: True for voucher in VOUCHER_DISPLAY_ORDER if run_options.wants_voucher(voucher)},
            'cross_group_issues': cross_group_issues,
            'cross_group_checked': cross_group_checked,
        }

        if cross_group_checked and cross_group_issues:
            for issue in cross_group_issues:
                log(f'{company} 跨凭证校验异常：{issue}', 'warn')
        elif cross_group_checked:
            log(f'{company} A1-A4 科目对冲校验通过', 'ok')
        else:
            log(f'{company} 未选择完整的 A1-A4，已跳过跨凭证对冲校验', 'warn')

    return voucher_paths, validation_summary


def _clear_bank_summary(ws, end_row=12):
    for row_idx in range(2, end_row + 1):
        for col_idx in range(22, 30):
            ws.cell(row=row_idx, column=col_idx).value = None


def _save_company_workbooks(base_wb, base_dir, input_path, payroll_year, payroll_month, company_results, log):
    buffer = io.BytesIO()
    base_wb.save(buffer)
    workbook_bytes = buffer.getvalue()
    output_paths = {}

    for company, bank_results in company_results.items():
        company_output_path = _build_company_output_path(base_dir, company, payroll_year, payroll_month, input_path)
        company_wb = load_workbook(io.BytesIO(workbook_bytes))
        ws = company_wb[company_wb.sheetnames[0]]
        header_row_idx = _find_payroll_header_row(ws)
        data_start_row = header_row_idx + 1

        total_row_idx = _find_total_row(ws)
        if total_row_idx is None:
            raise ValueError('拆分公司文件时未找到“总计”行')

        if ws.max_row > total_row_idx:
            ws.delete_rows(total_row_idx + 1, ws.max_row - total_row_idx)

        for row_idx in range(total_row_idx - 1, data_start_row - 1, -1):
            row_company = _normalize_text(ws.cell(row=row_idx, column=1).value)
            row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 21)]
            if all(_is_blank(value) for value in row_values):
                ws.delete_rows(row_idx, 1)
                continue
            if row_company != company:
                ws.delete_rows(row_idx, 1)

        new_total_row_idx = _find_total_row(ws)
        if new_total_row_idx is None:
            raise ValueError(f'拆分 {company} 文件时总计行丢失')

        _recalculate_total_row(ws, new_total_row_idx, data_start_row)
        _clear_bank_summary(ws)
        header_ref = ws.cell(row=header_row_idx, column=18 if ws.max_column >= 18 else 17)
        _write_bank_summary(ws, bank_results, header_ref)

        company_wb.save(company_output_path)
        output_paths[company] = company_output_path
        log(f'{company} 文件已生成：{company_output_path}', 'ok')

    return output_paths


def fill_first_sheet_ab(input_path, base_dir, mapping_path, bank_dir, log, run_options=None):
    run_options = normalize_run_options(run_options)
    mapping_path = mapping_path or _get_mapping_path(base_dir)
    bank_dir = bank_dir or _get_bank_dir(base_dir, run_options.processing_year, run_options.processing_month)
    external_log = log
    log_entries = []

    def log(text, tag=''):
        log_entries.append({'tag': tag or 'info', 'text': text})
        external_log(text, tag)

    log('正在读取原始工资单…')
    wb = load_workbook(input_path)
    ws = wb[wb.sheetnames[0]]
    header_row_idx = _find_payroll_header_row(ws)
    data_start_row = header_row_idx + 1
    log(f'首个工作表：{ws.title}，共 {ws.max_row} 行')
    log(f'工资单表头位于第 {header_row_idx} 行，数据从第 {data_start_row} 行开始', 'ok')
    log('正在读取 Mapping 表…')
    cost_center_map = _load_mappings(mapping_path)
    log('Mapping 表读取完成', 'ok')
    timesheet_path = _get_timesheet_path(base_dir, run_options.processing_year, run_options.processing_month)
    bonus_path = _get_bonus_path(base_dir, run_options.processing_year, run_options.processing_month)
    if not os.path.isfile(timesheet_path):
        raise FileNotFoundError(f'找不到工时数据文件（请先放置）：{timesheet_path}')
    if requires_bonus_data(run_options) and not os.path.isfile(bonus_path):
        raise FileNotFoundError(f'找不到奖金数据文件（请先放置）：{bonus_path}')
    log('正在匹配工时数据中的内部订单…')
    _, timesheet_match_summary = _match_timesheet_internal_orders(timesheet_path, save_changes=False)
    if timesheet_match_summary['unmatched_rows']:
        raise ValueError(
            f'工时数据仍有 {len(timesheet_match_summary["unmatched_rows"])} 行项目未匹配到内部订单，无法继续后续处理'
        )
    timesheet_context = _load_timesheet_match_context(timesheet_path)
    bonus_context = _load_bonus_context(bonus_path, cost_center_map) if requires_bonus_data(run_options) else None
    log(
        f'工时内部订单匹配完成：已匹配 {timesheet_match_summary["matched_rows"]} 行，'
        f'补充逻辑命中 {timesheet_match_summary["used_extra_rows"]} 行',
        'ok',
    )
    bank_records = {}
    bank_scan_stats = {
        'source_file_count': 0,
        'scanned_row_count': 0,
        'deduped_row_count': 0,
        'kept_row_count': 0,
    }
    if requires_bank_data(run_options):
        log('正在读取银行流水…')
        bank_load_result = _load_bank_records_with_stats(bank_dir)
        bank_records = bank_load_result['records']
        bank_scan_stats = bank_load_result['stats']
        log(
            '银行流水读取完成：'
            f'扫描 {bank_scan_stats["source_file_count"]} 个文件，'
            f'保留 {bank_scan_stats["kept_row_count"]} 行，'
            f'去重 {bank_scan_stats["deduped_row_count"]} 行',
            'ok',
        )
    else:
        log('本次未选择 A1-A3，已跳过银行流水读取', 'warn')

    last_a = None
    last_b = None
    last_a_style_cell = None
    last_b_style_cell = None
    fill_a = 0
    fill_b = 0
    total_row_idx = ws.max_row + 1
    issue_rows = {}
    validation_issue_count = 0
    cost_center_issue_count = 0
    internal_order_issue_count = 0
    payroll_summary = {}
    rows_by_company = {}

    if _is_blank(ws.cell(row=data_start_row, column=1).value):
        for probe_row_idx in range(data_start_row, ws.max_row + 1):
            company_value = _normalize_text(ws.cell(row=probe_row_idx, column=1).value)
            if _is_total_row(company_value):
                break
            if company_value:
                if company_value == '耐数信息':
                    last_a = '耐数电子'
                    log('首个公司分组未写公司名称，按标准顺序识别为耐数电子', 'warn')
                break

    for row_idx in range(data_start_row, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1)
        cell_b = ws.cell(row=row_idx, column=2)

        if _is_total_row(cell_a.value):
            log(f'在第 {row_idx} 行识别到“总计”，从该行开始停止填充')
            total_row_idx = row_idx
            break

        if _is_blank(cell_a.value):
            if last_a is not None:
                cell_a.value = last_a
                if last_a_style_cell is not None:
                    _copy_cell_style(last_a_style_cell, cell_a)
                fill_a += 1
        else:
            last_a = cell_a.value
            last_a_style_cell = cell_a

        if _is_blank(cell_b.value):
            if last_b is not None:
                cell_b.value = last_b
                if last_b_style_cell is not None:
                    _copy_cell_style(last_b_style_cell, cell_b)
                fill_b += 1
        else:
            last_b = cell_b.value
            last_b_style_cell = cell_b

    log(f'A 列填充 {fill_a} 个空白单元格', 'ok')
    log(f'B 列填充 {fill_b} 个空白单元格', 'ok')

    header_ref = ws.cell(row=header_row_idx, column=18 if ws.max_column >= 18 else 17)
    data_style_col = 18 if ws.max_column >= 18 else 17
    s_header = ws.cell(row=header_row_idx, column=19)
    t_header = ws.cell(row=header_row_idx, column=20)
    s_header.value = '成本中心'
    _copy_cell_style(header_ref, s_header)
    t_header.value = '内部订单'
    _copy_cell_style(header_ref, t_header)

    if ws.column_dimensions['R'].width:
        ws.column_dimensions['S'].width = ws.column_dimensions['R'].width
    if ws.column_dimensions['Q'].width:
        ws.column_dimensions['T'].width = ws.column_dimensions['Q'].width

    for row_idx in range(data_start_row, total_row_idx):
        row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, 19)]
        if all(_is_blank(value) for value in row_values):
            continue

        row_issues = []
        company = _normalize_text(ws.cell(row=row_idx, column=1).value)
        dept = _normalize_text(ws.cell(row=row_idx, column=2).value)
        project = _normalize_text(ws.cell(row=row_idx, column=3).value)
        item_type = _normalize_text(ws.cell(row=row_idx, column=4).value)
        rows_by_company.setdefault(company, []).append(row_idx)

        expected_actual = _to_decimal(ws.cell(row=row_idx, column=5).value)
        for col_idx in range(11, 17):
            expected_actual -= _to_decimal(ws.cell(row=row_idx, column=col_idx).value)
        expected_actual = expected_actual.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        actual_value = _to_money(ws.cell(row=row_idx, column=17).value)
        company_summary = payroll_summary.setdefault(
            company,
            {
                'salary': Decimal('0'),
                'fund': Decimal('0'),
                'tax': Decimal('0'),
                'social': Decimal('0'),
            },
        )
        company_summary['salary'] += _to_decimal(ws.cell(row=row_idx, column=17).value)
        company_summary['fund'] += _to_decimal(ws.cell(row=row_idx, column=10).value) + _to_decimal(
            ws.cell(row=row_idx, column=14).value
        )
        company_summary['tax'] += _to_decimal(ws.cell(row=row_idx, column=15).value)
        company_summary['social'] += sum(
            (_to_decimal(ws.cell(row=row_idx, column=col_idx).value) for col_idx in (6, 7, 8, 9, 11, 12, 13)),
            Decimal('0'),
        )

        if expected_actual != actual_value:
            validation_issue_count += 1
            row_issues.append(f'实发校验失败(Q={actual_value}, 计算值={expected_actual})')

        company_cost_center_map = cost_center_map.get(company)
        if company_cost_center_map is None:
            row_issues.append(f'无法识别公司：{company or "空"}')
            cost_center_value = ''
            internal_order_value = ''
            cost_center_issue_count += 1
            internal_order_issue_count += 1
        else:
            cost_center_value = company_cost_center_map.get(dept, '')
            if not cost_center_value:
                cost_center_issue_count += 1
                row_issues.append(f'成本中心未匹配：{dept or "空"}')

            internal_order_value, project_for_match = _lookup_internal_order_from_timesheet(
                timesheet_context,
                company,
                dept,
                project,
            )
            if not internal_order_value:
                internal_order_issue_count += 1
                row_issues.append(f'内部订单未匹配：{project_for_match or "空"}')

        s_cell = ws.cell(row=row_idx, column=19)
        t_cell = ws.cell(row=row_idx, column=20)
        ref_cell = ws.cell(row=row_idx, column=data_style_col)
        _copy_cell_style(ref_cell, s_cell)
        _copy_cell_style(ref_cell, t_cell)
        s_cell.value = cost_center_value
        t_cell.value = internal_order_value

        if row_issues:
            issue_rows[row_idx] = row_issues

    for row_idx in issue_rows:
        _mark_row_red(ws, row_idx, 20)

    if internal_order_issue_count:
        raise ValueError(
            f'工资单中有 {internal_order_issue_count} 行内部订单(T列)未匹配，已停止处理，请修正工时数据或工资单项目后再运行'
        )

    if issue_rows:
        log(f'共发现 {len(issue_rows)} 行异常，已在输出文件中标红', 'warn')
    else:
        log('所有数据校验和匹配均通过', 'ok')

    payroll_year, payroll_month = run_options.payroll_period
    next_year, next_month = _shift_month(payroll_year, payroll_month, 1)
    bank_results = []
    company_bank_results = {}
    bank_issue_count = 0
    bank_salary_issue_count = 0
    bank_fund_issue_count = 0
    bank_tax_issue_count = 0
    bank_social_issue_count = 0
    bank_issue_companies = set()

    selected_companies = [company for company in payroll_summary if run_options.wants_company(company)]
    if requires_bank_data(run_options):
        log('开始核对银行流水…')
        for company in selected_companies:
            summary = payroll_summary[company]
            company_code = COMPANY_NAME_TO_CODE.get(company)
            if not company_code:
                continue

            salary_tax_bank = _summarize_bank_records(bank_records, company_code, next_year, next_month)
            fund_bank = _summarize_bank_records(bank_records, company_code, payroll_year, payroll_month)
            salary_match = _match_salary_combination(
                salary_tax_bank['salary_records'],
                summary['salary'].quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
            )
            tax_match = _match_bank_record_combination(
                salary_tax_bank['treasury_records'],
                (
                    summary['tax']
                    + _get_bonus_tax_adjustment(company_code, next_year, next_month)
                ).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
                '个税国库流水',
            )
            social_match = _match_bank_record_combination(
                salary_tax_bank['treasury_records'],
                summary['social'].quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
                '社保国库流水',
            )

            checks = [
                ('salary', '实发工资', summary['salary'].quantize(Decimal('0.01'), rounding=ROUND_HALF_UP), salary_match['matched_amount'], f'{next_year}-{next_month:02d}'),
                ('fund', '公积金', summary['fund'].quantize(Decimal('0.01'), rounding=ROUND_HALF_UP), fund_bank['fund'], f'{payroll_year}-{payroll_month:02d}'),
                (
                    'tax',
                    '个税',
                    (
                        summary['tax']
                        + _get_bonus_tax_adjustment(company_code, next_year, next_month)
                    ).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
                    tax_match['matched_amount'],
                    f'{next_year}-{next_month:02d}',
                ),
                ('social', '社保', summary['social'].quantize(Decimal('0.01'), rounding=ROUND_HALF_UP), social_match['matched_amount'], f'{next_year}-{next_month:02d}'),
            ]

            for item_key, item_label, payroll_amount, bank_amount, period_label in checks:
                diff = (payroll_amount - bank_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                passed = diff == Decimal('0.00')
                note = ''

                if item_key == 'salary':
                    passed = salary_match['matched']
                    diff = (payroll_amount - bank_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    note = salary_match['note']
                elif item_key == 'tax':
                    passed = tax_match['matched']
                    diff = (payroll_amount - bank_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    note = tax_match['note']
                    bonus_tax_adjustment = _get_bonus_tax_adjustment(company_code, next_year, next_month)
                    if bonus_tax_adjustment:
                        note = f'{note}；已包含上月年终奖个税 {bonus_tax_adjustment}'
                elif item_key == 'social':
                    passed = social_match['matched']
                    diff = (payroll_amount - bank_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    note = social_match['note']

                if item_key == 'fund' and company_code == '2050' and payroll_year == 2026 and payroll_month == 2:
                    passed = abs(diff) == Decimal('3650.00')
                    if passed:
                        note = '202602 特殊规则：公积金允许绝对差额 3650'

                source_files = salary_tax_bank['files'] if item_key in ('salary', 'tax', 'social') else fund_bank['files']
                if not source_files:
                    note = f'未找到 {period_label} 的匹配银行流水'
                elif not note:
                    note = '来源：' + '、'.join(source_files)
                elif item_key in ('salary', 'tax', 'social'):
                    note = note + '；来源：' + '、'.join(source_files)

                if not passed:
                    bank_issue_count += 1
                    bank_issue_companies.add(company)
                    if item_key == 'salary':
                        bank_salary_issue_count += 1
                    elif item_key == 'fund':
                        bank_fund_issue_count += 1
                    elif item_key == 'tax':
                        bank_tax_issue_count += 1
                    elif item_key == 'social':
                        bank_social_issue_count += 1
                    log(f'{company}{item_label}核对异常：工资表={payroll_amount}，流水={bank_amount}，差额={diff}', 'warn')
                else:
                    log(f'{company}{item_label}核对通过：{payroll_amount}', 'ok')

                bank_results.append(
                    {
                        'company': company,
                        'item_label': item_label,
                        'payroll_amount': payroll_amount,
                        'bank_amount': bank_amount,
                        'diff': diff,
                        'period_label': period_label,
                        'passed': passed,
                        'note': note,
                    }
                )
            company_bank_results[company] = [item for item in bank_results if item['company'] == company]
    else:
        for company in selected_companies:
            company_bank_results[company] = []
        log('本次未选择 A1-A3，已跳过银行核对', 'warn')

    _write_bank_summary(ws, bank_results, header_ref)

    for company in bank_issue_companies:
        for row_idx in rows_by_company.get(company, []):
            _mark_row_red(ws, row_idx, 20)

    output_paths = _save_company_workbooks(wb, base_dir, input_path, payroll_year, payroll_month, company_bank_results, log)
    key_tax_source_paths = {}
    try:
        tax_source_bank_records = bank_records if bank_records else _load_bank_records(bank_dir)
        key_tax_source_paths = _write_key_tax_source_report(
            base_dir,
            output_paths,
            tax_source_bank_records,
            payroll_year,
            payroll_month,
        )
        for path in key_tax_source_paths.values():
            log(f'重点税源采集信息已生成：{path}', 'ok')
    except Exception as exc:
        log(f'重点税源采集信息生成失败，已跳过：{exc}', 'warn')
    voucher_paths, voucher_validation_summary = _generate_voucher_files(
        base_dir,
        output_paths,
        bank_records,
        payroll_year,
        payroll_month,
        log,
        timesheet_context,
        bonus_context,
        run_options=run_options,
    )
    compensation_report_result = {}
    try:
        compensation_report_result = generate_compensation_report(base_dir)
        report_path = compensation_report_result.get('path')
        row_counts = compensation_report_result.get('rows_by_company', {})
        log(
            '实发薪酬表已生成：'
            f'{report_path}（2050 {row_counts.get("2050", 0)} 行，2060 {row_counts.get("2060", 0)} 行）',
            'ok',
        )
    except Exception as exc:
        log(f'实发薪酬表生成失败，已跳过：{exc}', 'warn')
    artifact_paths = _write_run_artifacts(
        base_dir,
        run_options,
        input_path,
        mapping_path,
        bank_dir,
        timesheet_path,
        bonus_path,
        _get_co_workorder_path(base_dir, payroll_year, payroll_month),
        _get_shared_expense_path(base_dir, payroll_year, payroll_month),
        output_paths,
        voucher_paths,
        voucher_validation_summary,
        log_entries,
        bank_scan_stats,
        compensation_report_result.get('path'),
    )
    if key_tax_source_paths:
        artifact_paths['key_tax_source'] = key_tax_source_paths
    if compensation_report_result:
        artifact_paths['compensation_report'] = compensation_report_result

    return {
        'sheet_name': ws.title,
        'row_count': ws.max_row,
        'fill_a': fill_a,
        'fill_b': fill_b,
        'issue_count': len(issue_rows),
        'validation_issue_count': validation_issue_count,
        'cost_center_issue_count': cost_center_issue_count,
        'internal_order_issue_count': internal_order_issue_count,
        'bank_issue_count': bank_issue_count,
        'bank_salary_issue_count': bank_salary_issue_count,
        'bank_fund_issue_count': bank_fund_issue_count,
        'bank_tax_issue_count': bank_tax_issue_count,
        'bank_social_issue_count': bank_social_issue_count,
        'bank_scan_stats': bank_scan_stats,
        'processing_period_label': format_period_label(run_options.processing_year, run_options.processing_month),
        'payroll_period_label': format_period_label(payroll_year, payroll_month),
        'output_paths': output_paths,
        'voucher_paths': voucher_paths,
        'voucher_validation_summary': voucher_validation_summary,
        'artifact_paths': artifact_paths,
    }
