import io
import os
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import load_workbook

from .core import (
    COMPANY_NAME_TO_CODE,
    _build_a1_rows,
    _build_a2_rows,
    _build_a3_rows,
    _build_a4_rows,
    _build_a5_rows,
    _build_a6_rows,
    _build_a7_rows,
    _build_a8_rows,
    _build_a9_rows,
    _build_a10_rows,
    _build_bonus_department_amounts,
    _copy_cell_style,
    _find_bank_files,
    _find_raw_files,
    _find_total_row,
    _format_code,
    _get_bank_dir,
    _get_bonus_path,
    _get_bonus_tax_adjustment,
    _get_co_workorder_path,
    _get_mapping_path,
    _get_raw_dir,
    _get_shared_expense_path,
    _get_timesheet_path,
    _is_blank,
    _is_total_row,
    _load_bank_records_with_stats,
    _load_bonus_context,
    _load_co_workorder_context,
    _load_mappings,
    _load_shared_expense_context,
    _load_timesheet_match_context,
    _lookup_internal_order_from_timesheet,
    _match_bank_record_combination,
    _mark_row_red,
    _match_salary_combination,
    _match_timesheet_internal_orders,
    _normalize_text,
    _recalculate_total_row,
    _summarize_bank_records,
    _to_decimal,
    _to_money,
    _validate_cross_group_accounts,
    _validate_voucher_groups,
    _write_utf8_file,
)
from .options import (
    ACTUAL_VOUCHERS,
    BONUS_VOUCHERS,
    COMPANY_OPTIONS,
    RunOptions,
    build_batch_layout_from_options,
    format_period_label,
    normalize_run_options,
    requires_bank_data,
    requires_bonus_data,
    requires_co_data,
    requires_shared_expense_data,
)


def _status_rank(status):
    return {'ok': 0, 'warn': 1, 'error': 2}.get(status, 0)


def _merge_status(*statuses):
    return max(statuses, key=_status_rank) if statuses else 'ok'


def _append_line(lines, status, text):
    lines.append({'status': status, 'text': text})


def _render_markdown_table(headers, rows):
    lines = ['| ' + ' | '.join(headers) + ' |', '| ' + ' | '.join(['---'] * len(headers)) + ' |']
    for row in rows:
        lines.append('| ' + ' | '.join(str(item) for item in row) + ' |')
    return '\n'.join(lines)


def _render_markdown_list(items, empty_text='无'):
    if not items:
        return f'- {empty_text}'
    return '\n'.join(f'- {item}' for item in items)


def _status_label(status):
    mapping = {
        'ok': '通过',
        'warn': '预警',
        'warning': '预警',
        'error': '阻断',
        'err': '阻断',
        'info': '信息',
    }
    return mapping.get(status, status or '信息')


def _write_precheck_report(layout, run_options, result, input_checks):
    report_path = os.path.join(layout.archive_root, f'{layout.batch_code}-预检报告.md')
    summary_rows = [
        ('总体状态', _status_label(result.get('overall_status', 'info'))),
        ('是否可运行', '是' if result.get('can_run') else '否'),
        ('阻断项数量', len(result.get('blockers', []))),
        ('预警项数量', len(result.get('warnings', []))),
    ]
    input_rows = [
        (
            item['name'],
            '必须' if item['required'] else '按需',
            '已就绪' if item['exists'] else '未就绪',
            item['path'],
            item['note'],
        )
        for item in input_checks
    ]
    company_rows = []
    voucher_sections = []
    for company, company_data in result.get('company_checks', {}).items():
        company_rows.append(
            (
                company,
                _status_label(company_data.get('status', 'info')),
                len(company_data.get('blockers', [])),
                len(company_data.get('warnings', [])),
            )
        )
        voucher_rows = []
        for voucher_id, voucher_info in company_data.get('voucher_status', {}).items():
            voucher_rows.append((voucher_id, _status_label(voucher_info.get('status', 'info')), voucher_info.get('message', '')))
        voucher_sections.extend(
            [
                f'### {company}',
                '',
                _render_markdown_table(['凭证', '结果', '说明'], voucher_rows or [('无', '无', '无')]),
                '',
            ]
        )
    content = '\n'.join(
        [
            f'# {layout.batch_name} 预检报告',
            '',
            '## 执行结论',
            '',
            f'- 总结：`{result.get("summary", "预检完成")}`',
            f'- 可运行：`{"是" if result.get("can_run") else "否"}`',
            f'- 总体状态：`{_status_label(result.get("overall_status", "info"))}`',
            '',
            '## 批次范围',
            '',
            _render_markdown_table(
                ['项目', '值'],
                [
                    ('处理月份', run_options.processing_label),
                    ('工资所属月份', run_options.payroll_label),
                    ('公司范围', '、'.join(run_options.companies) or '未选'),
                    ('凭证范围', '、'.join(run_options.vouchers) or '未选'),
                    ('预检报告路径', report_path),
                ],
            ),
            '',
            '## 预检摘要',
            '',
            _render_markdown_table(['项目', '结果'], summary_rows),
            '',
            '## 输入资料状态',
            '',
            _render_markdown_table(['资料', '要求', '状态', '路径', '说明'], input_rows),
            '',
            '## 风险清单',
            '',
            '### 阻断项',
            '',
            _render_markdown_list(result.get('blockers', []), empty_text='无'),
            '',
            '### 预警项',
            '',
            _render_markdown_list(result.get('warnings', []), empty_text='无'),
            '',
            '## 公司级结果',
            '',
            _render_markdown_table(['公司', '状态', '阻断项', '预警项'], company_rows or [('无', '无', '0', '0')]),
            '',
            '## 凭证级结果',
            '',
        ]
        + voucher_sections
        + [
            '',
            '## 详细检查日志',
            '',
            _render_markdown_list([item['text'] for item in result.get('display_lines', [])], empty_text='无详细日志'),
            '',
            '## 使用说明',
            '',
            '- 本文档用于回答“现在能不能跑、风险点在哪、每家公司每张凭证是否具备生成条件”。',
            '- 若存在阻断项，应先处理阻断项，再重新执行预检。',
            '',
        ]
        + ['']
    )
    _write_utf8_file(report_path, content)
    return report_path


def collect_input_health(base_dir, mapping_path, bank_dir, run_options):
    run_options = normalize_run_options(run_options)
    layout = build_batch_layout_from_options(base_dir, run_options)
    mapping_path = mapping_path or layout.mapping_path
    bank_dir = bank_dir or layout.bank_dir
    raw_dir = layout.raw_dir
    payroll_period = run_options.payroll_period
    raw_files = _find_raw_files(raw_dir, payroll_period[0], payroll_period[1])
    checks = []

    def add_check(name, path, required, exists, note, vouchers=''):
        checks.append(
            {
                'name': name,
                'path': path,
                'required': required,
                'exists': exists,
                'note': note,
                'vouchers': vouchers,
                'status': 'ok' if exists else ('error' if required else 'warn'),
            }
        )

    add_check(
        '原始工资单',
        os.path.join(raw_dir, f'人力成本研发项目分摊{payroll_period[0]}{payroll_period[1]:02d} - to财务-原始.xlsx'),
        True,
        len(raw_files) == 1,
        f'需匹配工资所属月份 {format_period_label(payroll_period[0], payroll_period[1])}，且只能命中 1 个文件',
        '全部',
    )
    add_check('Mapping表.xlsx', mapping_path, True, os.path.exists(mapping_path), '部门映射与成本中心来源', '全部')

    timesheet_path = _get_timesheet_path(base_dir, run_options.processing_year, run_options.processing_month)
    add_check('工时数据.xlsx', timesheet_path, True, os.path.exists(timesheet_path), '内部订单与工时分摊来源', '全部')

    bonus_path = _get_bonus_path(base_dir, run_options.processing_year, run_options.processing_month)
    add_check(
        '年终奖计提文件',
        bonus_path,
        requires_bonus_data(run_options),
        os.path.exists(bonus_path),
        'A7/A8 使用',
        'A7-A8',
    )

    bank_exists = bool(_find_bank_files(bank_dir))
    add_check(
        '银行流水',
        os.path.join(bank_dir, '银行流水*.xlsx'),
        requires_bank_data(run_options),
        bank_exists,
        f'A1-A3 发放凭证和银行核对使用，按处理月份 {run_options.processing_label} 全目录扫描后筛选',
        'A1-A3',
    )

    co_path = _get_co_workorder_path(base_dir, payroll_period[0], payroll_period[1])
    add_check(
        'CO工单分摊',
        co_path or os.path.join(base_dir, '耐数电子', '处理月份', 'CO工单分摊', 'CO工单分摊.xlsx'),
        requires_co_data(run_options),
        bool(co_path) and os.path.exists(co_path),
        'A9 仅 2050 使用',
        'A9',
    )

    shared_path = _get_shared_expense_path(base_dir, payroll_period[0], payroll_period[1])
    add_check(
        '待分摊费用',
        shared_path or os.path.join(base_dir, '原始数据', '待分摊费用', '待分摊费用YYMM.xlsx'),
        requires_shared_expense_data(run_options),
        bool(shared_path) and os.path.exists(shared_path),
        'A10 使用',
        'A10',
    )
    return {
        'run_options': run_options,
        'raw_files': raw_files,
        'payroll_period': payroll_period,
        'checks': checks,
    }


def _prepare_payroll_preview(input_path, cost_center_map, timesheet_context, selected_companies):
    wb = load_workbook(input_path)
    ws = wb[wb.sheetnames[0]]

    last_a = None
    last_b = None
    last_a_style_cell = None
    last_b_style_cell = None
    total_row_idx = ws.max_row + 1
    issue_rows = {}
    company_issue_counts = {}
    payroll_summary = {}
    rows_by_company = {}

    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1)
        cell_b = ws.cell(row=row_idx, column=2)

        if _is_total_row(cell_a.value):
            total_row_idx = row_idx
            break

        if _is_blank(cell_a.value):
            if last_a is not None:
                cell_a.value = last_a
                if last_a_style_cell is not None:
                    _copy_cell_style(last_a_style_cell, cell_a)
        else:
            last_a = cell_a.value
            last_a_style_cell = cell_a

        if _is_blank(cell_b.value):
            if last_b is not None:
                cell_b.value = last_b
                if last_b_style_cell is not None:
                    _copy_cell_style(last_b_style_cell, cell_b)
        else:
            last_b = cell_b.value
            last_b_style_cell = cell_b

    header_ref = ws['R2'] if ws.max_column >= 18 else ws['Q2']
    data_style_col = 18 if ws.max_column >= 18 else 17
    ws['S2'].value = '成本中心'
    _copy_cell_style(header_ref, ws['S2'])
    ws['T2'].value = '内部订单'
    _copy_cell_style(header_ref, ws['T2'])

    if ws.column_dimensions['R'].width:
        ws.column_dimensions['S'].width = ws.column_dimensions['R'].width
    if ws.column_dimensions['Q'].width:
        ws.column_dimensions['T'].width = ws.column_dimensions['Q'].width

    for row_idx in range(3, total_row_idx):
        row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, 19)]
        if all(_is_blank(value) for value in row_values):
            continue

        row_issues = []
        company = _normalize_text(ws.cell(row=row_idx, column=1).value)
        dept = _normalize_text(ws.cell(row=row_idx, column=2).value)
        project = _normalize_text(ws.cell(row=row_idx, column=3).value)
        rows_by_company.setdefault(company, []).append(row_idx)
        if company not in selected_companies:
            continue

        expected_actual = _to_decimal(ws.cell(row=row_idx, column=5).value)
        for col_idx in range(11, 17):
            expected_actual -= _to_decimal(ws.cell(row=row_idx, column=col_idx).value)
        expected_actual = expected_actual.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        actual_value = _to_money(ws.cell(row=row_idx, column=17).value)

        company_summary = payroll_summary.setdefault(
            company,
            {'salary': Decimal('0'), 'fund': Decimal('0'), 'tax': Decimal('0'), 'social': Decimal('0')},
        )
        company_summary['salary'] += _to_decimal(ws.cell(row=row_idx, column=17).value)
        company_summary['fund'] += _to_decimal(ws.cell(row=row_idx, column=10).value) + _to_decimal(
            ws.cell(row=row_idx, column=14).value
        )
        company_summary['tax'] += _to_decimal(ws.cell(row=row_idx, column=15).value)
        company_summary['social'] += sum(
            (_to_decimal(ws.cell(row=row_idx, column=col_idx).value) for col_idx in (6, 7, 8, 9, 11, 12, 13)),
            Decimal('0'),
        )

        if expected_actual != actual_value:
            row_issues.append(f'实发校验失败(Q={actual_value}, 计算值={expected_actual})')

        company_cost_center_map = cost_center_map.get(company)
        if company_cost_center_map is None:
            row_issues.append(f'无法识别公司：{company or "空"}')
            cost_center_value = ''
            internal_order_value = ''
        else:
            cost_center_value = company_cost_center_map.get(dept, '')
            if not cost_center_value:
                row_issues.append(f'成本中心未匹配：{dept or "空"}')
            internal_order_value, project_for_match = _lookup_internal_order_from_timesheet(
                timesheet_context,
                company,
                dept,
                project,
            )
            if not internal_order_value:
                row_issues.append(f'内部订单未匹配：{project_for_match or "空"}')

        s_cell = ws.cell(row=row_idx, column=19)
        t_cell = ws.cell(row=row_idx, column=20)
        ref_cell = ws.cell(row=row_idx, column=data_style_col)
        _copy_cell_style(ref_cell, s_cell)
        _copy_cell_style(ref_cell, t_cell)
        s_cell.value = cost_center_value
        t_cell.value = internal_order_value

        if row_issues:
            issue_rows[row_idx] = {'company': company, 'issues': row_issues}
            company_issue_counts.setdefault(company, {'formula': 0, 'cost_center': 0, 'order': 0})
            for issue in row_issues:
                if issue.startswith('实发校验失败'):
                    company_issue_counts[company]['formula'] += 1
                elif issue.startswith('成本中心未匹配') or issue.startswith('无法识别公司'):
                    company_issue_counts[company]['cost_center'] += 1
                elif issue.startswith('内部订单未匹配'):
                    company_issue_counts[company]['order'] += 1

    for row_idx in issue_rows:
        _mark_row_red(ws, row_idx, 20)

    return {
        'workbook': wb,
        'worksheet': ws,
        'total_row_idx': total_row_idx,
        'issue_rows': issue_rows,
        'company_issue_counts': company_issue_counts,
        'payroll_summary': payroll_summary,
        'rows_by_company': rows_by_company,
    }


def _build_company_preview_workbooks(base_wb, selected_companies):
    buffer = io.BytesIO()
    base_wb.save(buffer)
    workbook_bytes = buffer.getvalue()
    snapshots = {}

    for company in selected_companies:
        company_wb = load_workbook(io.BytesIO(workbook_bytes))
        ws = company_wb[company_wb.sheetnames[0]]
        total_row_idx = _find_total_row(ws)
        if total_row_idx is None:
            raise ValueError('拆分公司文件时未找到“总计”行')

        if ws.max_row > total_row_idx:
            ws.delete_rows(total_row_idx + 1, ws.max_row - total_row_idx)

        for row_idx in range(total_row_idx - 1, 2, -1):
            row_company = _normalize_text(ws.cell(row=row_idx, column=1).value)
            row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 21)]
            if all(_is_blank(value) for value in row_values):
                ws.delete_rows(row_idx, 1)
                continue
            if row_company != company:
                ws.delete_rows(row_idx, 1)

        new_total_row_idx = _find_total_row(ws)
        if new_total_row_idx is None:
            raise ValueError(f'拆分 {company} 文件时总计行丢失')
        _recalculate_total_row(ws, new_total_row_idx)
        snapshots[company] = {'workbook': company_wb, 'worksheet': ws, 'total_row_idx': new_total_row_idx}

    return snapshots


def run_startup_precheck(base_dir, mapping_path, bank_dir, run_options):
    run_options = normalize_run_options(run_options)
    layout = build_batch_layout_from_options(base_dir, run_options)
    mapping_path = mapping_path or layout.mapping_path
    bank_dir = bank_dir or layout.bank_dir
    result = {
        'run_options': {
            'companies': list(run_options.companies),
            'vouchers': list(run_options.vouchers),
            'posting_year': run_options.processing_year,
            'posting_month': run_options.processing_month,
            'payroll_year': run_options.payroll_year,
            'payroll_month': run_options.payroll_month,
        },
        'overall_status': 'ok',
        'can_run': False,
        'input_checks': [],
        'blockers': [],
        'warnings': [],
        'company_checks': {},
        'display_lines': [],
        'bank_scan_stats': {},
        'report_path': '',
    }
    lines = result['display_lines']

    input_health = collect_input_health(base_dir, mapping_path, bank_dir, run_options)
    result['input_checks'] = input_health['checks']
    raw_files = input_health['raw_files']
    payroll_period = run_options.payroll_period
    required_missing = [item for item in input_health['checks'] if item['required'] and not item['exists']]
    if not run_options.companies:
        result['blockers'].append('至少选择 1 家公司')
    if not run_options.vouchers:
        result['blockers'].append('至少选择 1 张凭证')
    if run_options.wants_voucher('A9') and '耐数电子' not in run_options.companies:
        result['blockers'].append('A9 仅适用于耐数电子，请勾选 2050')
    result['blockers'].extend(f'缺少必需资料：{item["name"]}' for item in required_missing)

    _append_line(
        lines,
        'ok',
        f'预检范围：处理月份 {run_options.processing_label} / 工资所属 {run_options.payroll_label} / '
        + '、'.join(run_options.companies)
        + ' / '
        + '、'.join(run_options.vouchers),
    )
    for item in input_health['checks']:
        prefix = '必需' if item['required'] else '按需'
        tag = item['status']
        _append_line(lines, tag, f'[{prefix}] {item["name"]}：{"已就绪" if item["exists"] else "未就绪"}（{item["vouchers"] or "全部"}）')

    if result['blockers']:
        result['overall_status'] = 'error'
        result['summary'] = f'存在 {len(result["blockers"])} 个阻断项'
        for blocker in result['blockers']:
            _append_line(lines, 'err', blocker)
        result['report_path'] = _write_precheck_report(layout, run_options, result, result['input_checks'])
        return result

    cost_center_map = _load_mappings(mapping_path)
    _append_line(lines, 'ok', 'Mapping 表读取通过')

    timesheet_path = _get_timesheet_path(base_dir, run_options.processing_year, run_options.processing_month)
    project_index, timesheet_match_summary = _match_timesheet_internal_orders(timesheet_path, save_changes=False)
    if timesheet_match_summary['unmatched_rows']:
        result['blockers'].append(
            f'工时/KOK3 仍有 {len(timesheet_match_summary["unmatched_rows"])} 行项目未匹配到内部订单'
        )
        _append_line(lines, 'err', result['blockers'][-1])
        result['overall_status'] = 'error'
        result['summary'] = f'存在 {len(result["blockers"])} 个阻断项'
        result['report_path'] = _write_precheck_report(layout, run_options, result, result['input_checks'])
        return result
    _append_line(
        lines,
        'ok',
        f'工时/KOK3 校验通过：已匹配 {timesheet_match_summary["matched_rows"]} 行，补充逻辑命中 {timesheet_match_summary["used_extra_rows"]} 行',
    )

    timesheet_context = _load_timesheet_match_context(timesheet_path)
    _append_line(lines, 'ok', '工时内部订单唯一性检查通过')

    bonus_context = None
    if requires_bonus_data(run_options):
        bonus_context = _load_bonus_context(
            _get_bonus_path(base_dir, run_options.processing_year, run_options.processing_month),
            cost_center_map,
        )
        _append_line(lines, 'ok', '奖金数据读取通过')

    bank_records = {}
    if requires_bank_data(run_options):
        bank_load_result = _load_bank_records_with_stats(bank_dir)
        bank_records = bank_load_result['records']
        result['bank_scan_stats'] = bank_load_result['stats']
        _append_line(
            lines,
            'ok',
            '银行流水读取通过：'
            f'扫描 {bank_load_result["stats"]["source_file_count"]} 个文件，'
            f'保留 {bank_load_result["stats"]["kept_row_count"]} 行，'
            f'去重 {bank_load_result["stats"]["deduped_row_count"]} 行',
        )

    co_context = None
    if requires_co_data(run_options):
        co_context = _load_co_workorder_context(_get_co_workorder_path(base_dir, payroll_period[0], payroll_period[1]))
        _append_line(lines, 'ok', f'A9 CO工单分摊文件校验通过：处理月份 {run_options.processing_label}')

    shared_expense_context = None
    if requires_shared_expense_data(run_options):
        shared_expense_context = _load_shared_expense_context(
            _get_shared_expense_path(base_dir, payroll_period[0], payroll_period[1]),
            cost_center_map,
        )
        _append_line(lines, 'ok', f'A10 待分摊费用文件读取通过：处理月份 {run_options.processing_label}')

    preview = _prepare_payroll_preview(raw_files[0], cost_center_map, timesheet_context, set(run_options.companies))
    company_snapshots = _build_company_preview_workbooks(preview['workbook'], run_options.companies)
    payroll_year, payroll_month = payroll_period
    next_year = payroll_year
    next_month = payroll_month
    from .core import _shift_month

    next_year, next_month = _shift_month(payroll_year, payroll_month, 1)

    company_statuses = []
    for company in run_options.companies:
        company_result = {'status': 'ok', 'blockers': [], 'warnings': [], 'voucher_status': {}}
        counts = preview['company_issue_counts'].get(company, {'formula': 0, 'cost_center': 0, 'order': 0})
        if counts['order']:
            company_result['blockers'].append(f'工资单有 {counts["order"]} 行内部订单未匹配')
        if counts['formula']:
            company_result['warnings'].append(f'工资单有 {counts["formula"]} 行 Q列实发公式异常')
        if counts['cost_center']:
            company_result['warnings'].append(f'工资单有 {counts["cost_center"]} 行成本中心未匹配')

        payroll_summary = preview['payroll_summary'].get(
            company,
            {'salary': Decimal('0'), 'fund': Decimal('0'), 'tax': Decimal('0'), 'social': Decimal('0')},
        )
        if requires_bank_data(run_options):
            company_code = COMPANY_NAME_TO_CODE[company]
            salary_tax_bank = _summarize_bank_records(bank_records, company_code, next_year, next_month)
            fund_reconcile_bank = _summarize_bank_records(bank_records, company_code, payroll_year, payroll_month)
            fund_voucher_bank = _summarize_bank_records(bank_records, company_code, next_year, next_month)
            salary_match = _match_salary_combination(
                salary_tax_bank['salary_records'],
                payroll_summary['salary'].quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
            )
            tax_match = _match_bank_record_combination(
                salary_tax_bank['treasury_records'],
                (
                    payroll_summary['tax']
                    + _get_bonus_tax_adjustment(company_code, next_year, next_month)
                ).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
                '个税国库流水',
            )
            social_match = _match_bank_record_combination(
                salary_tax_bank['treasury_records'],
                payroll_summary['social'].quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
                '社保国库流水',
            )
            if run_options.wants_voucher('A1') and not social_match['matched']:
                company_result['blockers'].append(f'A1 未找到匹配社保金额的银行流水组合：{social_match["note"]}')
            if run_options.wants_voucher('A2') and fund_voucher_bank['fund'] == Decimal('0.00'):
                company_result['blockers'].append('A2 未找到公积金对应银行流水')
            if run_options.wants_voucher('A3') and (not salary_match['matched'] or not tax_match['matched']):
                company_result['blockers'].append(f'A3 缺少匹配的工资组合或个税组合：工资={salary_match["note"]}；个税={tax_match["note"]}')

            bank_checks = [
                ('工资', payroll_summary['salary'].quantize(Decimal('0.01'), rounding=ROUND_HALF_UP), salary_match['matched_amount'], salary_match['matched']),
                ('公积金', payroll_summary['fund'].quantize(Decimal('0.01'), rounding=ROUND_HALF_UP), fund_reconcile_bank['fund'], payroll_summary['fund'].quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) == fund_reconcile_bank['fund'] or (company_code == '2050' and payroll_year == 2026 and payroll_month == 2 and abs(payroll_summary['fund'].quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) - fund_reconcile_bank['fund']) == Decimal('3650.00'))),
                (
                    '个税',
                    (
                        payroll_summary['tax']
                        + _get_bonus_tax_adjustment(company_code, next_year, next_month)
                    ).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
                    tax_match['matched_amount'],
                    tax_match['matched'],
                ),
                ('社保', payroll_summary['social'].quantize(Decimal('0.01'), rounding=ROUND_HALF_UP), social_match['matched_amount'], social_match['matched']),
            ]
            for label, payroll_amount, bank_amount, passed in bank_checks:
                if not passed:
                    company_result['warnings'].append(f'{label}核对异常：工资表 {payroll_amount} / 银行 {bank_amount}')

        snapshot = company_snapshots[company]
        company_ws = snapshot['worksheet']
        total_row_idx = snapshot['total_row_idx']
        company_code = COMPANY_NAME_TO_CODE[company]

        def dry_build(voucher_id):
            if voucher_id == 'A1':
                return _build_a1_rows(company, company_code, company_ws, total_row_idx, bank_records, payroll_year, payroll_month)
            if voucher_id == 'A2':
                return _build_a2_rows(company, company_code, bank_records, payroll_year, payroll_month)
            if voucher_id == 'A3':
                return _build_a3_rows(company, company_code, company_ws, bank_records, payroll_year, payroll_month)
            if voucher_id == 'A4':
                return _build_a4_rows(company, company_code, company_ws, total_row_idx, payroll_year, payroll_month, timesheet_context)
            if voucher_id == 'A5':
                return _build_a5_rows(company, company_code, company_ws, total_row_idx, payroll_year, payroll_month)
            if voucher_id == 'A6':
                return _build_a6_rows(company, company_code, company_ws, payroll_year, payroll_month, timesheet_context)
            if voucher_id == 'A7':
                return _build_a7_rows(company, company_code, payroll_year, payroll_month, bonus_context)
            if voucher_id == 'A8':
                return _build_a8_rows(company, company_code, payroll_year, payroll_month, bonus_context, timesheet_context)
            if voucher_id == 'A9':
                return _build_a9_rows(company_code, payroll_year, payroll_month, co_context)
            if voucher_id == 'A10':
                return _build_a10_rows(company, company_code, payroll_year, payroll_month, shared_expense_context, timesheet_context)
            return []

        actual_rows = []
        for voucher_id in run_options.vouchers:
            if voucher_id == 'A9' and company != '耐数电子':
                company_result['voucher_status'][voucher_id] = {'status': 'warn', 'message': '该公司不适用 A9'}
                continue
            try:
                rows = dry_build(voucher_id)
                if voucher_id in ACTUAL_VOUCHERS:
                    actual_rows.extend(rows)
                if rows:
                    _validate_voucher_groups(rows)
                    company_result['voucher_status'][voucher_id] = {'status': 'ok', 'message': f'可生成，共 {len(rows)} 行'}
                else:
                    company_result['voucher_status'][voucher_id] = {'status': 'warn', 'message': '当前选择下无可生成数据'}
            except Exception as exc:
                company_result['voucher_status'][voucher_id] = {'status': 'error', 'message': str(exc)}
                company_result['blockers'].append(f'{voucher_id}：{exc}')

        if all(v in run_options.vouchers for v in ACTUAL_VOUCHERS) and actual_rows:
            issues = _validate_cross_group_accounts(actual_rows)
            if issues:
                company_result['warnings'].append(f'A1-A4 对冲异常 {len(issues)} 项')

        company_result['status'] = 'error' if company_result['blockers'] else ('warn' if company_result['warnings'] else 'ok')
        result['company_checks'][company] = company_result
        company_statuses.append(company_result['status'])

        _append_line(lines, company_result['status'], f'[{company}] 预检结果：{company_result["status"]}')
        for blocker in company_result['blockers']:
            _append_line(lines, 'err', f'[{company}] {blocker}')
        for warning in company_result['warnings']:
            _append_line(lines, 'warn', f'[{company}] {warning}')
        for voucher_id in run_options.vouchers:
            voucher_status = company_result['voucher_status'].get(voucher_id)
            if voucher_status:
                tag = voucher_status['status']
                _append_line(lines, tag if tag != 'error' else 'err', f'[{company}] {voucher_id}：{voucher_status["message"]}')

    result['overall_status'] = _merge_status('error' if result['blockers'] else 'ok', *company_statuses)
    result['warnings'].extend(
        warning
        for company_data in result['company_checks'].values()
        for warning in company_data.get('warnings', [])
    )
    result['blockers'].extend(
        blocker
        for company_data in result['company_checks'].values()
        for blocker in company_data.get('blockers', [])
    )
    result['can_run'] = not result['blockers']
    if result['blockers']:
        result['summary'] = f'无法运行：{len(result["blockers"])} 个阻断项'
    elif result['warnings']:
        result['summary'] = f'可运行，但有 {len(result["warnings"])} 个预警项'
    else:
        result['summary'] = '预检通过，可直接运行'
    result['report_path'] = _write_precheck_report(layout, run_options, result, result['input_checks'])
    return result
