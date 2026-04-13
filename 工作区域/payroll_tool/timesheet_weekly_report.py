import calendar
import os
from datetime import datetime

from openpyxl import load_workbook

from .core import _copy_cell_style, _get_timesheet_path, _normalize_text


TEMPLATE_FILENAME = '耐数周报工时-模板.xlsx'
FALLBACK_TEMPLATE_FILENAME = '耐数周报工时-202602.xlsx'
OUTPUT_FILENAME_TEMPLATE = '耐数周报工时-{year}{month:02d}.xlsx'
SOURCE_SHEET_NAME = '工时汇总'
OUTPUT_SHEET_NAME = '汇总'
DATA_START_ROW = 2
DATA_COLUMNS = 8
DEFAULT_MONTHS = (1, 3)


def _get_template_path(base_dir):
    timesheet_dir = os.path.join(base_dir, '原始数据', '工时数据')
    primary_path = os.path.join(timesheet_dir, TEMPLATE_FILENAME)
    if os.path.exists(primary_path):
        return primary_path
    return os.path.join(timesheet_dir, FALLBACK_TEMPLATE_FILENAME)


def _get_output_path(base_dir, year, month):
    return os.path.join(
        base_dir,
        '原始数据',
        '工时数据',
        OUTPUT_FILENAME_TEMPLATE.format(year=year, month=month),
    )


def _parse_month_label(month_value):
    text = _normalize_text(month_value)
    digits = ''.join(ch for ch in text if ch.isdigit())
    return int(digits) if digits else None


def _build_report_date(year, month):
    return datetime(year, month, calendar.monthrange(year, month)[1])


def _build_period_value(year, month):
    return int(f'{year % 100:02d}{month:02d}')


def _iter_month_rows(timesheet_path, target_year, target_month):
    wb = load_workbook(timesheet_path, data_only=True, read_only=True)
    ws = wb[SOURCE_SHEET_NAME]

    for row in ws.iter_rows(min_row=2, values_only=True):
        year_value = row[1]
        month_value = _parse_month_label(row[2])
        hours = row[10]
        if year_value != target_year or month_value != target_month:
            continue
        if hours is None or float(hours) <= 0:
            continue
        yield (
            _build_report_date(target_year, target_month),
            row[3],
            row[4],
            row[5],
            hours,
            row[6],
            row[8],
            _build_period_value(target_year, target_month),
        )


def list_available_months(base_dir, year=2026):
    timesheet_path = _get_timesheet_path(base_dir)
    wb = load_workbook(timesheet_path, data_only=True, read_only=True)
    ws = wb[SOURCE_SHEET_NAME]
    months = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        year_value = row[1]
        month_value = _parse_month_label(row[2])
        hours = row[10]
        if year_value != year or month_value is None:
            continue
        if hours is None or float(hours) <= 0:
            continue
        months.add(month_value)

    return tuple(sorted(months))


def _copy_row_style(ws, source_row_idx, target_row_idx):
    for col_idx in range(1, DATA_COLUMNS + 1):
        _copy_cell_style(ws.cell(row=source_row_idx, column=col_idx), ws.cell(row=target_row_idx, column=col_idx))
    ws.row_dimensions[target_row_idx].height = ws.row_dimensions[source_row_idx].height
    ws.row_dimensions[target_row_idx].hidden = ws.row_dimensions[source_row_idx].hidden


def _prepare_output_sheet(ws):
    original_last_row = ws.max_row
    for sheet_row_idx in range(DATA_START_ROW, original_last_row + 1):
        for col_idx in range(1, DATA_COLUMNS + 1):
            ws.cell(row=sheet_row_idx, column=col_idx).value = None
    return original_last_row


def export_timesheet_weekly_report(base_dir, year, month):
    timesheet_path = _get_timesheet_path(base_dir)
    template_path = _get_template_path(base_dir)
    output_path = _get_output_path(base_dir, year, month)

    rows = list(_iter_month_rows(timesheet_path, year, month))
    if not rows:
        raise ValueError(f'工时数据中未找到 {year} 年 {month:02d} 月的非零工时记录')

    wb = load_workbook(template_path)
    first_sheet = wb[wb.sheetnames[0]]
    for sheet_name in wb.sheetnames[1:]:
        del wb[sheet_name]
    ws = first_sheet
    ws.title = OUTPUT_SHEET_NAME

    original_last_row = _prepare_output_sheet(ws)
    style_row_idx = DATA_START_ROW

    for offset, row_values in enumerate(rows, start=0):
        target_row_idx = DATA_START_ROW + offset
        if target_row_idx > original_last_row:
            _copy_row_style(ws, style_row_idx, target_row_idx)
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=target_row_idx, column=col_idx).value = value

    final_last_row = DATA_START_ROW + len(rows) - 1
    if final_last_row < original_last_row:
        ws.delete_rows(final_last_row + 1, original_last_row - final_last_row)

    wb.active = 0
    wb.save(output_path)
    return output_path


def generate_default_timesheet_weekly_reports(base_dir, year=2026, months=DEFAULT_MONTHS):
    output_paths = []
    for month in months:
        output_paths.append(export_timesheet_weekly_report(base_dir, year, month))
    return output_paths
